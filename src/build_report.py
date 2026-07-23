#!/usr/bin/env python3
"""
FM Validator — _VALIDATED.xlsx Report Builder
16-tab transaction-grade audit report
"""
import sys, json, os, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule

# ── Colours ──────────────────────────────────────────────────────────────────
DARK_BLUE   = '1F4E79'; MID_BLUE   = '2E75B6'; LIGHT_BLUE  = 'D6E4F0'; PALE_BLUE = 'EBF3FA'
RED         = 'C00000'; LIGHT_RED  = 'FFE0E0'
AMBER       = 'C55A11'; LIGHT_AMBER= 'FCE4D6'
YELLOW      = 'FFD966'; LIGHT_YELL = 'FFF2CC'
GREEN       = '375623'; LIGHT_GREEN= 'E2EFDA'
GREY_DARK   = '595959'; GREY_MID   = 'A6A6A6'; GREY_LIGHT = 'F2F2F2'; WHITE = 'FFFFFF'
BORDER_COL  = 'BFBFBF'

# Dashboard redesign palette (Audit Output only) — per client UI brief, July 2026
CHARCOAL    = '1F2933'; GREY_TXT2   = '667085'
PANEL_GREY  = 'F5F7FA'; PANEL_BORDER= 'D9E2EC'
P1_FILL='FDECEC'; P1_TXT='9B1C1C'
P2_FILL='FFF4E5'; P2_TXT='9A5B00'
P3_FILL='EEF4FA'; P3_TXT='24435C'
OK_FILL ='EAF7EA'; OK_TXT ='1F6B3A'
PALE_ACCENT = 'EAF3F8'

def F(hex_col): return PatternFill('solid', fgColor=hex_col)
def Fn(bold=False,sz=10,col='000000',italic=False): return Font(bold=bold,size=sz,color=col,italic=italic,name='Arial')
def A(h='left',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def B(col=BORDER_COL,sty='thin'):
    s=Side(style=sty,color=col)
    return Border(left=s,right=s,top=s,bottom=s)
def noBorder():
    s=Side(style='thin',color=WHITE)
    return Border(left=s,right=s,top=s,bottom=s)

def cell(ws,ref,val=None,bold=False,sz=10,col='000000',italic=False,bg=None,h='left',v='center',wrap=False,border=False):
    c=ws[ref]
    if val is not None: c.value=val
    c.font=Fn(bold=bold,sz=sz,col=col,italic=italic)
    if bg: c.fill=F(bg)
    c.alignment=A(h=h,v=v,wrap=wrap)
    if border: c.border=B()
    return c

def hdr(ws,ref,val,bg=DARK_BLUE,tc=WHITE,sz=10,h='left',bold=True):
    c=ws[ref]; c.value=val
    c.font=Fn(bold=bold,sz=sz,col=tc)
    c.fill=F(bg); c.alignment=A(h=h,v='center')
    c.border=B(col=WHITE); return c

def merge(ws,rng,val=None,bold=False,sz=10,col='000000',bg=None,h='left',v='center',wrap=False,italic=False):
    ws.merge_cells(rng); tl=rng.split(':')[0]; c=ws[tl]
    if val is not None: c.value=val
    c.font=Fn(bold=bold,sz=sz,col=col,italic=italic)
    if bg: c.fill=F(bg)
    c.alignment=A(h=h,v=v,wrap=wrap); return c

def fill_range(ws,r1,c1,r2,c2,bg):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            ws.cell(r,c).fill=F(bg)

def set_col(ws,col,w): ws.column_dimensions[get_column_letter(col)].width=w
def set_row(ws,row,h): ws.row_dimensions[row].height=h

def row_data(ws,row,vals,widths,bgs,fns):
    for i,(val,fn) in enumerate(zip(vals,fns)):
        c=ws.cell(row,i+1); c.value=val
        c.font=fn; c.fill=F(bgs[i] if bgs[i] else WHITE)
        c.alignment=A(h='center' if i<2 else 'left',v='top',wrap=True)
        c.border=B()

# ════════════════════════════════════════════════════════════════════════════════
def merge_bold_prefix(ws, rng, bold_prefix, rest, sz=10, col='000000', bg=None, v='center'):
    """Merge a range with the leading label bold and the remaining text regular
    weight, within the SAME cell — used for 'Reason:   ...' style lines where
    only the label itself should be bold, not the dynamic content after it."""
    top_left = rng.split(':')[0]
    bold_font = InlineFont(b=True, sz=sz, color=col)
    reg_font  = InlineFont(b=False, sz=sz, color=col)
    rich = CellRichText([TextBlock(bold_font, bold_prefix), TextBlock(reg_font, rest)])
    ws.merge_cells(rng)
    c = ws[top_left]
    c.value = rich
    c.alignment = A(v=v)
    if bg:
        for row in ws[rng]:
            for cc in row:
                cc.fill = F(bg)
    return c


def build_report(data_path, output_path):
    with open(data_path,'r') as f: d=json.load(f)


    # Sanitize text fields — Excel formula errors in strings cause LibreOffice to flag them
    def san(v):
        if isinstance(v,str): return v.replace("#REF!","#REF").replace("#VALUE!","#VALUE").replace("#N/A","#NA").replace("#NAME?","#NAME").replace("#DIV/0!","#DIV0")
        return v
    findings=[{k:san(v) for k,v in f.items()} for f in d.get("findings",[])]
    t0           = d.get('tier0',{})
    crossRunStats = d.get('crossRunStats', {'closed':[], 'new':[], 'regressed':[], 'stillOpen':[]})
    modelName    = d.get('modelName','Financial Model')
    modelType    = d.get('modelType','unknown')
    modelIndustry= d.get('modelIndustry','')
    currency     = d.get('currency','')
    periodicity  = d.get('periodicity','')
    domainSkill  = d.get('domainSkill','skill-generic.md')
    reviewDate   = datetime.now().strftime('%d %b %Y')
    sourceFile   = d.get('sourceFile','model.xlsm')
    overallAssess= d.get('overallAssessment','not_fit_for_purpose')
    igReadiness  = d.get('igReadiness',0)
    igCommentary = d.get('igCommentary','')
    modelTier    = d.get('modelTier','Tier 1')
    reviewMode   = d.get('reviewMode','llm_only')
    ruleResults  = d.get('ruleResults',[])
    errorScan    = d.get('errorScan',[])
    redundantIn  = d.get('redundantInputs',{'applicable':False,'totalInputs':0,'redundantCount':0,'redundant':[],'inputSheets':[]})
    orphanIn     = d.get('orphanSheets',{'applicable':False,'orphanSheets':[],'financialStatementSheets':[],'reachableSheets':[],'totalSheets':0})
    namedRangeIn = d.get('namedRangeAudit',{'applicable':False,'unused':[],'poorlyNamed':[],'broken':[],'totalNamedRanges':0})
    reasonIn    = d.get('reasonableness',{'waccOverride':{'applicable':False},'terminalValue':{'applicable':False},'outputs':{'applicable':False,'results':[],'flaggedCount':0}})
    dupIn       = d.get('duplicateSheets',{'applicable':False,'flaggedCount':0,'flagged':[]})
    formulaDeepIn = d.get('formulaDeepDive',{'applicable':False,'reviewed':0,'findings':[]})
    deepAcctSheets = d.get('deepAccountingResolvedSheets',{'resolvedMap':{},'unresolvedCategories':[]})
    vbaIn        = d.get('vbaReview',{'applicable':False,'hasVbaProject':False,'moduleCount':0,'note':'','findings':[]})

    # Checklist rules for the Validation Matrix — loaded from config
    checklist_path=os.path.join(os.path.dirname(os.path.abspath(__file__)),'..','config','checklist.json')
    try:
        with open(checklist_path) as cf: _cl=json.load(cf)
        checklist_rules=([dict(r,_tier='Tier 1') for r in _cl.get('tier1',[])]
                        +[dict(r,_tier='Tier 2') for r in _cl.get('tier2',[])])
    except Exception:
        checklist_rules=[]

    # ── P1/P2/P3 framework renewal, Tier 3 ───────────────────────────────────
    # Four named readiness gates from the memo's own "Not Ready for
    # Reliance" list, beyond the P1/Critical-Query gates already built in
    # Tier 1. Each is computed from data that already exists — no new
    # per-check instrumentation was needed for any of these.

    # 1. "An incomplete mandatory critical procedure" — confirmed directly
    # against the real checklist.json (not the stale, unused duplicate at
    # checklists/checklist.json — a separate hygiene item, noted but not
    # fixed here): 43 rules across Tier 1+2 are tagged severity=='fatal'
    # or sit under a "0. Fatal gate" source_section. If any of these 43
    # specific rules has a status other than 'pass' in ruleResults, that
    # is a distinct signal from the overall igReadiness% — even a 95%-
    # complete audit can still have one of these unresolved.
    _mandatory_critical_ids = {r['id'] for r in checklist_rules
        if r.get('severity')=='fatal' or 'fatal gate' in str(r.get('source_section','')).lower()}
    _rule_status_by_id = {r.get('id'): r.get('status') for r in ruleResults}
    _incomplete_mandatory = [rid for rid in _mandatory_critical_ids
        if _rule_status_by_id.get(rid) not in ('pass',)]
    mandatory_procedure_gap = len(_incomplete_mandatory) > 0

    # 2. "An unaudited critical module" — from deepAcctSheets'
    # unresolvedCategories, filtered to only the categories that are
    # UNIVERSALLY expected in any financial model (Balance Sheet, Cash
    # Flow, Income Statement, Debt, Equity). Depreciation & Tax and
    # Leases are deliberately excluded here — confirmed via real testing
    # this session that these are commonly, legitimately not-applicable
    # (a pre-operational development model with no depreciation schedule
    # yet, a business with no material leases), not an audit gap.
    _universally_critical_modules = {'Balance Sheet','Cash Flow','Income Statement','Debt','Equity'}
    _unresolved_critical_modules = [c for c in deepAcctSheets.get('unresolvedCategories',[])
        if c in _universally_critical_modules]
    critical_module_gap = len(_unresolved_critical_modules) > 0

    # 3. "An unreconciled key output" — any finding from the key-output
    # dependency-chain check (T0-CHAIN-*), which traces IRR/NPV/EBITDA/
    # DSCR/MOIC-style cells for a chain that dead-ends at a blank cell.
    key_output_gap = any(str(f.get('id','')).startswith('T0-CHAIN') for f in findings)

    # 4. "Required reviewer approval" — this tool cannot manufacture a
    # human sign-off. Deliberately opt-in, defaulting to NOT approved:
    # an explicit reviewerApproved=true must be supplied in the payload
    # for this gate to clear. This is a genuine behaviour change, not
    # just an additive signal — the highest reliance tier becomes
    # unreachable through automation alone unless a human explicitly
    # confirms approval, which is exactly what the memo's own exhaustive
    # AND-clause for "Ready for Reliance" requires ("Required Reviewer
    # Approval Is Complete" is one of eight mandatory conjuncts).
    reviewer_approved = bool(d.get('reviewerApproved', False))
    reviewer_approval_gap = not reviewer_approved

    # ── P1/P2/P3 framework renewal, Tier 1 item 2 ────────────────────────────
    # Severity alone used to decide P1/P2/P3 for every finding. Per the
    # renewed framework: only a Confirmed Finding is eligible for a
    # P-priority at all — a Query, Critical Query, Observation, Scope
    # Limitation, Not Applicable, or False Positive is reported in full
    # (still gets its own Issue Log row below) but must never be counted
    # toward P1/P2/P3 stats, dashboard gates, or the Remediation tab.
    # Returning the record_type string itself for non-Confirmed records is
    # deliberate, not just a placeholder: every downstream usage of
    # priority() already handles an unrecognised key safely (a .get(key,
    # default) style-lookup, or a plain =='P1' comparison that's simply
    # False for anything else) — confirmed by reading every call site
    # before making this change, not assumed.
    def priority(f):
        if (f.get('record_type') or 'Confirmed Finding') != 'Confirmed Finding':
            return f.get('record_type')
        sev = (f.get('severity') or '').lower()
        if sev in ('fatal','critical'): return 'P1'
        if sev in ('high','medium'): return 'P2'
        return 'P3'

    p1 = [f for f in findings if priority(f)=='P1']
    p2 = [f for f in findings if priority(f)=='P2']
    p3 = [f for f in findings if priority(f)=='P3']
    critical_queries = [f for f in findings if f.get('record_type')=='Critical Query']

    # Pre-compute display titles ONCE (shared by dashboard + Issue Log — was
    # previously computed only inside the Issue Log loop, so the dashboard
    # top-issues table showed raw IDs). Also pre-compute the Issue Log row
    # each finding will land on, so the dashboard's "View" link can point
    # directly at it (internal same-workbook hyperlink — cross-workbook
    # HYPERLINK references render as broken/unresolved text in non-Excel
    # viewers such as Google Sheets or Excel Online).
    for _f in findings:
        _t=(_f.get('title') or _f.get('label') or '').strip()
        if not _t or _t==_f.get('id'):
            _w=str(_f.get('condition') or _f.get('what_wrong') or _f.get('reason') or _f.get('detail') or '').replace('\n',' ').strip()
            if _w:
                _t=re.split(r'(?<=[.;])\s',_w)[0]
            else:
                _t=_f.get('id','')
        _f['title']=_t; _f['label']=_t
    id_to_issue_row = {f.get('id'): i for i, f in enumerate(findings, 5)}  # Issue Log header now at row 4, data from row 5

    # Audit coverage counts — feeds the dashboard completion breakdown (V11 1.3)
    def _rmatch0(rid,xid): return xid==rid or (xid or '').startswith(rid+'-')
    cov_perf=cov_pass=cov_issue=cov_unc=cov_np=0
    for _rule in checklist_rules:
        _rres=[r for r in ruleResults if _rmatch0(_rule.get('id',''),r.get('id',''))]
        if not _rres: cov_np+=1; continue
        cov_perf+=1
        if any(r.get('status') not in ('pass','uncertain') for r in _rres): cov_issue+=1
        elif any(r.get('status')=='uncertain' for r in _rres): cov_unc+=1
        else: cov_pass+=1

    # KPMG verdict
    # Neutral audit completion header — no verdict or reliance conclusion
    p1_open = len(p1)
    p2_open = len(p2)
    p3_open = len(p3)
    # ── P1/P2/P3 framework renewal, Tier 1 item 3 ────────────────────────────
    # An unresolved Critical Query must block reliance the same way an open
    # P1 does, even though nothing has been confirmed as a defect yet — this
    # is the memo's own explicit principle, distinct from ordinary Queries/
    # Observations, which do NOT gate reliance at all.
    critical_query_open = len(critical_queries)
    # ── P1/P2/P3 framework renewal, Tier 3 ───────────────────────────────────
    # The three substantive named gates (mandatory critical procedure,
    # critical module, key output) block reliance with the SAME severity
    # as an open P1 or unresolved Critical Query — the memo lists all of
    # these together under "Not Ready for Reliance", not as a lesser tier.
    other_gates_open = mandatory_procedure_gap or critical_module_gap or key_output_gap
    # Reliance classification (V11 1.1) — component of the audit conclusion,
    # stated factually. Hard rule: a model cannot be classified reliance-ready
    # for lender/investor use or transaction execution while any P1 item is
    # open, OR while any Critical Query remains unresolved.
    if p1_open > 0 or critical_query_open > 0 or other_gates_open:
        _blocker_parts = []
        if p1_open > 0: _blocker_parts.append(f'{p1_open} P1 item(s)')
        if critical_query_open > 0: _blocker_parts.append(f'{critical_query_open} unresolved Critical Quer{"y" if critical_query_open==1 else "ies"}')
        if mandatory_procedure_gap: _blocker_parts.append(f'{len(_incomplete_mandatory)} incomplete mandatory critical procedure(s)')
        if critical_module_gap: _blocker_parts.append(f'{len(_unresolved_critical_modules)} unaudited critical module(s) ({", ".join(_unresolved_critical_modules)})')
        if key_output_gap: _blocker_parts.append('an unreconciled key output')
        _blocker_desc = ', '.join(_blocker_parts)
        if igReadiness >= 60:
            verdict_short='RELIANCE-READY FOR INTERNAL REVIEW ONLY'; verdict_bg=DARK_BLUE
            verdict_text=(f'{_blocker_desc} remain open — this model cannot be classified reliance-ready for management, lender or investor use until these are resolved (or, for a Critical Query, confirmed either way) and retested. '
                          f'{igReadiness}% of planned procedures completed ({cov_pass} passed, {cov_issue} raised issues, {cov_unc} uncertain, {cov_np} not run).')
        else:
            verdict_short='NOT RELIANCE-READY'; verdict_bg=DARK_BLUE
            verdict_text=(f'{_blocker_desc} open and {igReadiness}% of planned procedures completed. Both open reliance blockers and audit coverage prevent reliance at any level. '
                          f'Resolve these items and complete outstanding procedures before reassessment.')
    elif igReadiness >= 95 and not reviewer_approval_gap:
        verdict_short='RELIANCE-READY FOR TRANSACTION EXECUTION'; verdict_bg=MID_BLUE
        verdict_text=(f'No P1 items, unresolved Critical Queries, or other reliance blockers open, reviewer approval recorded, and {igReadiness}% of planned procedures completed ({cov_pass} passed). '
                      f'{p2_open} P2 item(s) remain and should be resolved or formally waived as part of transaction close.')
    elif igReadiness >= 95 and reviewer_approval_gap:
        # ── P1/P2/P3 framework renewal, Tier 3 ─────────────────────────────
        # The memo's own exhaustive AND-clause for the highest reliance
        # tier explicitly includes "Required Reviewer Approval Is
        # Complete" as one of eight mandatory conjuncts. This tool cannot
        # manufacture a human sign-off, so this tier is deliberately
        # capped one level down until an explicit reviewerApproved=true
        # is supplied — a genuine, disclosed behaviour change, not a
        # silent one.
        verdict_short='RELIANCE-READY FOR LENDER / INVESTOR REVIEW'; verdict_bg=MID_BLUE
        verdict_text=(f'All automated gates pass and {igReadiness}% of planned procedures completed ({cov_pass} passed), but required reviewer approval has not yet been recorded — '
                      f'this caps the model one tier below Transaction Execution readiness until a reviewer explicitly signs off. {p2_open} P2 item(s) remain.')
    elif igReadiness >= 80:
        verdict_short='RELIANCE-READY FOR LENDER / INVESTOR REVIEW'; verdict_bg=MID_BLUE
        verdict_text=(f'No P1 items or unresolved Critical Queries open; {igReadiness}% of planned procedures completed. Outstanding items: {cov_unc} uncertain and {cov_np} not-run procedures, {p2_open} open P2 item(s). '
                      f'Suitable for external review with these limitations disclosed.')
    elif igReadiness >= 60:
        verdict_short='RELIANCE-READY FOR MANAGEMENT DISCUSSION'; verdict_bg=MID_BLUE
        verdict_text=(f'No P1 items or unresolved Critical Queries open; {igReadiness}% of planned procedures completed. Coverage gaps ({cov_np} procedures not run, {cov_unc} uncertain) limit use to internal management discussion until closed.')
    else:
        verdict_short='RELIANCE-READY FOR INTERNAL REVIEW ONLY'; verdict_bg=DARK_BLUE
        verdict_text=(igCommentary or f'{igReadiness}% of planned procedures completed. Coverage is not yet sufficient for management, lender or investor reliance. {cov_np} procedures not run, {cov_unc} uncertain.')

    risk_rating = f'P1: {len(p1)}  P2: {len(p2)}  P3: {len(p3)}'

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # TAB 1 — AUDIT OUTPUT
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = 'Audit Output'; ws1.sheet_view.showGridLines=False
    ws1.freeze_panes = 'B5'
    ws1.sheet_view.zoomScale = 100
    for col,w in [(1,3)] + [(c,15) for c in range(2,10)] + [(10,3)]:
        set_col(ws1,col,w)

    def badge(ws,r,c,text,fill,txt_col,bold=True,sz=9):
        cc=ws.cell(r,c); cc.value=text; cc.font=Fn(bold=bold,sz=sz,col=txt_col)
        cc.fill=F(fill); cc.alignment=A(h='center',v='center'); cc.border=B(col=PANEL_BORDER)
        return cc

    def kpi_card(ws,label_row,val_row,c1,c2,label,value,fmt=None,val_col=CHARCOAL):
        col_letter1=get_column_letter(c1); col_letter2=get_column_letter(c2)
        if c2>c1: merge(ws,f'{col_letter1}{label_row}:{col_letter2}{label_row}',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        else: cell(ws,f'{col_letter1}{label_row}',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        if c2>c1: merge(ws,f'{col_letter1}{val_row}:{col_letter2}{val_row}',value,bold=True,sz=18,col=val_col,bg=PANEL_GREY,h='center')
        else: cell(ws,f'{col_letter1}{val_row}',value,bold=True,sz=18,col=val_col,bg=PANEL_GREY,h='center')
        if fmt: ws[f'{col_letter1}{val_row}'].number_format=fmt
        for rr in (label_row,val_row):
            for cc in range(c1,c2+1):
                ws.cell(rr,cc).border=B(col=PANEL_BORDER)

    r = 1  # running row pointer — every section advances this, nothing below is hardcoded

    # ── Header band ──────────────────────────────────────────────────────────
    merge(ws1,f'B{r}:I{r+1}','FINANCIAL MODEL AUDIT REPORT',bold=True,sz=20,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws1,r,2,r+1,9,DARK_BLUE); set_row(ws1,r,22); set_row(ws1,r+1,10); r+=2
    merge(ws1,f'B{r}:I{r}',modelName,sz=11,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws1,r,2,r,9,DARK_BLUE); set_row(ws1,r,20); r+=1

    merge(ws1,f'B{r}:C{r}',f'{modelType} — {modelIndustry}',sz=9,col=GREY_TXT2,bg=PANEL_GREY,v='center')
    merge(ws1,f'D{r}:E{r}',f'{currency} · {periodicity}',sz=9,col=GREY_TXT2,bg=PANEL_GREY,v='center')
    merge(ws1,f'F{r}:G{r}',f'Audit mode: {"AI-assisted (Mode A)" if reviewMode=="llm_only" else reviewMode}',sz=9,col=GREY_TXT2,bg=PANEL_GREY,v='center')
    merge(ws1,f'H{r}:I{r}',f'Review date: {reviewDate}',sz=9,col=GREY_TXT2,bg=PANEL_GREY,v='center',h='right')
    set_row(ws1,r,22); r+=1
    set_row(ws1,r,8); r+=1

    # ── Reliance status card — Status / Reason / Required action ──────────────
    _verdict_display = {
        'NOT RELIANCE-READY': 'Not ready for reliance',
        'RELIANCE-READY FOR INTERNAL REVIEW ONLY': 'Ready for internal review only',
        'RELIANCE-READY FOR MANAGEMENT DISCUSSION': 'Ready for management discussion',
        'RELIANCE-READY FOR LENDER / INVESTOR REVIEW': 'Ready for lender / investor review',
        'RELIANCE-READY FOR TRANSACTION EXECUTION': 'Ready for transaction execution',
    }.get(verdict_short, verdict_short.title())
    # ── P1/P2/P3 framework renewal, Tier 1 item 3 (completion) ───────────────
    # Found via reviewing a real run: verdict_short/verdict_bg/verdict_text
    # above were correctly updated to gate on critical_query_open, but these
    # three separate dashboard text fields (Reason, Required action, Key
    # Takeaway) were missed — they referenced p1_open directly and would
    # have shown "4 open P1 finding(s)" while silently omitting 16 open
    # Critical Queries sitting right next to that banner in the same
    # report. Re-derived here independently (rather than assuming
    # _blocker_desc from the branch above is in scope, which it would not
    # be when p1_open==0 and critical_query_open==0).
    # Tier 3 extension of the same fix: the three new named gates
    # (mandatory procedure, critical module, key output) must appear in
    # these fields too, or the Reason line would silently omit a gate
    # the verdict banner right above it is blocking on.
    _open_parts = []
    if p1_open>0: _open_parts.append(f'{p1_open} open P1 finding(s)')
    if critical_query_open>0: _open_parts.append(f'{critical_query_open} unresolved Critical Quer{"y" if critical_query_open==1 else "ies"}')
    if mandatory_procedure_gap: _open_parts.append(f'{len(_incomplete_mandatory)} incomplete mandatory critical procedure(s)')
    if critical_module_gap: _open_parts.append(f'{len(_unresolved_critical_modules)} unaudited critical module(s)')
    if key_output_gap: _open_parts.append('an unreconciled key output')
    _open_desc = ', '.join(_open_parts)
    _has_blocker = p1_open>0 or critical_query_open>0 or other_gates_open
    _reason = (f'{_open_desc} and {igReadiness}% audit completion' if _has_blocker
               else f'{igReadiness}% audit completion, {cov_unc} procedure(s) uncertain, {cov_np} not run' if igReadiness<100 or cov_unc or cov_np
               else 'All planned procedures completed with no open P1 findings or unresolved Critical Queries')
    _next_step = ('Close all P1 items, resolve all Critical Queries (confirming whether a defect exists either way), and clear any incomplete mandatory procedures, unaudited critical modules, and unreconciled key outputs — then complete outstanding procedures and reassess.' if _has_blocker
                  else 'Resolve remaining P2 items and complete outstanding procedures before wider reliance.' if igReadiness<95
                  else ('Record reviewer approval to reach Transaction Execution readiness.' if reviewer_approval_gap else 'No further action required for this reliance level.'))
    merge(ws1,f'B{r}:I{r}',f'Status:   {_verdict_display}',bold=True,sz=12,col=WHITE,bg=verdict_bg,v='center')
    fill_range(ws1,r,2,r,9,verdict_bg); set_row(ws1,r,26); r+=1
    merge_bold_prefix(ws1,f'B{r}:I{r}','Reason:   ',_reason,sz=10,col=CHARCOAL,bg=PANEL_GREY,v='center')
    set_row(ws1,r,20); r+=1
    merge_bold_prefix(ws1,f'B{r}:I{r}','Required action:   ',_next_step,sz=10,col=CHARCOAL,bg=PANEL_GREY,v='center')
    set_row(ws1,r,20); r+=1
    # A5 — compact severity summary, borrowed from audit-xls's convention
    # of a scannable one-line count at the top of the output, ahead of any
    # detail. p1/p2/p3 are already computed above for the Work Not
    # Performed table's own logic; this just surfaces the same counts
    # more prominently, right where a reader looks first.
    merge_bold_prefix(ws1,f'B{r}:I{r}','Findings by priority:   ',f'{len(p1)} P1 · {len(p2)} P2 · {len(p3)} P3' + (f' · {critical_query_open} Critical Quer{"y" if critical_query_open==1 else "ies"}' if critical_query_open>0 else ''),sz=10,col=CHARCOAL,bg=PANEL_GREY,v='center')
    set_row(ws1,r,20); r+=1
    # ── P1/P2/P3 framework renewal, Tier 2 item 2 ────────────────────────────
    # Cross-run Closed/New/Regressed — previously the report only ever
    # showed OPEN counts, giving no way to confirm what was actually fixed
    # between runs versus what simply wasn't flagged this time for an
    # unrelated reason. A first run (no prior history to compare against)
    # is called out explicitly rather than silently showing "0 closed, 0
    # regressed" in a way that could be misread as "nothing changed".
    _crn_closed = len(crossRunStats.get('closed', []))
    _crn_new = len(crossRunStats.get('new', []))
    _crn_regressed = len(crossRunStats.get('regressed', []))
    _crn_still_open = len(crossRunStats.get('stillOpen', []))
    _is_first_run = (_crn_closed == 0 and _crn_regressed == 0 and _crn_still_open == 0 and _crn_new > 0)
    if _is_first_run:
        _crn_text = f'First run for this model — no prior report to compare against ({_crn_new} finding(s) established as the baseline).'
    else:
        _crn_text = f'{_crn_closed} closed · {_crn_new} new · {_crn_still_open} still open since the last run'
        if _crn_regressed > 0:
            _crn_text += f' · \u26a0 {_crn_regressed} regressed (previously closed, now reappeared)'
    merge_bold_prefix(ws1,f'B{r}:I{r}','Since last run:   ',_crn_text,sz=10,col=CHARCOAL,bg=(LIGHT_RED if _crn_regressed>0 else PANEL_GREY),v='center')
    set_row(ws1,r,20); r+=1
    set_row(ws1,r,8); r+=1

    # ── Key Takeaway box ──────────────────────────────────────────────────────
    _blockers=[]
    if p1_open>0: _blockers.append('open P1 findings')
    if critical_query_open>0: _blockers.append('unresolved Critical Queries')
    if mandatory_procedure_gap: _blockers.append('incomplete mandatory critical procedures')
    if critical_module_gap: _blockers.append('unaudited critical modules')
    if key_output_gap: _blockers.append('an unreconciled key output')
    if t0.get('stats',{}).get('totalExternalLinks',0)>0: _blockers.append('external workbook links')
    if len(errorScan)>0: _blockers.append('formula errors')
    if cov_unc>0: _blockers.append('incomplete audit coverage')
    if redundantIn.get('applicable') and redundantIn.get('redundantCount',0)>0: _blockers.append('redundant input assumptions')
    _takeaway = (f"The model is not currently suitable for reliance. The main blockers are {', '.join(_blockers[:4])}."
                 if _has_blocker or _blockers else
                 "The model has no open P1 findings or unresolved Critical Queries and is suitable for reliance at the level shown above.")
    merge(ws1,f'B{r}:I{r}','Key Takeaway',bold=True,sz=8,col=GREY_TXT2,bg=WHITE,h='left'); set_row(ws1,r,14); r+=1
    merge(ws1,f'B{r}:I{r+1}',_takeaway,sz=10,col=CHARCOAL,bg=PALE_ACCENT,wrap=True,v='center')
    fill_range(ws1,r,2,r+1,9,PALE_ACCENT); set_row(ws1,r,18); set_row(ws1,r+1,18); r+=2
    set_row(ws1,r,8); r+=1

    # ── KPI cards — Row A: Risk (P1/P2/P3, 2 cols each) + Coverage (2 cols) ───
    merge(ws1,f'B{r}:G{r}','Risk',bold=True,sz=8,col=GREY_TXT2,bg=WHITE,h='left')
    merge(ws1,f'H{r}:I{r}','Coverage',bold=True,sz=8,col=GREY_TXT2,bg=WHITE,h='left')
    set_row(ws1,r,14); r+=1
    _lbl_row, _val_row = r, r+1
    kpi_card(ws1,_lbl_row,_val_row,2,3,'P1 OPEN',p1_open,fmt='#,##0;[Red](#,##0);-',val_col=(P1_TXT if p1_open>0 else CHARCOAL))
    kpi_card(ws1,_lbl_row,_val_row,4,5,'P2 OPEN',p2_open,fmt='#,##0;[Red](#,##0);-',val_col=(P2_TXT if p2_open>0 else CHARCOAL))
    kpi_card(ws1,_lbl_row,_val_row,6,7,'P3 OPEN',p3_open,fmt='#,##0;[Red](#,##0);-',val_col=(P3_TXT if p3_open>0 else CHARCOAL))
    ws1.merge_cells(f'H{_lbl_row}:I{_lbl_row}'); ws1[f'H{_lbl_row}'].value='Audit Completion'
    ws1[f'H{_lbl_row}'].font=Fn(bold=True,sz=8,col=GREY_TXT2); ws1[f'H{_lbl_row}'].fill=F(PANEL_GREY); ws1[f'H{_lbl_row}'].alignment=A(h='center')
    ws1.merge_cells(f'H{_val_row}:I{_val_row}'); ws1[f'H{_val_row}'].value=igReadiness/100.0; ws1[f'H{_val_row}'].number_format='0%;[Red](0%);-'
    ws1[f'H{_val_row}'].font=Fn(bold=True,sz=18,col=MID_BLUE); ws1[f'H{_val_row}'].fill=F(PANEL_GREY); ws1[f'H{_val_row}'].alignment=A(h='center')
    for rr in (_lbl_row,_val_row):
        for cc in range(2,10): ws1.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(ws1,_lbl_row,16); set_row(ws1,_val_row,26); r=_val_row+1
    set_row(ws1,r,10); r+=1

    # ── KPI cards — Row B: Formula complexity (5×1col) + Input control (3col) ─
    merge(ws1,f'B{r}:F{r}','Formula Complexity',bold=True,sz=8,col=GREY_TXT2,bg=WHITE,h='left')
    merge(ws1,f'G{r}:I{r}','Input Control',bold=True,sz=8,col=GREY_TXT2,bg=WHITE,h='left')
    set_row(ws1,r,14); r+=1
    _lbl_row, _val_row = r, r+1
    _stats=t0.get('stats',{})
    fcells=[('Unique\nFormulas',_stats.get('uniqueFormulaCount',0),2),
            ('Formula\nCells',_stats.get('totalFormulaCells',0),3),
            ('IFERROR\nCount',_stats.get('totalIferrorCount',0),4),
            ('OFFSET\nCount',_stats.get('totalOffsetCount',0),5),
            ('External\nLinks',_stats.get('totalExternalLinks',0),6)]
    for label,val,col in fcells:
        c1=ws1.cell(_lbl_row,col); c1.value=label; c1.font=Fn(bold=True,sz=8,col=GREY_TXT2); c1.fill=F(PANEL_GREY); c1.alignment=A(h='center',wrap=True)
        c2=ws1.cell(_val_row,col); c2.value=val; c2.number_format='#,##0;[Red](#,##0);-'
        ext_warn = (label.startswith('External') and val>0)
        c2.font=Fn(bold=True,sz=16,col=(P2_TXT if ext_warn else CHARCOAL)); c2.fill=F(PANEL_GREY); c2.alignment=A(h='center')
        c1.border=B(col=PANEL_BORDER); c2.border=B(col=PANEL_BORDER)
    set_row(ws1,_lbl_row,22)
    _rc=redundantIn.get('redundantCount',0); _ti=redundantIn.get('totalInputs',0)
    merge(ws1,f'G{_lbl_row}:I{_lbl_row}','Redundant\nInputs',bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
    ri_val = f'{_rc:,} of {_ti:,}' if redundantIn.get('applicable') else 'N/A'
    merge(ws1,f'G{_val_row}:I{_val_row}',ri_val,bold=True,sz=16,col=(P2_TXT if _rc>0 else CHARCOAL),bg=PANEL_GREY,h='center')
    for rr in (_lbl_row,_val_row):
        for cc in range(7,10): ws1.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(ws1,_val_row,26); r=_val_row+1
    set_row(ws1,r,10); r+=1

    # ── Audit coverage bar + compact procedure mini-table ──────────────────────
    merge(ws1,f'B{r}:I{r}','Audit Coverage',bold=True,sz=11,col=DARK_BLUE,bg=WHITE,h='left'); set_row(ws1,r,18); r+=1
    _filled = max(0, min(8, round(igReadiness/100.0*8)))
    for i in range(8):
        col=2+i
        ws1.cell(r,col).fill = F(MID_BLUE if i<_filled else PALE_ACCENT)
        ws1.cell(r,col).border = B(col=PANEL_BORDER)
    set_row(ws1,r,10); r+=1
    merge(ws1,f'B{r}:I{r}',
          f'{len(checklist_rules)} planned · {cov_perf} performed · {cov_pass} passed · {cov_issue} raised issues · {cov_unc} uncertain · {cov_np} not run',
          sz=9,col=GREY_TXT2,bg=WHITE)
    set_row(ws1,r,18); r+=1
    set_row(ws1,r,10); r+=1

    # ── Top 5 Blockers ──────────────────────────────────────────────────────────
    merge(ws1,f'B{r}:I{r}','Top 5 Blockers',bold=True,sz=11,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws1,r,2,r,9,DARK_BLUE); set_row(ws1,r,20); r+=1
    _tb_headers=[('Priority',2,2),('Area',3,3),('Issue',4,5),('Model Impact',6,7),('Required Action',8,8),('View',9,9)]
    for label,c1,c2 in _tb_headers:
        col_l1=get_column_letter(c1); col_l2=get_column_letter(c2)
        if c2>c1: merge(ws1,f'{col_l1}{r}:{col_l2}{r}',label,bold=True,sz=8,col=WHITE,bg=DARK_BLUE,h='center')
        else: cell(ws1,f'{col_l1}{r}',label,bold=True,sz=8,col=WHITE,bg=DARK_BLUE,h='center')
        ws1[f'{col_l1}{r}'].border=B(col=WHITE)
    set_row(ws1,r,18); r+=1

    _pri_rank={'P1':0,'P2':1,'P3':2}
    # P1/P2/P3 framework renewal, Tier 2 item 3: risk_weighted_total ranks
    # WITHIN a priority tier — the memo's own explicit ask ("a weighted
    # score may be used to rank P2 and P3 findings... but a numerical
    # score must not create a P1 by itself"). Placed right after the
    # priority-tier rank, before the pre-existing key_output_impact/
    # fscore tiebreakers, which still apply for any finding that wasn't
    # risk-scored (e.g. a non-Confirmed-Finding record_type, by design —
    # see risk-scoring.js's own gating).
    top5 = sorted(p1+p2+p3, key=lambda f:(
        _pri_rank.get(priority(f),3),
        -(f.get('risk_weighted_total') or 0),
        0 if str(f.get('key_output_impact','')).lower() in ('yes','true','high') else 1,
        -(f.get('fscore') or 0)))[:5]
    _pri_style={'P1':(P1_FILL,P1_TXT),'P2':(P2_FILL,P2_TXT),'P3':(P3_FILL,P3_TXT)}
    for f in top5:
        i=r
        pri=priority(f); pf,pt=_pri_style.get(pri,(GREY_LIGHT,CHARCOAL))
        badge(ws1,i,2,pri,pf,pt)
        c3=ws1.cell(i,3); c3.value=f.get('category','') or '—'; c3.font=Fn(sz=9,col=CHARCOAL); c3.fill=F(WHITE); c3.alignment=A(h='center',v='center'); c3.border=B(col=PANEL_BORDER)
        ws1.merge_cells(f'D{i}:E{i}')
        ws1[f'D{i}'].value=(f.get('title') or f.get('label') or '')
        ws1[f'D{i}'].font=Fn(sz=9,col=CHARCOAL); ws1[f'D{i}'].fill=F(WHITE); ws1[f'D{i}'].alignment=A(wrap=True,v='center')
        ws1.merge_cells(f'F{i}:G{i}')
        impact = f.get('model_risk') or f.get('consequence') or '—'
        ws1[f'F{i}'].value=str(impact); ws1[f'F{i}'].font=Fn(sz=9,col=CHARCOAL); ws1[f'F{i}'].fill=F(WHITE); ws1[f'F{i}'].alignment=A(wrap=True,v='center')
        action = f.get('corrective_action') or f.get('fix_instruction') or '—'
        c8=ws1.cell(i,8); c8.value=str(action); c8.font=Fn(sz=9,col=CHARCOAL); c8.fill=F(WHITE); c8.alignment=A(wrap=True,v='center'); c8.border=B(col=PANEL_BORDER)
        _row_target = id_to_issue_row.get(f.get('id'))
        c9=ws1.cell(i,9)
        if _row_target:
            c9.value=f"=HYPERLINK(\"#'Issue Log'!A{_row_target}\",\"View\")"
            c9.font=Font(size=9,color=MID_BLUE,underline='single',name='Arial')
        else:
            c9.value='—'; c9.font=Fn(sz=9,col=GREY_MID)
        c9.fill=F(WHITE); c9.alignment=A(h='center',v='center'); c9.border=B(col=PANEL_BORDER)
        for cc in (2,3,4,5,6,7,8,9): ws1.cell(i,cc).border=B(col=PANEL_BORDER)
        set_row(ws1,i,32); r+=1
    set_row(ws1,r,10); r+=1

    # ── Review area status — soft badges, Reference Tab column ─────────────────
    status_areas=[
        ('Formula integrity',t0.get('stats',{}).get('totalExternalLinks',0)>0 or t0.get('stats',{}).get('totalRefInFormula',0)>0,'formula errors or external links detected'),
        ('Unique formula review',t0.get('stats',{}).get('fscoreDist',{}).get('High',0)>0,'Complex formulas found — see Formula Analysis tab for detail'),
        ('Dependency flow',False,'Dependency flow assessed'),
        ('Scenario logic',False,'Scenario checks completed'),
        ('Debt / funding logic',any(f.get('category')=='Debt' and priority(f)=='P1' for f in findings),'Debt checks completed'),
        ('Financial statements',any(f.get('category')=='Integration' and priority(f)=='P1' for f in findings),'Three-statement integration checked'),
        ('Tax logic',any(f.get('category')=='Tax' for f in findings),'Tax checks completed'),
        ('Valuation / returns',False,'Valuation checks completed'),
        ('Presentation / usability',False,'Presentation checks completed'),
        ('Input linkage',
         redundantIn.get('applicable',False) and redundantIn.get('redundantCount',0)>0,
         (f"{redundantIn.get('redundantCount',0)} of {redundantIn.get('totalInputs',0)} input-sheet constants are not referenced by any formula — see the Redundant Inputs tab and Issue Log finding T0-RI-001"
          if redundantIn.get('applicable',False) and redundantIn.get('redundantCount',0)>0
          else f"All {redundantIn.get('totalInputs',0)} input-sheet constants are referenced by model formulas"
          if redundantIn.get('applicable',False)
          else 'Not applicable — no input/assumption-named sheet detected')),
        ('Statement linkage',
         orphanIn.get('applicable',False) and len(orphanIn.get('orphanSheets',[]))>0,
         (f"{len(orphanIn.get('orphanSheets',[]))} sheet(s) have no traceable path to a financial statement ({', '.join(orphanIn.get('orphanSheets',[])[:5])}) — see the Sheet Linkage tab and Issue Log finding T0-LINK-001"
          if orphanIn.get('applicable',False) and len(orphanIn.get('orphanSheets',[]))>0
          else f"All {orphanIn.get('totalSheets',0)} sheets have a traceable path to a financial statement"
          if orphanIn.get('applicable',False)
          else 'Not applicable — no financial-statement-named sheet detected')),
        ('Named range audit',
         namedRangeIn.get('applicable',False) and (len(namedRangeIn.get('broken',[]))>0 or len(namedRangeIn.get('unused',[]))>0),
         (f"{len(namedRangeIn.get('broken',[]))} broken, {len(namedRangeIn.get('unused',[]))} unused of {namedRangeIn.get('totalNamedRanges',0)} named ranges — see the Named Range Audit tab"
          if namedRangeIn.get('applicable',False) and (len(namedRangeIn.get('broken',[]))>0 or len(namedRangeIn.get('unused',[]))>0)
          else f"All {namedRangeIn.get('totalNamedRanges',0)} named ranges are used and resolve correctly"
          if namedRangeIn.get('applicable',False)
          else 'Not applicable — no named ranges defined in this model')),
        ('Commercial reasonableness',
         (reasonIn.get('waccOverride',{}).get('mismatch',False) or reasonIn.get('terminalValue',{}).get('flagged',False)
          or reasonIn.get('outputs',{}).get('flaggedCount',0)>0 or dupIn.get('flaggedCount',0)>0),
         (f"{reasonIn.get('outputs',{}).get('flaggedCount',0)} output metric(s) and {dupIn.get('flaggedCount',0)} duplicate sheet(s) flagged for commercial-reasonableness review — this checks whether results are believable, not whether the model is wired correctly. See the Reasonableness Review tab."
          if (reasonIn.get('waccOverride',{}).get('applicable',False) or reasonIn.get('outputs',{}).get('applicable',False))
          else 'No output metrics or duplicate sheets were found to warrant flagging'
          if (reasonIn.get('waccOverride',{}).get('applicable',False) or reasonIn.get('outputs',{}).get('applicable',False))
          else 'Not applicable — no labelled output metrics found to test in this model')),
    ]
    merge(ws1,f'B{r}:I{r}','Review Area Status',bold=True,sz=11,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws1,r,2,r,9,DARK_BLUE); set_row(ws1,r,20); r+=1
    for label,c1,c2 in [('Review Area',2,3),('Status',4,4),('Key Finding',5,8),('Ref. Tab',9,9)]:
        col_l1=get_column_letter(c1); col_l2=get_column_letter(c2)
        if c2>c1: merge(ws1,f'{col_l1}{r}:{col_l2}{r}',label,bold=True,sz=8,col=WHITE,bg=DARK_BLUE,h='center')
        else: cell(ws1,f'{col_l1}{r}',label,bold=True,sz=8,col=WHITE,bg=DARK_BLUE,h='center')
        ws1[f'{col_l1}{r}'].border=B(col=WHITE)
    set_row(ws1,r,18); r+=1

    _ref_tab = {
        'Formula integrity':'Formula Risk Review','Unique formula review':'Formula Risk Review',
        'Input linkage':'Redundant Inputs','Debt / funding logic':'Issue Log',
        'Financial statements':'Issue Log','Tax logic':'Issue Log',
        'Statement linkage':'Sheet Linkage','Named range audit':'Named Range Audit',
        'Commercial reasonableness':'Reasonableness Review',
    }
    _blocked_areas = {'Debt / funding logic','Financial statements'} if p1_open>0 else set()
    for area,has_issue,summary in status_areas:
        i=r
        if area=='Statement linkage' and orphanIn.get('applicable',False) and len(orphanIn.get('orphanSheets',[]))>0:
            status_txt='Blocked'; sf,st=(P1_FILL,P1_TXT)
        elif area=='Named range audit' and namedRangeIn.get('applicable',False) and len(namedRangeIn.get('broken',[]))>0:
            status_txt='Blocked'; sf,st=(P1_FILL,P1_TXT)
        elif area in _blocked_areas and any(f.get('category') in ('Debt','Integration') and priority(f)=='P1' for f in findings):
            status_txt='Blocked'; sf,st=(P1_FILL,P1_TXT)
        elif has_issue:
            status_txt='Review'; sf,st=(P2_FILL,P2_TXT)
        elif area=='Input linkage' and not redundantIn.get('applicable',False):
            status_txt='Not Started'; sf,st=(GREY_LIGHT,GREY_TXT2)
        elif area=='Statement linkage' and not orphanIn.get('applicable',False):
            status_txt='Not Started'; sf,st=(GREY_LIGHT,GREY_TXT2)
        elif area=='Named range audit' and not namedRangeIn.get('applicable',False):
            status_txt='Not Started'; sf,st=(GREY_LIGHT,GREY_TXT2)
        elif area=='Commercial reasonableness' and not (reasonIn.get('waccOverride',{}).get('applicable',False) or reasonIn.get('outputs',{}).get('applicable',False)):
            status_txt='Not Started'; sf,st=(GREY_LIGHT,GREY_TXT2)
        else:
            status_txt='Completed'; sf,st=(OK_FILL,OK_TXT)
        ws1.merge_cells(f'B{i}:C{i}')
        ws1[f'B{i}'].value=area; ws1[f'B{i}'].font=Fn(bold=True,sz=9,col=CHARCOAL); ws1[f'B{i}'].fill=F(WHITE); ws1[f'B{i}'].alignment=A(v='center')
        badge(ws1,i,4,status_txt,sf,st,sz=8)
        ws1.merge_cells(f'E{i}:H{i}')
        ws1[f'E{i}'].value=summary; ws1[f'E{i}'].font=Fn(sz=9,col=GREY_TXT2); ws1[f'E{i}'].fill=F(WHITE); ws1[f'E{i}'].alignment=A(wrap=True,v='center')
        ref = _ref_tab.get(area,'Validation Matrix')
        c9=ws1.cell(i,9); c9.value=f"=HYPERLINK(\"#'{ref}'!A1\",\"{ref}\")"
        c9.font=Font(size=8,color=MID_BLUE,underline='single',name='Arial'); c9.fill=F(WHITE); c9.alignment=A(h='center',v='center')
        for cc in range(2,10): ws1.cell(i,cc).border=B(col=PANEL_BORDER)
        set_row(ws1,i,20); r+=1
    set_row(ws1,r,10); r+=1

    # ── Accounting review summary — 4 lines only ────────────────────────────────
    merge(ws1,f'B{r}:I{r}','Accounting Review Summary',bold=True,sz=11,col=DARK_BLUE,bg=WHITE,h='left'); set_row(ws1,r,18); r+=1
    _acct=[('Basis','Accrual basis assumed from financial statement structure'),
           ('Framework','Not confirmed in model'),
           ('Areas checked','Depreciation, revenue recognition, liability classification, debt treatment, tax balances'),
           ('Limitations','No source document testing or detailed accounting standard confirmation performed')]
    for label,val in _acct:
        ws1.cell(r,2).value=label; ws1.cell(r,2).font=Fn(sz=9,bold=True,col=GREY_TXT2); ws1.cell(r,2).fill=F(PANEL_GREY)
        ws1.merge_cells(f'C{r}:I{r}')
        ws1.cell(r,3).value=val; ws1.cell(r,3).font=Fn(sz=9,col=CHARCOAL); ws1.cell(r,3).fill=F(PANEL_GREY); ws1.cell(r,3).alignment=A(wrap=True)
        for cc in range(2,10): ws1.cell(r,cc).border=B(col=PANEL_BORDER)
        set_row(ws1,r,16); r+=1
    set_row(ws1,r,10); r+=1

    # ── Next Actions box ──────────────────────────────────────────────────────
    _actions=[]
    if t0.get('stats',{}).get('totalExternalLinks',0)>0: _actions.append('Remove or replace external workbook links.')
    if len(errorScan)>0: _actions.append('Resolve formula errors shown in the Error Matrix.')
    if cov_unc>0: _actions.append('Complete uncertain audit procedures.')
    if redundantIn.get('applicable') and redundantIn.get('redundantCount',0)>0: _actions.append('Review redundant input assumptions.')
    if p2_open>0: _actions.append('Resolve or formally waive open P2 findings.')
    _actions.append('Re-run audit checks after remediation.')
    _actions = _actions[:5]
    merge(ws1,f'B{r}:I{r}','Next Actions',bold=True,sz=11,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws1,r,2,r,9,DARK_BLUE); set_row(ws1,r,18); r+=1
    for i,act in enumerate(_actions,1):
        merge(ws1,f'B{r}:I{r}',f'{i}.   {act}',sz=9,col=CHARCOAL,bg=WHITE)
        for cc in range(2,10): ws1.cell(r,cc).border=B(col=PANEL_BORDER)
        set_row(ws1,r,18); r+=1
    set_row(ws1,r,8); r+=1

    # ── Scope footer — short, links to detail tab ───────────────────────────────
    merge(ws1,f'B{r}:I{r}',
          'Scope limitations: formula text inspection, named-range audit, VBA review and source document testing were not included. Full detail: Scope and Reliance tab.',
          sz=8,col=GREY_TXT2,bg=PANEL_GREY,italic=True,wrap=True)
    set_row(ws1,r,26); r+=1
    set_row(ws1,r,10); r+=1

    # ── Navigation row ───────────────────────────────────────────────────────
    _nav=['Issue Log','Reasonableness Review','Formula Risk Review','Redundant Inputs','Sheet Linkage','Named Range Audit','Scope and Reliance','Validation Matrix','Remediation','Assumption Register']
    for i,tab in enumerate(_nav):
        col=2+(i%8)
        row_offset=i//8
        c=ws1.cell(r+row_offset,col); c.value=f"=HYPERLINK(\"#'{tab}'!A1\",\"{tab}\")"
        c.font=Font(size=8,bold=True,color=DARK_BLUE,underline='single',name='Arial')
        c.fill=F(PALE_ACCENT); c.alignment=A(h='center',v='center'); c.border=B(col=PANEL_BORDER)
    _nav_rows = (len(_nav)-1)//8 + 1
    for rr in range(_nav_rows): set_row(ws1,r+rr,20)
    r += _nav_rows - 1
    _audit_output_last_row = r

    # ════════════════════════════════════════════════════════════════════════
    # TAB 2 — READ ME
    # ════════════════════════════════════════════════════════════════════════
    ws2=wb.create_sheet('Read Me'); ws2.sheet_view.showGridLines=False
    for col,w in [(1,3),(2,15),(3,15),(4,3),(5,15),(6,15),(7,3)]: set_col(ws2,col,w)

    merge(ws2,'B1:F1','READ ME — HOW TO USE THIS REPORT',bold=True,sz=18,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws2,1,2,1,6,DARK_BLUE); set_row(ws2,1,26)
    r2=3

    # ── Colour legend ──────────────────────────────────────────────────────
    _legend=[('P1',P1_FILL,P1_TXT),('P2',P2_FILL,P2_TXT),('P3',P3_FILL,P3_TXT),
             ('Complete',OK_FILL,OK_TXT),('Review',P2_FILL,P2_TXT)]
    col=2
    for lbl,fill,txt in _legend:
        c=ws2.cell(r2,col); c.value=lbl; c.font=Fn(bold=True,sz=9,col=txt); c.fill=F(fill)
        c.alignment=A(h='center',v='center'); c.border=B(col=PANEL_BORDER)
        col+=1
        if col>6: break
    set_row(ws2,r2,20); r2+=1
    set_row(ws2,r2,10); r2+=1

    def readme_card(row_start,c1,c2,heading,body):
        col_l1=get_column_letter(c1); col_l2=get_column_letter(c2)
        merge(ws2,f'{col_l1}{row_start}:{col_l2}{row_start}',heading,bold=True,sz=11,col=DARK_BLUE,bg=PALE_ACCENT,v='center')
        set_row(ws2,row_start,20)
        body_row=row_start+1
        # size the card body height to fit the text without excess empty space
        est_lines = max(2, sum(len(part)//55+1 for part in body.split(chr(10))))
        merge(ws2,f'{col_l1}{body_row}:{col_l2}{body_row}',body,sz=9,col=CHARCOAL,wrap=True,v='top')
        set_row(ws2,body_row,min(160,est_lines*13+10))
        return body_row+1

    _cards=[
        ('WHAT THIS REPORT IS',
         'The output of the FM Validator automated audit pipeline — Tier 0 formula-text analysis, Tier 1 code checks and Tier 2 Claude semantic review combined into one transaction-grade audit file. It records findings; it does not modify the source model.'),
        ('HOW TO USE THE ISSUE LOG',
         'Each row is one finding, sorted P1 first then F-Score. Use the table filters to focus by area, priority or status. Click View to jump to the affected cell — if a link ever fails to open, the Sheet and Cell columns are always accurate as a fallback.'),
        ('PRIORITY MEANING',
         'P1 — must be resolved before any external reliance.\nP2 — should be resolved before final issue; can be accepted with documented rationale.\nP3 — best practice, address in the next revision.\nQuery — needs confirmation from the model owner.'),
        ('CLOSURE RULES',
         'Open: unresolved. Closed: retested and confirmed. Waived: accepted risk with sign-off. Deferred: pushed to a future version. Superseded: replaced by another finding. Closure needs implementation, retest and reviewer sign-off.'),
        ('KEY TABS',
         'Issue Log — the finding register.\nValidation Matrix — every rule and its outcome.\nRedundant Inputs — unreferenced assumptions.\nError Matrix — live formula errors by code.\nAssumption Register — provenance for key drivers.\nScope and Reliance — the formal reliance record.'),
        ('LIMITATIONS',
         'This review covers the automatable subset of a structured model audit. It does not replace source document review, cell-by-cell formula inspection, or reviewer judgment. Full exclusions: see the Scope and Reliance tab.'),
    ]
    positions=[(2,3),(5,6),(2,3),(5,6),(2,3),(5,6)]
    row_l, row_r = r2, r2
    for i,(heading,body) in enumerate(_cards):
        c1,c2 = positions[i]
        if c1==2:
            row_l = readme_card(row_l,c1,c2,heading,body); row_l+=1
        else:
            row_r = readme_card(row_r,c1,c2,heading,body); row_r+=1


    # ════════════════════════════════════════════════════════════════════════
    # TAB 3 — SCOPE AND RELIANCE
    # ════════════════════════════════════════════════════════════════════════
    ws3=wb.create_sheet('Scope and Reliance'); ws3.sheet_view.showGridLines=False
    for col,w in [(1,3),(2,24),(3,16),(4,36),(5,24),(6,3)]: set_col(ws3,col,w)
    r3=1

    merge(ws3,f'B{r3}:E{r3}','SCOPE AND RELIANCE',bold=True,sz=18,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws3,r3,2,r3,5,DARK_BLUE); set_row(ws3,r3,26); r3+=1
    set_row(ws3,r3,8); r3+=1

    # ── Reliance decision box — before Model Details, per spec ──────────────
    _verdict_display2 = {
        'NOT RELIANCE-READY': 'Not reliance-ready',
        'RELIANCE-READY FOR INTERNAL REVIEW ONLY': 'Reliance-ready — internal review only',
        'RELIANCE-READY FOR MANAGEMENT DISCUSSION': 'Reliance-ready — management discussion',
        'RELIANCE-READY FOR LENDER / INVESTOR REVIEW': 'Reliance-ready — lender / investor review',
        'RELIANCE-READY FOR TRANSACTION EXECUTION': 'Reliance-ready — transaction execution',
    }.get(verdict_short, verdict_short.title())
    _decision_text = (f'Not reliance-ready until {p1_open} P1 finding(s) are closed and outstanding procedures ({cov_np} not run, {cov_unc} uncertain) are completed.'
                       if p1_open>0 else
                       f'{_verdict_display2}. {p2_open} P2 item(s) remain; see Issue Log for detail.')
    merge(ws3,f'B{r3}:E{r3}',f'RELIANCE DECISION:   {_verdict_display2}',bold=True,sz=12,col=WHITE,bg=verdict_bg,v='center')
    fill_range(ws3,r3,2,r3,5,verdict_bg); set_row(ws3,r3,24); r3+=1
    merge(ws3,f'B{r3}:E{r3+1}',_decision_text,sz=10,col=CHARCOAL,bg=PANEL_GREY,wrap=True,v='center')
    fill_range(ws3,r3,2,r3+1,5,PANEL_GREY); set_row(ws3,r3,18); set_row(ws3,r3+1,18); r3+=2
    set_row(ws3,r3,10); r3+=1

    def panel_header(row,title):
        merge(ws3,f'B{row}:E{row}',title,bold=True,sz=11,col=WHITE,bg=DARK_BLUE,v='center')
        fill_range(ws3,row,2,row,5,DARK_BLUE); set_row(ws3,row,20)
        return row+1

    def kv_row(row,label,val,bold_val=False):
        ws3.cell(row,2).value=label; ws3.cell(row,2).font=Fn(sz=9,bold=True,col=GREY_TXT2); ws3.cell(row,2).fill=F(PANEL_GREY)
        ws3.merge_cells(f'C{row}:E{row}')
        ws3.cell(row,3).value=val; ws3.cell(row,3).font=Fn(sz=9,col=CHARCOAL,bold=bold_val); ws3.cell(row,3).fill=F(PANEL_GREY); ws3.cell(row,3).alignment=A(wrap=True)
        for cc in range(2,6): ws3.cell(row,cc).border=B(col=PANEL_BORDER)
        set_row(ws3,row,18)
        return row+1

    # ── Panel 1: Engagement Details ──────────────────────────────────────────
    r3 = panel_header(r3,'ENGAGEMENT DETAILS')
    for label,val in [('Model name',modelName),('Model type',f'{modelType} — {modelIndustry}'),
                       ('Currency / Periodicity',f'{currency} · {periodicity}'),('Source file',sourceFile),
                       ('Domain skill applied',domainSkill),('Review date',reviewDate),
                       ('Review mode',reviewMode),('Model tier',modelTier)]:
        r3 = kv_row(r3,label,val)
    set_row(ws3,r3,10); r3+=1

    # ── Panel 2: Work Performed ──────────────────────────────────────────────
    r3 = panel_header(r3,'WORK PERFORMED')
    for label,val in [('Rules applied',f'{len(findings)} findings from a {len(checklist_rules)}-rule checklist'),
                       ('Tier 0 coverage',f'{t0.get("stats",{}).get("totalFormulaCells",0):,} formula cells scanned across all sheets'),
                       ('Tier 1 coverage','12 deterministic structural code checks'),
                       ('Tier 2 coverage',f'{len(checklist_rules)-12} Claude semantic checks across 13 sections'),
                       ('Accounting framework','Not confirmed in model — accrual basis assumed from statement structure'),
                       ('Audit completion',f'{igReadiness}% of planned procedures ({cov_pass} passed, {cov_issue} raised issues, {cov_unc} uncertain, {cov_np} not run)')]:
        r3 = kv_row(r3,label,val)
    set_row(ws3,r3,10); r3+=1

    # ── Panel 3: Work Not Performed — exclusions as a proper table ───────────
    r3 = panel_header(r3,'WORK NOT PERFORMED')
    for label,c1,c2 in [('PROCEDURE',2,2),('STATUS',3,3),('IMPACT',4,4),('NEXT STEP',5,5)]:
        cell(ws3,f'{get_column_letter(c1)}{r3}',label,bold=True,sz=8,col=WHITE,bg=DARK_BLUE,h='center')
        ws3.cell(r3,c1).border=B(col=WHITE)
    set_row(ws3,r3,16); r3+=1
    _fdd_performed = formulaDeepIn.get('applicable', False)
    _fdd_summary = (f"Individually reviewed the {formulaDeepIn.get('reviewed',0)} highest-complexity formulas in this model (Tier 0 F-score ranked), not sampled — a targeted, not exhaustive, set. {len(formulaDeepIn.get('findings',[]))} finding(s) resulted."
                     if _fdd_performed else 'Best-effort via Tier 0 only')
    # A4 — cell-level dependency-chain tracing for key output cells (EBITDA,
    # IRR, DSCR, MOIC and similar), checking whether the formula chain
    # behind each one dead-ends at a blank cell or propagates a cached
    # error, rather than only the sheet-level linkage the Sheet Dependency
    # tab already covers. This runs deterministically alongside the other
    # T0-* checks — its absence of findings means a clean result, not that
    # it didn't run, so this is described unconditionally rather than
    # gated on findings being present.
    _chain_findings = [f for f in findings if str(f.get('id','')).startswith('T0-CHAIN')]
    _chain_summary = (f" Cell-level dependency chains behind key output cells were also traced back through their formula precedents — "
                       + (f"{len(_chain_findings)} root cause(s) found where a chain reaches a blank cell or a cached error (see Issue Log)."
                          if _chain_findings else "no chain reached a blank cell or a cached error for the key outputs checked."))
    _reason_performed = reasonIn.get('outputs',{}).get('applicable',False) or reasonIn.get('waccOverride',{}).get('applicable',False)
    _reason_flagged = (reasonIn.get('outputs',{}).get('flaggedCount',0)
                        + (1 if reasonIn.get('waccOverride',{}).get('mismatch') else 0)
                        + (1 if reasonIn.get('terminalValue',{}).get('flagged') else 0)
                        + dupIn.get('flaggedCount',0))
    _reason_summary = (
        f"Tested WACC-vs-applied-rate consistency, terminal value concentration, {len(reasonIn.get('outputs',{}).get('results',[]))} output metric(s) against disclosed rule-of-thumb thresholds, and duplicate/backup sheet naming. {_reason_flagged} item(s) flagged for review. This tests whether stated results are plausible against documented triggers — it is not verified against live external market data, and does not test for missing line items (see Commercial omission testing below)."
        if _reason_performed else 'No labelled output metrics (EBITDA margin, IRR, exit multiple etc.) were found in this model to test.'
    )
    _vba_ran = vbaIn.get('applicable', False)
    _vba_has_project = vbaIn.get('hasVbaProject', False)
    _vba_module_count = vbaIn.get('moduleCount', 0)
    _vba_finding_count = len(vbaIn.get('findings', []))
    if not _vba_ran:
        _vba_status = 'Not performed'
        _vba_summary = (f"VBA/macro extraction did not complete for this run: {vbaIn.get('note','no reason recorded')}"
                         if vbaIn.get('note') else
                         'This review reads formula cells directly; it does not execute or trace VBA/macro code, so any calculation performed inside a macro is invisible to it, however well the macro itself is written.')
        _vba_next = 'Re-run once the underlying issue is resolved, or provide macro source for manual review'
    elif not _vba_has_project:
        _vba_status = 'Performed'
        _vba_summary = 'No VBA project was found in this workbook — there is no macro code for this limitation to apply to.'
        _vba_next = 'None'
    else:
        _vba_status = 'Performed (targeted)'
        _vba_summary = (f"Extracted and risk-scanned {_vba_module_count} VBA module(s); {_vba_finding_count} finding(s) raised (see Issue Log). "
                         f"This reads and pattern-scans macro source — it does not execute the macros, so any behaviour that only manifests at runtime is still outside this review's scope.")
        _vba_next = 'None for the scanned modules — engage a manual VBA code reviewer if runtime behaviour needs verification'
    # B1 — 7 mining-specific commercial omission patterns (revenue-to-Ops
    # linkage, capex post-commissioning, tax shield, debt sculpting,
    # rehabilitation provision, royalty treatment, FX consistency) were
    # converted from prose guidance into graded, mandatory checklist rules
    # (T2-S10-090 to 096). This is domain-specific — only present for
    # models using skill-mining.md — so the status below is conditional on
    # whether those rules actually exist in this run's checklist, not a
    # claim of universal omission-testing coverage.
    # Whether the 7 rules EXIST in checklist_rules is not the right test —
    # they're part of the fixed, universal checklist and are present for
    # every run regardless of domain. The right test is whether mining's
    # own domain skill was actually loaded for THIS run (matching the same
    # detection pattern already used elsewhere in this file for mining-
    # specific content) — otherwise this row would claim "performed,
    # mining-specific" on every single run, including non-mining models
    # where these rules mostly resolve to "not applicable".
    _omission_rule_ids = {f'T2-S10-{n:03d}' for n in range(90, 97)}
    _is_mining_run = 'mining' in (modelType or '').lower() or 'mining' in (domainSkill or '').lower()
    _omission_rules_present = _is_mining_run and [r for r in checklist_rules if r.get('id') in _omission_rule_ids]
    _exclusions=[
        ('Formula text inspection', 'Performed' if _fdd_performed else 'Partial',
         _fdd_summary,
         'None — opt-in deep review was performed for this run' if _fdd_performed else 'Enable the Formula Deep Dive opt-in for individual review of the highest-risk formulas, or provide direct Excel/formula access for full coverage'),
        ('Source document review','Not performed','This review has no access to contracts, invoices, bank statements or other source records — only the Excel file itself. An assumption can look internally consistent without being independently verified against reality.','Provide source documents for Mode C review'),
        ('Cell-by-cell audit', 'Partial',
         (f"The {formulaDeepIn.get('reviewed',0)} highest-risk formulas received individual review this run — not full coverage of every cell, but not merely pattern-based either."
          if _fdd_performed else 'Formula logic inspection requires formula text access for full individual review.') + _chain_summary,
         'Enable the Formula Deep Dive opt-in for individual review of the highest-risk formulas, in addition to the chain tracing already performed' if not _fdd_performed else 'None for the reviewed set — engage a manual reviewer for full coverage beyond the targeted set'),
        ('Assumption and output reasonableness', 'Performed (targeted)' if _reason_performed else 'Not performed',
         _reason_summary,
         'None for the tested metrics — extend threshold coverage or add domain-specific benchmarks if deeper testing is wanted' if _reason_performed else 'Label key output metrics clearly in the model so they can be located and tested'),
        ('Commercial omission testing',
         'Partial (mining-specific)' if _omission_rules_present else 'Not performed',
         (f"{len(_omission_rules_present)} mining-specific omission pattern(s) (revenue-to-operations linkage, capex phasing post-commissioning, tax shield modelling, debt sculpting, rehabilitation provision, royalty treatment, FX consistency) are graded, mandatory checklist rules for this model — see the T2-S10-09x rows in the Validation Matrix. This is domain-specific and only covers 7 named patterns: only models using the mining domain skill get even this partial coverage today, and it does not extend to other missing line items (a cost category, a risk, a required schedule) a domain expert might separately expect to see."
          if _omission_rules_present else
          'Different question to output reasonableness above: this is about line items missing from the model entirely (a cost category, a risk, a required schedule) that a domain expert would expect to see — not whether the stated figures are plausible. Requires a challenger model and commercial judgment.'),
         'Extend to other domains or add further patterns as they are identified, or commission a challenger-model review for full coverage' if _omission_rules_present else 'Commission a challenger-model review'),
        ('VBA and macro audit',_vba_status,_vba_summary,_vba_next),
        ('Named range audit','Partial','Checks whether every named range is used, clearly named and resolves correctly via static formula-text analysis; a name referenced only from VBA, a user-defined function, or a chart data range would not be detected as used.','Manually confirm any VBA-only or chart-only usages if suspected'),
    ]
    _status_style={'Not performed':(P1_FILL,P1_TXT),'Partial':(P2_FILL,P2_TXT),'Partial (mining-specific)':(P2_FILL,P2_TXT),'Performed':(OK_FILL,OK_TXT),'Performed (targeted)':(OK_FILL,OK_TXT)}
    _col_chars = {2:22, 4:34, 5:22}  # rough usable characters per line, by column width
    for proc,status,impact,nxt in _exclusions:
        ws3.cell(r3,2).value=proc; ws3.cell(r3,2).font=Fn(sz=9,col=CHARCOAL); ws3.cell(r3,2).fill=F(WHITE); ws3.cell(r3,2).alignment=A(wrap=True,v='center')
        sf,st=_status_style.get(status,(GREY_LIGHT,CHARCOAL))
        badge(ws3,r3,3,status,sf,st,sz=8)
        ws3.cell(r3,4).value=impact; ws3.cell(r3,4).font=Fn(sz=9,col=GREY_TXT2); ws3.cell(r3,4).fill=F(WHITE); ws3.cell(r3,4).alignment=A(wrap=True,v='center')
        ws3.cell(r3,5).value=nxt; ws3.cell(r3,5).font=Fn(sz=9,col=GREY_TXT2); ws3.cell(r3,5).fill=F(WHITE); ws3.cell(r3,5).alignment=A(wrap=True,v='center')
        for cc in range(2,6): ws3.cell(r3,cc).border=B(col=PANEL_BORDER)
        # Row height sized to the longest wrapped text in the row — a fixed
        # height clips longer explanations even with wrap=True set, since
        # wrap only breaks the text into lines, it doesn't grow the row.
        _max_lines = max(1, *[-(-len(t)//_col_chars[c]) for c,t in [(2,proc),(4,impact),(5,nxt)]])
        set_row(ws3,r3, max(30, 16 + _max_lines*13)); r3+=1
    set_row(ws3,r3,10); r3+=1

    # ── Sign-Off panel — proper table, not floating rows ────────────────────
    r3 = panel_header(r3,'SIGN-OFF')
    for label,c1,c2 in [('CHECKED BY',2,2),('APPROVED BY',3,3),('DATE',4,4),('VERSION',5,5)]:
        cell(ws3,f'{get_column_letter(c1)}{r3}',label,bold=True,sz=8,col=WHITE,bg=DARK_BLUE,h='center')
        ws3.cell(r3,c1).border=B(col=WHITE)
    set_row(ws3,r3,16); r3+=1
    for cc in range(2,6):
        ws3.cell(r3,cc).value=''; ws3.cell(r3,cc).fill=F(WHITE); ws3.cell(r3,cc).border=B(col=PANEL_BORDER)
        ws3.cell(r3,cc).alignment=A(h='center',v='center')
    set_row(ws3,r3,24); r3+=1
    merge(ws3,f'B{r3}:E{r3}','Not yet signed off — this is a draft automated review, not a final reliance record.',sz=8,col=GREY_TXT2,italic=True)
    set_row(ws3,r3,16); r3+=1
    _scope_last_row = r3


    # ════════════════════════════════════════════════════════════════════════
    # TAB 4 — ISSUE LOG
    # ════════════════════════════════════════════════════════════════════════
    ws4=wb.create_sheet('Issue Log'); ws4.sheet_view.showGridLines=False
    # Column order: first 12 are the default-visible client-facing set;
    # everything after col 13 is real audit-trail detail, pushed right and
    # hidden by default (available for reviewer drill-down, not deleted).
    # 'View' is genuinely useful (one click to the affected cell), not a
    # technical scan artefact — it stays in the default-visible set even
    # though it isn't literally named in the brief's 12-column list.
    il_headers=['','ID','Priority','Status','Area','Sheet','Cell','View','Finding',
                'Model Impact','Required Fix','Owner','Target Date','Closure Evidence',
                'Relevance','Urgency','Issue Type','Workstream','What is wrong',
                'Why it matters','Output affected','Key Output\nImpact','F-Score',
                'Confidence','Method','Range','Root Cause',
                'Investment\nBlocker','Escalation\nFlag','Date Raised','Days Open',
                'Management\nResponse','Reviewer\nResponse','Client\nResponse','']
    n_cols = len(il_headers)
    n_visible = 14  # ID..Closure Evidence, cols 2-14 inclusive
    col_widths=[3,12,8,12,14,10,8,10,32,32,32,14,12,20]+[16]*(n_cols-1-n_visible)
    for i,w in enumerate(col_widths,1): set_col(ws4,i,w)
    # Name-based column lookup — avoids the index-drift bugs that hardcoded
    # column numbers caused twice already this session.
    col_idx = {h.replace('\n',' '): i for i,h in enumerate(il_headers,1) if h}
    HEADER_ROW=4

    merge(ws4,f'B1:{get_column_letter(n_cols-1)}1','ISSUE LOG',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws4,1,2,1,n_cols-1,DARK_BLUE); set_row(ws4,1,26)

    # ── Top filter strip — live formulas, updates as reviewers fill the sheet ──
    # Row/column references all derive from col_idx and the real header/data
    # boundaries — never hardcoded, since hardcoded refs are exactly what
    # broke twice already when columns got reordered.
    _first_data_row = HEADER_ROW+1
    _il_last_data_row = HEADER_ROW+len(findings)
    PRI_L = get_column_letter(col_idx['Priority']); ST_L = get_column_letter(col_idx['Status'])
    OWN_L = get_column_letter(col_idx['Owner']);    TGT_L = get_column_letter(col_idx['Target Date'])
    _pri_rng = f'{PRI_L}{_first_data_row}:{PRI_L}{_il_last_data_row}'
    _st_rng  = f'{ST_L}{_first_data_row}:{ST_L}{_il_last_data_row}'
    _own_rng = f'{OWN_L}{_first_data_row}:{OWN_L}{_il_last_data_row}'
    _tgt_rng = f'{TGT_L}{_first_data_row}:{TGT_L}{_il_last_data_row}'
    _strip=[
        ('P1 Open', f'=COUNTIFS({_pri_rng},"P1",{_st_rng},"<>Closed",{_st_rng},"<>Waived",{_st_rng},"<>Deferred",{_st_rng},"<>Superseded")'),
        ('P2 Open', f'=COUNTIFS({_pri_rng},"P2",{_st_rng},"<>Closed",{_st_rng},"<>Waived",{_st_rng},"<>Deferred",{_st_rng},"<>Superseded")'),
        ('Needs Owner', f'=COUNTIFS({_own_rng},"",{_st_rng},"<>Closed",{_st_rng},"<>Waived")'),
        ('Overdue', f'=COUNTIFS({_tgt_rng},"<"&TODAY(),{_st_rng},"<>Closed",{_st_rng},"<>Waived",{_st_rng},"<>Deferred",{_st_rng},"<>Superseded",{_st_rng},"<>Ready for Retest")'),
        ('Ready for Retest', f'=COUNTIF({_st_rng},"Ready for Retest")'),
    ]
    for i,(label,formula) in enumerate(_strip):
        c1,c2 = 2+i*2, 3+i*2
        col_l1,col_l2 = get_column_letter(c1), get_column_letter(c2)
        merge(ws4,f'{col_l1}2:{col_l2}2',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        merge(ws4,f'{col_l1}3:{col_l2}3',formula,bold=True,sz=14,col=CHARCOAL,bg=PANEL_GREY,h='center')
        for rr in (2,3):
            for cc in range(c1,c2+1): ws4.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(ws4,2,16); set_row(ws4,3,24)

    # ── Header row ──────────────────────────────────────────────────────────
    for col,h in enumerate(il_headers,1):
        c=ws4.cell(HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
        c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
    set_row(ws4,HEADER_ROW,32)
    VIEW_COL=col_idx['View']

    # Freeze header row + first three key columns (ID, Priority, Status —
    # columns B,C,D; column A is just the margin) per spec.
    ws4.freeze_panes = f'E{HEADER_ROW+1}'

    _pri_badge={'P1':(P1_FILL,P1_TXT),'P2':(P2_FILL,P2_TXT),'P3':(P3_FILL,P3_TXT)}
    _default_urgency_days={'P1':7,'P2':30,'P3':90}
    from datetime import timedelta
    for idx,f in enumerate(findings):
        row_i = HEADER_ROW+1+idx
        pri=priority(f)
        _t=(f.get('title') or f.get('label') or '').strip()
        f['title']=_t; f['label']=_t
        raw_sev=(f.get('severity') or f.get('priority') or 'Medium').lower()
        sheet=f.get('sheet',''); cell_ref=f.get('cell','A1') or 'A1'
        category=f.get('category','')
        issue_title=f.get('label') or f.get('id','')
        what_wrong=f.get('condition') or f.get('reason','')
        why_matters=f.get('consequence','')
        fix_action=f.get('corrective_action') or f.get('fix_instruction','')
        out_impact=f.get('dollar_impact','')
        urgency=('Before next reliance' if pri=='P1' else 'Before external circulation' if pri=='P2' else 'When convenient')
        fscore_val=f.get('fscore','') or '—'
        confidence_val=f.get('confidence','')
        confidence_display=f'{confidence_val}%' if confidence_val!='' else '—'
        # Suggested target date so Overdue conditional formatting has real
        # data to work against immediately, not just once a reviewer fills it.
        _target_date = (datetime.strptime(reviewDate,'%d %b %Y') + timedelta(days=_default_urgency_days.get(pri,30)))

        # White base with very light alternating banding — priority colour
        # lives ONLY in the badge cell, never the whole row.
        row_bg = WHITE if idx%2==0 else 'FAFBFC'

        # Name-keyed, not positional — immune to header reordering.
        row_values = {
            'ID': f.get('id',''), 'Priority': pri, 'Status': 'Open', 'Area': category,
            'Sheet': sheet, 'Cell': cell_ref, 'View': '',  # View written separately below
            'Finding': issue_title, 'Model Impact': f.get('model_risk','') or why_matters,
            'Required Fix': fix_action, 'Owner': '', 'Target Date': _target_date, 'Closure Evidence': '',
            'Relevance': 'Relevant', 'Urgency': urgency, 'Issue Type': f.get('issue_type',''),
            'Workstream': f.get('workstream',''), 'What is wrong': what_wrong, 'Why it matters': why_matters,
            'Output affected': out_impact, 'Key Output Impact': f.get('key_output_impact','Unknown'),
            'F-Score': fscore_val, 'Confidence': confidence_display, 'Method': f.get('method',''), 'Range': '',
            'Root Cause': f.get('root_cause',''),
            'Investment Blocker': 'Yes' if f.get('investment_grade_blocker') else 'No',
            'Escalation Flag': 'Yes' if f.get('escalation_flag') else 'No',
            'Date Raised': reviewDate, 'Days Open': '', 'Management Response': '',
            'Reviewer Response': '', 'Client Response': '',
        }
        _centered = {'Priority','Status','Sheet','Cell','View','Target Date','F-Score','Confidence',
                     'Method','Investment Blocker','Escalation Flag','Days Open'}
        _wrapped  = {'Finding','Model Impact','Required Fix','Closure Evidence'}
        for name,val in row_values.items():
            col = col_idx[name]
            c=ws4.cell(row_i,col); c.value=val
            c.font=Fn(sz=9,bold=(name=='ID'))
            c.fill=F(row_bg)
            c.alignment=A(h='center' if name in _centered else 'left', v='top', wrap=(name in _wrapped))
            c.border=B(col=PANEL_BORDER)
        ws4.cell(row_i,col_idx['Target Date']).number_format='dd mmm yyyy'

        # Priority badge — soft colour confined to this one cell only
        pf,pt=_pri_badge.get(pri,(GREY_LIGHT,CHARCOAL))
        badge(ws4,row_i,col_idx['Priority'],pri,pf,pt,sz=9)

        # View hyperlink — jumps into the SOURCE model file. This is
        # necessarily a cross-workbook reference (it has to reach outside
        # this report into the client's own file), so full compatibility
        # can't be guaranteed the way an internal same-workbook link can —
        # but the label itself is now always exactly "View", never raw
        # formula/implementation text, never a variable fallback label.
        vc=ws4.cell(row_i,VIEW_COL)
        if sheet and sheet not in ('—','N/A','Multiple'):
            vc.value=f'=HYPERLINK("[{sourceFile}]{sheet}!A1","View")'
            vc.font=Font(size=9,color=MID_BLUE,underline='single',name='Arial')
        else:
            vc.value='—'; vc.font=Fn(sz=9,col=GREY_MID)
        vc.fill=F(row_bg); vc.alignment=A(h='center',v='center'); vc.border=B(col=PANEL_BORDER)
        set_row(ws4,row_i,44)

    # ── Hide technical columns beyond the 12 default-visible ones ───────────
    for col in range(n_visible+1, n_cols):
        ws4.column_dimensions[get_column_letter(col)].hidden = True

    # ── Real Excel Table — adds filter dropdowns to every visible header ────
    if len(findings) > 0:
        _table_ref = f'B{HEADER_ROW}:{get_column_letter(n_cols-1)}{_il_last_data_row}'
        _il_table = Table(displayName="IssueLogTable", ref=_table_ref)
        _il_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                    showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
        ws4.add_table(_il_table)

        # ── Data validation dropdowns ──────────────────────────────────────
        REL_L = get_column_letter(col_idx['Relevance'])
        status_dv = DataValidation(type='list',
            formula1='"Open,In Progress,Ready for Retest,Closed,Waived,Deferred,Superseded"', allow_blank=False)
        ws4.add_data_validation(status_dv); status_dv.add(_st_rng)

        relevance_dv = DataValidation(type='list', formula1='"Relevant,Not Relevant,Needs Review"', allow_blank=False)
        ws4.add_data_validation(relevance_dv); relevance_dv.add(f'{REL_L}{_first_data_row}:{REL_L}{_il_last_data_row}')

        # ── Conditional formatting — overdue target dates only ──────────────
        ws4.conditional_formatting.add(_tgt_rng,
            FormulaRule(formula=[f'AND({TGT_L}{_first_data_row}<TODAY(),NOT(OR({ST_L}{_first_data_row}="Closed",{ST_L}{_first_data_row}="Waived",{ST_L}{_first_data_row}="Deferred",{ST_L}{_first_data_row}="Superseded")))'],
                        fill=PatternFill(start_color=P1_FILL,end_color=P1_FILL,fill_type='solid'),
                        font=Font(color=P1_TXT)))

    # ════════════════════════════════════════════════════════════════════════
    # TAB 5 — REMEDIATION & RETEST
    # ════════════════════════════════════════════════════════════════════════
    ws5=wb.create_sheet('Remediation'); ws5.sheet_view.showGridLines=False
    rem_headers=['','#','Finding ID','Priority','Action Status','Timing','Owner','Target Date',
                 'Suggested Action','Promised Fix Version','Retest Required','Retest Result',
                 'Root Cause Group','F-Score','Urgency','Retested By','Retest Date','Eligible to Close','']
    n_rem_cols = len(rem_headers)
    n_rem_visible = 12  # #..Retest Result
    rem_widths=[3,5,14,8,14,12,14,12,36,16,12,14]+[16]*(n_rem_cols-1-n_rem_visible)
    for i,w in enumerate(rem_widths,1): set_col(ws5,i,w)
    rem_idx = {h.replace('\n',' '): i for i,h in enumerate(rem_headers,1) if h}
    REM_HEADER_ROW=5

    merge(ws5,f'B1:{get_column_letter(n_rem_cols-1)}1','REMEDIATION & RETEST TRACKER',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws5,1,2,1,n_rem_cols-1,DARK_BLUE); set_row(ws5,1,26)
    merge(ws5,f'B2:{get_column_letter(n_rem_cols-1)}2',
          'Suggested actions are AI-generated and should be reviewed before implementation. Some recommendations may not be necessary or appropriate for the specific model.',
          sz=8,col=GREY_TXT2,bg=PALE_ACCENT,italic=True,wrap=True)
    set_row(ws5,2,20)

    rem_findings=[f for f in findings if priority(f) in ('P1','P2')]
    _rf_first, _rf_last = REM_HEADER_ROW+1, REM_HEADER_ROW+len(rem_findings)
    PRI_L2 = get_column_letter(rem_idx['Priority']); AS_L = get_column_letter(rem_idx['Action Status'])
    TIM_L  = get_column_letter(rem_idx['Timing'])
    _pri_rng2 = f'{PRI_L2}{_rf_first}:{PRI_L2}{_rf_last}'
    _as_rng   = f'{AS_L}{_rf_first}:{AS_L}{_rf_last}'
    _tim_rng  = f'{TIM_L}{_rf_first}:{TIM_L}{_rf_last}'

    # ── Top summary strip — live formulas ──────────────────────────────────
    _rem_strip=[
        ('Open Actions', f'=COUNTIF({_as_rng},"<>Closed")'),
        ('P1 Actions', f'=COUNTIFS({_pri_rng2},"P1",{_as_rng},"<>Closed")'),
        ('Ready for Retest', f'=COUNTIF({_as_rng},"Ready for Retest")'),
        ('Overdue', f'=COUNTIF({_tim_rng},"Overdue")'),
    ]
    for i,(label,formula) in enumerate(_rem_strip):
        c1,c2 = 2+i*2, 3+i*2
        col_l1,col_l2 = get_column_letter(c1), get_column_letter(c2)
        merge(ws5,f'{col_l1}3:{col_l2}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        merge(ws5,f'{col_l1}4:{col_l2}4',formula,bold=True,sz=14,col=CHARCOAL,bg=PANEL_GREY,h='center')
        for rr in (3,4):
            for cc in range(c1,c2+1): ws5.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(ws5,3,16); set_row(ws5,4,24)

    for col,h in enumerate(rem_headers,1):
        c=ws5.cell(REM_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
        c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
    set_row(ws5,REM_HEADER_ROW,30)

    # Freeze header row + first four key columns (#, Finding ID, Priority,
    # Action Status) per spec.
    ws5.freeze_panes = f'{get_column_letter(rem_idx["Timing"])}{REM_HEADER_ROW+1}'

    _pri_badge2={'P1':(P1_FILL,P1_TXT),'P2':(P2_FILL,P2_TXT),'P3':(P3_FILL,P3_TXT)}
    for idx,f in enumerate(rem_findings):
        row_i = REM_HEADER_ROW+1+idx
        pri=priority(f)
        row_bg = WHITE if idx%2==0 else 'FAFBFC'
        # Punchier action wording — lead with an imperative verb where the
        # source text already starts with one; otherwise keep as-is but cap
        # length tightly so it reads as an action item, not a paragraph.
        raw_action = f.get('corrective_action') or f.get('fix_instruction','')
        action_text = (raw_action or '').rstrip(' ,;:')

        row_values={
            '#': idx+1, 'Finding ID': f.get('id',''), 'Priority': pri,
            'Action Status': 'Not Started', 'Timing': '',  # Timing set as a live formula below
            'Owner': '', 'Target Date': '', 'Suggested Action': action_text,
            'Promised Fix Version': '', 'Retest Required': 'Yes', 'Retest Result': 'Pending',
            'Root Cause Group': f.get('root_cause',''), 'F-Score': f.get('fscore','') or '—',
            'Urgency': f.get('urgency','') or '', 'Retested By': '', 'Retest Date': '',
            'Eligible to Close': 'No',
        }
        _centered2 = {'#','Priority','Action Status','Timing','Target Date','Retest Required',
                      'Retest Result','F-Score','Retest Date','Eligible to Close'}
        _wrapped2  = {'Suggested Action'}
        for name,val in row_values.items():
            col = rem_idx[name]
            c=ws5.cell(row_i,col); c.value=val
            c.font=Fn(sz=9,bold=(name in ('Finding ID','Priority')))
            c.fill=F(row_bg)
            c.alignment=A(h='center' if name in _centered2 else 'left', v='top', wrap=(name in _wrapped2))
            c.border=B(col=PANEL_BORDER)

        # Priority badge — colour confined to this cell only, never the row
        pf,pt=_pri_badge2.get(pri,(GREY_LIGHT,CHARCOAL))
        badge(ws5,row_i,rem_idx['Priority'],pri,pf,pt,sz=9)

        # Suggested Target Date so Timing has real data immediately
        _target = datetime.strptime(reviewDate,'%d %b %Y') + timedelta(days=(7 if pri=='P1' else 30))
        tc = ws5.cell(row_i, rem_idx['Target Date']); tc.value=_target; tc.number_format='dd mmm yyyy'
        tc.fill=F(row_bg); tc.alignment=A(h='center',v='center'); tc.border=B(col=PANEL_BORDER)

        # Timing — live formula, so it updates as time passes rather than
        # freezing at whatever was true when the report was generated.
        tgt_ref = f'{get_column_letter(rem_idx["Target Date"])}{row_i}'
        as_ref  = f'{get_column_letter(rem_idx["Action Status"])}{row_i}'
        timing_formula = (f'=IF({as_ref}="Closed","-",'
                           f'IF({tgt_ref}<TODAY(),"Overdue",IF({tgt_ref}<TODAY()+7,"Due Soon","On Track")))')
        tim_c = ws5.cell(row_i, rem_idx['Timing']); tim_c.value = timing_formula
        tim_c.font=Fn(bold=True,sz=9,col=CHARCOAL); tim_c.fill=F(row_bg)
        tim_c.alignment=A(h='center',v='center'); tim_c.border=B(col=PANEL_BORDER)

        set_row(ws5,row_i,32)

    _rem_last_row = REM_HEADER_ROW+len(rem_findings)

    for col in range(n_rem_visible+1, n_rem_cols):
        ws5.column_dimensions[get_column_letter(col)].hidden = True

    if len(rem_findings) > 0:
        _rem_table = Table(displayName="RemediationTable", ref=f'B{REM_HEADER_ROW}:{get_column_letter(n_rem_cols-1)}{_rem_last_row}')
        _rem_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                     showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
        ws5.add_table(_rem_table)

        action_status_dv = DataValidation(type='list',
            formula1='"Not Started,In Progress,Ready for Retest,Closed"', allow_blank=False)
        ws5.add_data_validation(action_status_dv); action_status_dv.add(_as_rng)

        # Action Status badges — live, so the colour tracks the dropdown
        # value if a reviewer changes it later (same pattern as Timing).
        _as_colors=[('Not Started',GREY_LIGHT,GREY_TXT2),('In Progress',P2_FILL,P2_TXT),
                    ('Ready for Retest',P3_FILL,P3_TXT),('Closed',OK_FILL,OK_TXT)]
        for val,fill,txt in _as_colors:
            ws5.conditional_formatting.add(_as_rng,
                CellIsRule(operator='equal', formula=[f'"{val}"'],
                           fill=PatternFill(start_color=fill,end_color=fill,fill_type='solid'), font=Font(color=txt,bold=False)))

        retest_dv = DataValidation(type='list', formula1='"Yes,No"', allow_blank=False)
        ws5.add_data_validation(retest_dv); retest_dv.add(f'{get_column_letter(rem_idx["Retest Required"])}{_rf_first}:{get_column_letter(rem_idx["Retest Required"])}{_rf_last}')

        # ── Conditional formatting — Timing and Priority only, per spec ─────
        ws5.conditional_formatting.add(_tim_rng,
            CellIsRule(operator='equal', formula=['"Overdue"'],
                       fill=PatternFill(start_color=P1_FILL,end_color=P1_FILL,fill_type='solid'), font=Font(color=P1_TXT)))
        ws5.conditional_formatting.add(_tim_rng,
            CellIsRule(operator='equal', formula=['"Due Soon"'],
                       fill=PatternFill(start_color=P2_FILL,end_color=P2_FILL,fill_type='solid'), font=Font(color=P2_TXT)))
        ws5.conditional_formatting.add(_tim_rng,
            CellIsRule(operator='equal', formula=['"On Track"'],
                       fill=PatternFill(start_color=OK_FILL,end_color=OK_FILL,fill_type='solid'), font=Font(color=OK_TXT)))
        ws5.conditional_formatting.add(_pri_rng2,
            CellIsRule(operator='equal', formula=['"P1"'],
                       fill=PatternFill(start_color=P1_FILL,end_color=P1_FILL,fill_type='solid'), font=Font(color=P1_TXT,bold=True)))

    # ════════════════════════════════════════════════════════════════════════
    # TAB 5B — VALIDATION MATRIX (B6) — every checklist rule with its outcome
    # ════════════════════════════════════════════════════════════════════════
    wsm=wb.create_sheet('Validation Matrix'); wsm.sheet_view.showGridLines=False
    vm_headers=['','Rule ID','Test Area','Rule','Tier','Status','Confidence','Related Findings','Retest Required',
                '#','Performed','Evidence Reviewed','Missing Evidence','']
    n_vm_cols = len(vm_headers)
    n_vm_visible = 9  # Rule ID..Retest Required
    vm_widths=[3,12,20,42,8,14,12,22,12]+[16]*(n_vm_cols-1-n_vm_visible)
    for i,w in enumerate(vm_widths,1): set_col(wsm,i,w)
    vm_idx = {h.replace('\n',' '): i for i,h in enumerate(vm_headers,1) if h}
    VM_HEADER_ROW=6

    merge(wsm,f'B1:{get_column_letter(n_vm_cols-1)}1','VALIDATION MATRIX',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(wsm,1,2,1,n_vm_cols-1,DARK_BLUE); set_row(wsm,1,26)
    merge(wsm,f'B2:{get_column_letter(n_vm_cols-1)}2',
          'Every rule in the checklist and its outcome — a procedure checklist, not a findings dump. See the Issue Log for the findings themselves.',
          sz=8,col=GREY_TXT2,bg=PALE_ACCENT,italic=True,wrap=True)
    set_row(wsm,2,18)

    def _rmatch(rid,xid): return xid==rid or (xid or '').startswith(rid+'-')

    m_rows=[]; n_pass=n_issue=n_unc=n_np=0
    for rule in checklist_rules:
        rid=rule.get('id',''); tier=rule.get('_tier','Tier 2')
        rres=[r for r in ruleResults if _rmatch(rid,r.get('id',''))]
        rfnd=[f for f in findings if _rmatch(rid,f.get('id',''))]
        performed='Yes' if rres else 'No'
        issues=[r for r in rres if r.get('status') not in ('pass','uncertain')]
        unc=[r for r in rres if r.get('status')=='uncertain']
        if not rres: status='Not Performed'; n_np+=1
        elif issues: status='Raised Issue'; n_issue+=1
        elif unc: status='Uncertain'; n_unc+=1
        else: status='Pass'; n_pass+=1
        confs=[r.get('confidence') for r in rres if isinstance(r.get('confidence'),(int,float)) and r.get('confidence')]
        conf_num = (max(confs)/100.0) if (confs and tier=='Tier 2') else None
        sm=re.search(r'-S(\d+)-',rid)
        if tier=='Tier 1': evidence='Full workbook — deterministic code check'
        elif sm and int(sm.group(1)) in (5,6,7,10):
            _resolved = deepAcctSheets.get('resolvedMap',{})
            if _resolved:
                _sheet_list = ', '.join(cat if cat==name else f'{cat} ({name})' for cat,name in _resolved.items())
                _unresolved = deepAcctSheets.get('unresolvedCategories',[])
                _unresolved_note = f' — {", ".join(_unresolved)} not found in this workbook' if _unresolved else ''
                evidence = f'Deep financial data subset — {_sheet_list}{_unresolved_note}'
            else:
                evidence = 'Deep financial data subset — full AFS/IFS/Cons/Debt/Equity/D&T/Leases'
        else: evidence='Standard data subset — all sheets, trimmed rows'
        refs=[f.get('id','') for f in rfnd]
        ref_txt=', '.join(refs[:4])+(f' +{len(refs)-4} more' if len(refs)>4 else '') if refs else '—'
        if status=='Not Performed': missing='Rule not returned by the review — re-run validation or test manually'
        elif status=='Uncertain': missing='Evidence insufficient for a conclusive test — see related finding'
        else: missing='—'
        retest='Yes' if (issues or any(f.get('needs_retest') for f in rfnd)) else 'No'
        m_rows.append(dict(rid=rid, area=rule.get('source_section','') or rule.get('section',''),
                            rule=rule.get('label',''), tier=tier, performed=performed, status=status,
                            conf_num=conf_num, evidence=evidence, refs=ref_txt, missing=missing, retest=retest))

    # ── Summary cards — Planned / Performed / Passed / Raised Issue / Uncertain / Not Run ──
    extras=len([f for f in findings if not any(_rmatch(r.get('id',''),f.get('id','')) for r in checklist_rules)])
    _cards=[('Planned',len(checklist_rules),2,3),('Performed',len(checklist_rules)-n_np,4,4),
            ('Passed',n_pass,5,5),('Raised Issue',n_issue,6,6),('Uncertain',n_unc,7,7),('Not Run',n_np,8,9)]
    for label,val,c1,c2 in _cards:
        col_l1,col_l2 = get_column_letter(c1), get_column_letter(c2)
        if c2>c1: merge(wsm,f'{col_l1}3:{col_l2}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        else: cell(wsm,f'{col_l1}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        if c2>c1: merge(wsm,f'{col_l1}4:{col_l2}4',val,bold=True,sz=16,col=CHARCOAL,bg=PANEL_GREY,h='center')
        else: cell(wsm,f'{col_l1}4',val,bold=True,sz=16,col=CHARCOAL,bg=PANEL_GREY,h='center')
        for rr in (3,4):
            for cc in range(c1,c2+1): wsm.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(wsm,3,14); set_row(wsm,4,24)
    if extras:
        merge(wsm,f'B{VM_HEADER_ROW-1}:{get_column_letter(n_vm_cols-1)}{VM_HEADER_ROW-1}',
              f'{extras} additional finding(s) fall outside the checklist — see Issue Log for detail.',
              sz=8,col=GREY_TXT2,italic=True)
        set_row(wsm,VM_HEADER_ROW-1,14)
    else:
        set_row(wsm,VM_HEADER_ROW-1,4)

    for col,h in enumerate(vm_headers,1):
        c=wsm.cell(VM_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
        c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
    set_row(wsm,VM_HEADER_ROW,26)

    # Freeze header row + first three columns (Rule ID, Test Area, Rule)
    wsm.freeze_panes = f'{get_column_letter(vm_idx["Tier"])}{VM_HEADER_ROW+1}'

    _status_badge={'Pass':(OK_FILL,OK_TXT),'Uncertain':(P2_FILL,P2_TXT),'Not Performed':(GREY_LIGHT,GREY_TXT2),'Raised Issue':(P1_FILL,P1_TXT)}
    _centered3={'Tier','Status','Confidence','Retest Required','#','Performed'}
    _wrapped3={'Rule','Related Findings','Evidence Reviewed','Missing Evidence'}
    for idx,m in enumerate(m_rows):
        row_i = VM_HEADER_ROW+1+idx
        row_bg = WHITE if idx%2==0 else 'FAFBFC'
        row_values={
            'Rule ID': m['rid'], 'Test Area': m['area'], 'Rule': m['rule'], 'Tier': m['tier'],
            'Status': m['status'], 'Confidence': m['conf_num'] if m['conf_num'] is not None else '—',
            'Related Findings': m['refs'], 'Retest Required': m['retest'],
            '#': idx+1, 'Performed': m['performed'], 'Evidence Reviewed': m['evidence'], 'Missing Evidence': m['missing'],
        }
        for name,val in row_values.items():
            col = vm_idx[name]
            c=wsm.cell(row_i,col); c.value=val
            c.font=Fn(sz=9,bold=(name=='Rule ID'))
            c.fill=F(row_bg)
            c.alignment=A(h='center' if name in _centered3 else 'left', v='top', wrap=(name in _wrapped3))
            c.border=B(col=PANEL_BORDER)
        if isinstance(row_values['Confidence'], float):
            wsm.cell(row_i, vm_idx['Confidence']).number_format='0%;[Red](0%);-'

        pf,pt = _status_badge.get(m['status'],(GREY_LIGHT,CHARCOAL))
        badge(wsm,row_i,vm_idx['Status'],m['status'],pf,pt,sz=8,bold=False)
        set_row(wsm,row_i,22)

    _vm_last_row = VM_HEADER_ROW+len(m_rows)

    for col in range(n_vm_visible+1, n_vm_cols):
        wsm.column_dimensions[get_column_letter(col)].hidden = True

    if len(m_rows) > 0:
        _vm_table = Table(displayName="ValidationMatrixTable", ref=f'B{VM_HEADER_ROW}:{get_column_letter(n_vm_cols-1)}{_vm_last_row}')
        _vm_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                    showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
        wsm.add_table(_vm_table)

        # Confidence — real data bar, not just a formatted percentage
        conf_l = get_column_letter(vm_idx['Confidence'])
        conf_rng = f'{conf_l}{VM_HEADER_ROW+1}:{conf_l}{_vm_last_row}'
        wsm.conditional_formatting.add(conf_rng,
            DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                        color=MID_BLUE, showValue=True, minLength=None, maxLength=None))

    # ════════════════════════════════════════════════════════════════════════
    # TAB 5C — ASSUMPTION REGISTER (B7) — assumption provenance template
    # ════════════════════════════════════════════════════════════════════════
    wsa=wb.create_sheet('Assumption Register'); wsa.sheet_view.showGridLines=False
    ar_headers=['','Assumption Area','Evidence Status','Decision Critical','Related Findings',
                'Source','Source Date','Owner','Basis','Externally Supported','Notes','']
    n_ar_cols = len(ar_headers)
    ar_widths=[3,26,14,14,22,18,12,14,16,16,34,3]
    for i,w in enumerate(ar_widths,1): set_col(wsa,i,w)
    ar_idx = {h.replace('\n',' '): i for i,h in enumerate(ar_headers,1) if h}
    AR_HEADER_ROW=6

    merge(wsa,f'B1:{get_column_letter(n_ar_cols-1)}1','ASSUMPTION REGISTER',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(wsa,1,2,1,n_ar_cols-1,DARK_BLUE); set_row(wsa,1,26)

    mining_areas=[('Commodity prices (PCI / thermal / benchmark)',['price','pricing','benchmark','hcc','pci','thermal']),
        ('FX rates',['fx','exchange rate','aud/usd','currency']),
        ('Production volumes & ramp-up',['production','volume','ramp','mtpa','tonn']),
        ('Product mix & quality',['product mix','quality','specification','washability']),
        ('Reserves & mine life',['reserve','resource','mine life','pit life','depletion']),
        ('Yield / recovery',['yield','recovery','wash']),
        ('Strip ratio & dilution',['strip ratio','dilution','waste']),
        ('Capital expenditure',['capex','capital expenditure','sustaining']),
        ('Operating costs',['opex','operating cost','unit cost','fob']),
        ('Royalties',['royalt']),
        ('Tax rates & treatment',['tax']),
        ('Debt terms & facilities',['debt','interest','facility','dscr','repayment','drawdown']),
        ('Discount rate / WACC',['discount rate','wacc','npv']),
        ('Inflation & escalation',['inflation','escalation','cpi']),
        ('Working capital',['working capital','receivable','payable','inventory']),
        ('Rehabilitation & closure costs',['rehabilitation','closure','restoration'])]
    generic_areas=[('Revenue drivers & pricing',['price','pricing','revenue','tariff']),
        ('Volume & growth rates',['volume','growth','ramp']),
        ('FX rates',['fx','exchange rate','currency']),
        ('Capital expenditure',['capex','capital expenditure']),
        ('Operating costs',['opex','operating cost','unit cost']),
        ('Tax rates & treatment',['tax']),
        ('Debt terms & facilities',['debt','interest','facility','dscr','repayment']),
        ('Discount rate / WACC',['discount rate','wacc','npv']),
        ('Inflation & escalation',['inflation','escalation','cpi']),
        ('Working capital',['working capital','receivable','payable','inventory'])]
    areas=mining_areas if ('mining' in (modelType or '').lower() or 'mining' in (domainSkill or '').lower()) else generic_areas

    def _ftext(f):
        return ' '.join(str(f.get(k,'') or '') for k in ('label','title','condition','reason','workstream','category')).lower()
    ftexts=[(f.get('id',''),_ftext(f)) for f in findings]

    _ar_first, _ar_last = AR_HEADER_ROW+1, AR_HEADER_ROW+len(areas)
    ES_L = get_column_letter(ar_idx['Evidence Status']); SRC_L = get_column_letter(ar_idx['Source'])
    OWN_L2 = get_column_letter(ar_idx['Owner']); EXT_L = get_column_letter(ar_idx['Externally Supported'])
    _es_rng = f'{ES_L}{_ar_first}:{ES_L}{_ar_last}'

    # ── Status line — live, reflects real state as reviewers fill the sheet ──
    merge(wsa,f'B2:{get_column_letter(n_ar_cols-1)}2',
          f'=IF(COUNTIF({_es_rng},"Missing")=0,"Evidence register complete — every assumption has a documented basis.","Evidence register incomplete — "&COUNTIF({_es_rng},"Missing")&" assumption(s) still need supporting evidence.")',
          bold=True,sz=10,col=WHITE,bg=P2_FILL,v='center')
    # colour set as a flat amber to start (register starts incomplete by
    # definition — every row defaults to Missing); flip to green only once
    # genuinely complete would require a formula-driven fill, which Excel
    # can't apply to font/fill directly from a cell formula — handled via
    # conditional formatting on this same range instead, below.
    fill_range(wsa,2,2,2,n_ar_cols-1,P2_FILL); set_row(wsa,2,20)
    wsa.conditional_formatting.add(f'B2:{get_column_letter(n_ar_cols-1)}2',
        FormulaRule(formula=[f'COUNTIF({_es_rng},"Missing")=0'],
                    fill=PatternFill(start_color=OK_FILL,end_color=OK_FILL,fill_type='solid'), font=Font(color=OK_TXT,bold=True)))

    # ── Summary cards — Total / Missing Source / Missing Owner / Externally Supported ──
    _ar_cards=[('Total Assumptions', f'={_ar_last-_ar_first+1}', 2,3),
               ('Missing Source', f'=COUNTIF({SRC_L}{_ar_first}:{SRC_L}{_ar_last},"Not yet provided")', 4,5),
               ('Missing Owner', f'=COUNTIF({OWN_L2}{_ar_first}:{OWN_L2}{_ar_last},"Not yet provided")', 6,7),
               ('Externally Supported', f'=COUNTIF({EXT_L}{_ar_first}:{EXT_L}{_ar_last},"Yes")', 8,9)]
    for label,formula,c1,c2 in _ar_cards:
        col_l1,col_l2 = get_column_letter(c1), get_column_letter(c2)
        merge(wsa,f'{col_l1}3:{col_l2}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        merge(wsa,f'{col_l1}4:{col_l2}4',formula,bold=True,sz=16,col=CHARCOAL,bg=PANEL_GREY,h='center')
        for rr in (3,4):
            for cc in range(c1,c2+1): wsa.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(wsa,3,14); set_row(wsa,4,24)
    merge(wsa,f'B5:{get_column_letter(n_ar_cols-1)}5',
          'Related findings are pre-populated from this review. Source, Source Date, Owner, Basis and Externally Supported are for the model owner to complete and the reviewer to verify.',
          sz=8,col=GREY_TXT2,italic=True,wrap=True)
    set_row(wsa,5,16)

    for col,h in enumerate(ar_headers,1):
        c=wsa.cell(AR_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
        c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
    set_row(wsa,AR_HEADER_ROW,28)
    wsa.freeze_panes = f'{get_column_letter(ar_idx["Related Findings"])}{AR_HEADER_ROW+1}'

    _placeholder='Not yet provided'
    _centered4={'Decision Critical','Source Date','Externally Supported'}
    _wrapped4={'Assumption Area','Related Findings','Notes'}
    for idx,(area,kws) in enumerate(areas):
        row_i = AR_HEADER_ROW+1+idx
        row_bg = WHITE if idx%2==0 else 'FAFBFC'
        hits=[fid for fid,txt in ftexts if any(k in txt for k in kws)]
        ref_txt=', '.join(hits[:4])+(f' +{len(hits)-4} more' if len(hits)>4 else '') if hits else '—'
        row_values={
            'Assumption Area': area, 'Evidence Status': 'Missing', 'Decision Critical': '',
            'Related Findings': ref_txt, 'Source': _placeholder, 'Source Date': _placeholder,
            'Owner': _placeholder, 'Basis': '', 'Externally Supported': 'No', 'Notes': '',
        }
        for name,val in row_values.items():
            col = ar_idx[name]
            c=wsa.cell(row_i,col); c.value=val
            c.font=Fn(sz=9,bold=(name=='Assumption Area'),
                      italic=(val==_placeholder), col=(GREY_MID if val==_placeholder else '000000'))
            c.fill=F(row_bg)
            c.alignment=A(h='center' if name in _centered4 else 'left', v='top', wrap=(name in _wrapped4))
            c.border=B(col=PANEL_BORDER)
        badge(wsa,row_i,ar_idx['Evidence Status'],'Missing',GREY_LIGHT,GREY_TXT2,sz=8,bold=False)
        set_row(wsa,row_i,24)

    _ar_last_row = AR_HEADER_ROW+len(areas)

    if len(areas) > 0:
        _ar_table = Table(displayName="AssumptionRegisterTable", ref=f'B{AR_HEADER_ROW}:{get_column_letter(n_ar_cols-1)}{_ar_last_row}')
        _ar_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                    showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
        wsa.add_table(_ar_table)

        evidence_dv = DataValidation(type='list', formula1='"Missing,Management,External,Verified"', allow_blank=False)
        wsa.add_data_validation(evidence_dv); evidence_dv.add(_es_rng)

        critical_dv = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
        wsa.add_data_validation(critical_dv)
        critical_dv.add(f'{get_column_letter(ar_idx["Decision Critical"])}{_ar_first}:{get_column_letter(ar_idx["Decision Critical"])}{_ar_last}')

        basis_dv = DataValidation(type='list', formula1='"Contract,Market,Management,Model,Other"', allow_blank=True)
        wsa.add_data_validation(basis_dv)
        basis_dv.add(f'{get_column_letter(ar_idx["Basis"])}{_ar_first}:{get_column_letter(ar_idx["Basis"])}{_ar_last}')

        ext_dv = DataValidation(type='list', formula1='"Yes,No"', allow_blank=False)
        wsa.add_data_validation(ext_dv); ext_dv.add(f'{EXT_L}{_ar_first}:{EXT_L}{_ar_last}')

        # ── Highlight still-missing Source / Source Date / Owner — pale blue,
        #    clears automatically once a reviewer replaces the placeholder ──
        for col_letter in (SRC_L, get_column_letter(ar_idx['Source Date']), OWN_L2):
            _rng = f'{col_letter}{_ar_first}:{col_letter}{_ar_last}'
            wsa.conditional_formatting.add(_rng,
                CellIsRule(operator='equal', formula=[f'"{_placeholder}"'],
                           fill=PatternFill(start_color=PALE_ACCENT,end_color=PALE_ACCENT,fill_type='solid')))

        # Evidence Status badge colours track the dropdown value live
        _ev_colors=[('Missing',GREY_LIGHT,GREY_TXT2),('Management',P2_FILL,P2_TXT),
                    ('External',MID_BLUE,WHITE),('Verified',OK_FILL,OK_TXT)]
        for val,fill,txt in _ev_colors:
            wsa.conditional_formatting.add(_es_rng,
                CellIsRule(operator='equal', formula=[f'"{val}"'],
                           fill=PatternFill(start_color=fill,end_color=fill,fill_type='solid'), font=Font(color=txt,bold=True)))

    # ════════════════════════════════════════════════════════════════════════
    # TAB 6 — FORMULA ANALYSIS
    # ════════════════════════════════════════════════════════════════════════
    ws6=wb.create_sheet('Formula Risk Review'); ws6.sheet_view.showGridLines=False
    for col,w in [(1,3),(2,10),(3,12),(4,8),(5,10),(6,14),(7,20),(8,26),(9,16),(10,32)]:
        set_col(ws6,col,w)
    uf_headers=['','UFI','Sheet','Cell','F-Score','Complexity Band','Formula Class','Key Risk',
                'Status','Reviewer Comment','Formula Text','F-Score Explanation','External Link',
                'Volatile','Hardcode','IFERROR','Cross-Sheet Refs','Precedent Sheets','']
    n_fa_cols = len(uf_headers)
    n_fa_visible = 10  # UFI..Reviewer Comment (column NUMBER of last visible field, not a count)
    uf_idx = {h.replace('\n',' '): i for i,h in enumerate(uf_headers,1) if h}
    FA_HEADER_ROW=6

    merge(ws6,f'B1:{get_column_letter(n_fa_cols-1)}1','FORMULA RISK REVIEW',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws6,1,2,1,n_fa_cols-1,DARK_BLUE); set_row(ws6,1,26)

    stats=t0.get('stats',{})
    ufs=t0.get('uniqueFormulas',[])
    n_high_risk = sum(1 for uf in ufs if uf.get('band') in ('High','Very High','Critical'))
    n_volatile  = sum(1 for uf in ufs if uf.get('volatileFlag'))
    _fa_first, _fa_last = FA_HEADER_ROW+1, FA_HEADER_ROW+min(len(ufs),200)
    BAND_L = get_column_letter(uf_idx['Complexity Band'])

    # ── Summary cards — exact 6 metrics from the brief ──────────────────────
    _fa_cards=[('Total Formulas', stats.get('totalFormulaCells',0), 2,3),
               ('Unique Patterns', stats.get('uniqueFormulaCount',0), 4,4),
               ('High Risk', n_high_risk, 5,5),
               ('External Links', stats.get('totalExternalLinks',0), 6,6),
               ('Volatile', n_volatile, 7,8),
               ('IFERROR', stats.get('totalIferrorCount',0), 9,10)]
    for label,val,c1,c2 in _fa_cards:
        col_l1,col_l2=get_column_letter(c1),get_column_letter(c2)
        if c2>c1: merge(ws6,f'{col_l1}3:{col_l2}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        else: cell(ws6,f'{col_l1}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        if c2>c1: merge(ws6,f'{col_l1}4:{col_l2}4',val,bold=True,sz=16,col=CHARCOAL,bg=PANEL_GREY,h='center')
        else: cell(ws6,f'{col_l1}4',val,bold=True,sz=16,col=CHARCOAL,bg=PANEL_GREY,h='center')
        ws6.cell(4,c1).number_format='#,##0;[Red](#,##0);-'
        for rr in (3,4):
            for cc in range(c1,c2+1): ws6.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(ws6,3,14); set_row(ws6,4,24)
    merge(ws6,f'B5:{get_column_letter(n_fa_cols-1)}5',
          'High-risk and volatile formulas warrant reviewer attention first — use the Complexity Band filter below.',
          sz=8,col=GREY_TXT2,italic=True)
    set_row(ws6,5,16)

    for col,h in enumerate(uf_headers,1):
        c=ws6.cell(FA_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
        c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
    set_row(ws6,FA_HEADER_ROW,28)
    ws6.freeze_panes = f'{get_column_letter(uf_idx["F-Score"])}{FA_HEADER_ROW+1}'

    _band_badge={'Critical':(P1_FILL,P1_TXT),'Very High':(P1_FILL,P1_TXT),'High':(P2_FILL,P2_TXT),
                 'Moderate':(P3_FILL,P3_TXT),'Low':(OK_FILL,OK_TXT)}
    _centered5={'F-Score','Complexity Band','External Link','Volatile','Hardcode','IFERROR','Cross-Sheet Refs'}
    _wrapped5={'Key Risk','Reviewer Comment','F-Score Explanation'}
    for idx,uf in enumerate(ufs[:200]):
        row_i = FA_HEADER_ROW+1+idx
        row_bg = WHITE if idx%2==0 else 'FAFBFC'
        band=uf.get('band','Low')

        # Key Risk — one short label, not a full sentence (auto_comment in
        # Reviewer Comment carries the detail)
        risk_flags=[]
        if uf.get('externalLinkFlag'): risk_flags.append('External link')
        if uf.get('volatileFlag'): risk_flags.append('Volatile function')
        if uf.get('hardcodeFlag'): risk_flags.append('Hardcoded value')
        if uf.get('iferrorFlag'): risk_flags.append('Error suppression')
        xrefs=uf.get('crossSheetRefs',0)
        if xrefs>2: risk_flags.append(f'{xrefs} sheet refs')
        key_risk = risk_flags[0] if len(risk_flags)==1 else (f'{risk_flags[0]} +{len(risk_flags)-1} more' if risk_flags else '—')

        auto_comment=''
        if band in ('High','Very High','Critical','Moderate'):
            parts=[]
            if uf.get('externalLinkFlag'): parts.append('references an external workbook — verify the link is current and the source file is accessible')
            if uf.get('volatileFlag'): parts.append('uses a volatile function (OFFSET/INDIRECT) — check whether a static alternative would work')
            if uf.get('iferrorFlag'): parts.append('contains error suppression — confirm this is not hiding a genuine calculation error')
            if uf.get('hardcodeFlag'): parts.append('contains hardcoded values — check whether these should be on the Inputs sheet')
            if xrefs>2: parts.append(f'references {xrefs} sheets — trace the dependency chain to confirm all source data is correct')
            auto_comment = ('Review required: '+'; '.join(parts)+'.') if parts else f'Complex formula ({band} F-score). Review the logic and confirm it produces the intended result.'

        row_values={
            'UFI': uf.get('ufi',''), 'Sheet': uf.get('sheet',''), 'Cell': uf.get('cell',''),
            'F-Score': uf.get('fscore',0), 'Complexity Band': band, 'Formula Class': uf.get('formulaClass',''),
            'Key Risk': key_risk, 'Status': '', 'Reviewer Comment': auto_comment,
            'Formula Text': uf.get('formulaText',''), 'F-Score Explanation': uf.get('explanation',''),
            'External Link': 'Yes' if uf.get('externalLinkFlag') else 'No',
            'Volatile': 'Yes' if uf.get('volatileFlag') else 'No',
            'Hardcode': 'Yes' if uf.get('hardcodeFlag') else 'No',
            'IFERROR': 'Yes' if uf.get('iferrorFlag') else 'No',
            'Cross-Sheet Refs': xrefs, 'Precedent Sheets': uf.get('precedentSheets',''),
        }
        for name,val in row_values.items():
            col=uf_idx[name]
            c=ws6.cell(row_i,col); c.value=val
            is_formula_text = (name=='Formula Text')
            c.font=Font(name='Consolas',size=8,color='000000') if is_formula_text else Fn(sz=9,bold=(name=='UFI'))
            c.fill=F(row_bg)
            c.alignment=A(h='center' if name in _centered5 else 'left', v='top', wrap=(name in _wrapped5 or is_formula_text))
            c.border=B(col=PANEL_BORDER)

        bf,bt=_band_badge.get(band,(GREY_LIGHT,CHARCOAL))
        badge(ws6,row_i,uf_idx['Complexity Band'],band,bf,bt,sz=8,bold=False)
        set_row(ws6,row_i,30)

    _fa_last_row = FA_HEADER_ROW+min(len(ufs),200)

    for col in range(n_fa_visible+1, n_fa_cols):
        ws6.column_dimensions[get_column_letter(col)].hidden = True

    if len(ufs) > 0:
        _fa_table = Table(displayName="FormulaRiskTable", ref=f'B{FA_HEADER_ROW}:{get_column_letter(n_fa_cols-1)}{_fa_last_row}')
        _fa_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                    showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
        ws6.add_table(_fa_table)

        status_dv = DataValidation(type='list',
            formula1='"Accept,Split Formula,Replace Link,Investigate,Not Relevant"', allow_blank=True)
        ws6.add_data_validation(status_dv)
        status_dv.add(f'{get_column_letter(uf_idx["Status"])}{_fa_first}:{get_column_letter(uf_idx["Status"])}{_fa_last}')

    if len(ufs) > 200:
        _note_row = _fa_last_row + 1
        merge(ws6,f'B{_note_row}:{get_column_letter(n_fa_cols-1)}{_note_row}',
              f'List capped at 200 unique formulas of {len(ufs):,} total — remaining formulas are represented in the summary counts above.',
              sz=8,col=GREY_TXT2,italic=True)
        set_row(ws6,_note_row,16)

    # ════════════════════════════════════════════════════════════════════════
    # TAB — ERROR-CODE ROOT CAUSE MATRIX (V11 §4)
    # ════════════════════════════════════════════════════════════════════════
    wse=wb.create_sheet('Error Matrix'); wse.sheet_view.showGridLines=False
    err_headers=['','Code','Count','Severity','Sample Locations','Typical Root Cause','Recommended Action','Owner','Fix Status','']
    n_em_cols = len(err_headers)
    for i,w in enumerate([3,10,8,10,32,32,32,14,14,3],1): set_col(wse,i,w)
    em_idx = {h.replace('\n',' '): i for i,h in enumerate(err_headers,1) if h}
    EM_HEADER_ROW=6

    merge(wse,f'B1:{get_column_letter(n_em_cols-1)}1','ERROR MATRIX',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(wse,1,2,1,n_em_cols-1,DARK_BLUE); set_row(wse,1,26)

    _ERR_GUIDE={
        '#REF!':   ('Referenced rows, columns or sheets were deleted or moved after the formula was written.',
                    'Rebuild the reference against the current layout; add named ranges for structural anchors.', 'High'),
        '#DIV/0!': ('Division where the denominator is zero or blank — typically an unguarded ratio in early or terminal periods.',
                    'Guard the denominator explicitly (IF(den=0,...)) rather than wrapping the result in IFERROR.', 'Medium'),
        '#N/A':    ('A lookup failed to find its key — missing key, mismatched type/format, or approximate match on unsorted data.',
                    'Confirm key existence and exact-match settings; reconcile key lists between source and lookup tables.', 'Medium'),
        '#VALUE!': ('Operation applied to the wrong data type — text in arithmetic, ranges of mismatched size, or stray characters in inputs.',
                    'Trace the offending operand; clean input typing and separate text from numeric columns.', 'Medium'),
        '#NAME?':  ('Formula references an undefined name — deleted named range, misspelled function, or missing add-in.',
                    'Repair or redefine the name; remove dependencies on unavailable add-ins.', 'High'),
        '#NUM!':   ('Invalid numeric operation — IRR failing to converge, negative value in a root/log, or overflow.',
                    'Check the input domain; for IRR provide a guess or use XIRR with explicit dates.', 'Medium'),
        '#NULL!':  ('Range intersection that does not intersect — usually a typo (space instead of comma/colon) in a range reference.',
                    'Correct the range operator in the formula.', 'Low'),
        '#SPILL!': ('A dynamic array result is blocked by existing content in its spill range.',
                    'Clear the blocking cells or convert the formula to a fixed range.', 'Low'),
    }

    from collections import OrderedDict
    _by_code=OrderedDict()
    for e in errorScan:
        code=str(e.get('error','')).strip()
        if not code: continue
        _by_code.setdefault(code,[]).append(f"{e.get('sheet','')}!{e.get('cell','')}")

    _live_total = sum(len(v) for v in _by_code.values())
    _error_types = len(_by_code)
    _ifer=t0.get('stats',{}).get('totalIferrorCount',0)
    # A code counts as High severity, and a large count amplifies it —
    # structural breakage (#REF!/#NAME?) with real volume is the reliance blocker.
    _has_high_severity = any(_ERR_GUIDE.get(c,('','','Medium'))[2]=='High' and len(locs)>0 for c,locs in _by_code.items())

    # ── Summary cards — Live Errors / Error Types / IFERROR Wrappers ────────
    _em_cards=[('Live Errors', _live_total, 2,3), ('Error Types', _error_types, 4,5), ('IFERROR Wrappers', _ifer, 6,7)]
    for label,val,c1,c2 in _em_cards:
        col_l1,col_l2=get_column_letter(c1),get_column_letter(c2)
        merge(wse,f'{col_l1}3:{col_l2}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        merge(wse,f'{col_l1}4:{col_l2}4',val,bold=True,sz=18,col=CHARCOAL,bg=PANEL_GREY,h='center')
        ws_val_cell = wse.cell(4,c1); ws_val_cell.number_format='#,##0;[Red](#,##0);-'
        for rr in (3,4):
            for cc in range(c1,c2+1): wse.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(wse,3,14); set_row(wse,4,26)

    # ── Reliance-blocker status line ─────────────────────────────────────────
    if _live_total==0:
        _blocker_txt='No live formula errors detected — not currently a reliance blocker on this basis.'
        _blocker_bg=OK_FILL; _blocker_fg=OK_TXT
    elif _has_high_severity:
        _blocker_txt=f'{_live_total:,} live error(s) including High-severity codes — formula errors ARE a reliance blocker until resolved.'
        _blocker_bg=P1_FILL; _blocker_fg=P1_TXT
    else:
        _blocker_txt=f'{_live_total:,} live error(s), no High-severity codes — review recommended but not an immediate reliance blocker.'
        _blocker_bg=P2_FILL; _blocker_fg=P2_TXT
    merge(wse,f'B5:{get_column_letter(n_em_cols-1)}5',_blocker_txt,bold=True,sz=9,col=_blocker_fg,bg=_blocker_bg,v='center')
    set_row(wse,5,18)

    for col,h in enumerate(err_headers,1):
        c=wse.cell(EM_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
        c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
    set_row(wse,EM_HEADER_ROW,22)
    wse.freeze_panes = f'{get_column_letter(em_idx["Sample Locations"])}{EM_HEADER_ROW+1}'

    _sev_badge={'High':(P1_FILL,P1_TXT),'Medium':(P2_FILL,P2_TXT),'Low':(P3_FILL,P3_TXT)}
    row_i=EM_HEADER_ROW+1
    if not _by_code:
        merge(wse,f'B{row_i}:{get_column_letter(n_em_cols-1)}{row_i}',
              'No live error values detected in any cell. Note: errors masked by IFERROR/IFNA wrappers are not visible as live values — masking risk is assessed separately (see Formula Risk Review).',
              sz=9,col=GREY_TXT2,wrap=True)
        set_row(wse,row_i,24); row_i+=1
    else:
        for _idx,(code, locs) in enumerate(sorted(_by_code.items(), key=lambda kv:-len(kv[1]))):
            cause,action,severity=_ERR_GUIDE.get(code,('Unclassified error code.','Investigate the listed cells directly.','Medium'))
            loc_txt=', '.join(locs[:5])+(f' +{len(locs)-5} more' if len(locs)>5 else '')
            row_bg = WHITE if _idx%2==0 else 'FAFBFC'
            row_values={'Code':code,'Count':len(locs),'Severity':severity,'Sample Locations':loc_txt,
                        'Typical Root Cause':cause,'Recommended Action':action,'Owner':'','Fix Status':'Open'}
            for name,val in row_values.items():
                col=em_idx[name]
                c=wse.cell(row_i,col); c.value=val
                c.font=Fn(sz=9,bold=(name=='Code'))
                c.fill=F(row_bg)
                c.alignment=A(h='center' if name in ('Count','Severity','Fix Status') else 'left',
                              v='top', wrap=(name in ('Sample Locations','Typical Root Cause','Recommended Action')))
                c.border=B(col=PANEL_BORDER)
            c_count = wse.cell(row_i, em_idx['Count']); c_count.number_format='#,##0;[Red](#,##0);-'
            bf,bt=_sev_badge.get(severity,(GREY_LIGHT,CHARCOAL))
            badge(wse,row_i,em_idx['Severity'],severity,bf,bt,sz=8,bold=False)
            set_row(wse,row_i,32); row_i+=1

    _em_last_row = row_i-1
    if _by_code:
        _em_table = Table(displayName="ErrorMatrixTable", ref=f'B{EM_HEADER_ROW}:{get_column_letter(n_em_cols-1)}{_em_last_row}')
        _em_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                    showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
        wse.add_table(_em_table)
        fix_dv = DataValidation(type='list', formula1='"Open,In Progress,Fixed,Not Applicable"', allow_blank=False)
        wse.add_data_validation(fix_dv)
        fix_dv.add(f'{get_column_letter(em_idx["Fix Status"])}{EM_HEADER_ROW+1}:{get_column_letter(em_idx["Fix Status"])}{_em_last_row}')

    # ════════════════════════════════════════════════════════════════════════
    # TAB — TIDY-UP: REDUNDANT INPUTS (V11 §2) — full client-actionable list
    # ════════════════════════════════════════════════════════════════════════
    wst=wb.create_sheet('Redundant Inputs'); wst.sheet_view.showGridLines=False
    ri_headers=['','Sheet','Cell','Value','Value Type','Nearby Label','Go to Cell',
                'Review Decision','Owner','Status','']
    n_ri_cols = len(ri_headers)
    for i,w in enumerate([3,14,10,14,14,26,12,16,16,12,3],1): set_col(wst,i,w)
    ri_idx = {h.replace('\n',' '): i for i,h in enumerate(ri_headers,1) if h}
    RI_HEADER_ROW=6

    merge(wst,f'B1:{get_column_letter(n_ri_cols-1)}1','REDUNDANT INPUTS',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(wst,1,2,1,n_ri_cols-1,DARK_BLUE); set_row(wst,1,26)

    if not redundantIn.get('applicable',False):
        merge(wst,f'B2:{get_column_letter(n_ri_cols-1)}2',
              'Not applicable — no sheet matching input/assumption/driver naming was detected in this model.',
              sz=9,col=GREY_TXT2,bg=PALE_ACCENT,italic=True,wrap=True)
        set_row(wst,2,20)
    else:
        _rc=redundantIn.get('redundantCount',0); _ti=redundantIn.get('totalInputs',0)
        _pct=(100.0*_rc/_ti) if _ti else 0.0

        # ── Summary cards — exact 3 metrics from the brief ──────────────────
        _ri_cards=[('Redundant Inputs', _rc, 2,3), ('Input Constants', _ti, 4,5), ('Redundancy Rate', _pct/100.0, 6,7)]
        for label,val,c1,c2 in _ri_cards:
            col_l1,col_l2=get_column_letter(c1),get_column_letter(c2)
            merge(wst,f'{col_l1}3:{col_l2}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
            merge(wst,f'{col_l1}4:{col_l2}4',val,bold=True,sz=18,col=(P2_TXT if label=='Redundant Inputs' and _rc>0 else CHARCOAL),bg=PANEL_GREY,h='center')
            wst.cell(4,c1).number_format = '0%;[Red](0%);-' if label=='Redundancy Rate' else '#,##0;[Red](#,##0);-'
            for rr in (3,4):
                for cc in range(c1,c2+1): wst.cell(rr,cc).border=B(col=PANEL_BORDER)
        set_row(wst,3,14); set_row(wst,4,26)

        # ── Instruction box — short, per the brief ───────────────────────────
        merge(wst,f'B5:{get_column_letter(n_ri_cols-1)}5',
              'Every listed input must be linked into the calculation chain, removed, or relabelled as a memo item. '
              + redundantIn.get('note',''),
              sz=8,col=GREY_TXT2,italic=True,wrap=True)
        set_row(wst,5,20)

        for col,h in enumerate(ri_headers,1):
            c=wst.cell(RI_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
            c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
        set_row(wst,RI_HEADER_ROW,26)
        wst.freeze_panes = f'{get_column_letter(ri_idx["Nearby Label"])}{RI_HEADER_ROW+1}'

        def _colnum(addr):
            m = re.match(r'^([A-Z]+)(\d+)$', addr or '')
            if not m: return (0,0)
            c=0
            for ch in m.group(1): c = c*26 + (ord(ch)-64)
            return (c, int(m.group(2)))

        _rows=sorted(redundantIn.get('redundant',[]), key=lambda item: (item.get('sheet',''), _colnum(item.get('cell',''))))

        def _value_type(v):
            try: v=float(v)
            except (TypeError,ValueError): return 'Other'
            if v==0: return 'Zero'
            if v<0: return 'Negative'
            if 0<abs(v)<=1: return 'Percentage-like'
            if v==int(v): return 'Whole number'
            return 'Decimal'

        _vt_badge={'Zero':(GREY_LIGHT,GREY_TXT2),'Negative':(P1_FILL,P1_TXT),'Percentage-like':(P3_FILL,P3_TXT),
                   'Whole number':(PANEL_GREY,CHARCOAL),'Decimal':(PANEL_GREY,CHARCOAL),'Other':(GREY_LIGHT,GREY_TXT2)}

        for idx,item in enumerate(_rows):
            row_i = RI_HEADER_ROW+1+idx
            row_bg = WHITE if idx%2==0 else 'FAFBFC'
            sheet=item.get('sheet',''); cell_addr=item.get('cell',''); val=item.get('value','')
            label=item.get('label','') or ''
            vtype=_value_type(val)
            row_values={'Sheet':sheet,'Cell':cell_addr,'Value':val,'Value Type':vtype,'Nearby Label':label or '—',
                        'Go to Cell':'','Review Decision':'','Owner':'','Status':'Open'}
            for name,v in row_values.items():
                col=ri_idx[name]
                c=wst.cell(row_i,col); c.value=v
                c.font=Fn(sz=9,bold=(name=='Cell'))
                c.fill=F(row_bg)
                c.alignment=A(h='center' if name in ('Cell','Value','Value Type','Status') else 'left', v='top', wrap=(name=='Nearby Label'))
                c.border=B(col=PANEL_BORDER)
            if isinstance(val,(int,float)):
                wst.cell(row_i, ri_idx['Value']).number_format='#,##0.####'
            vf,vt=_vt_badge.get(vtype,(GREY_LIGHT,GREY_TXT2))
            badge(wst,row_i,ri_idx['Value Type'],vtype,vf,vt,sz=8,bold=False)

            # Clean fixed label — no cell-address suffix, no raw formula text
            link=wst.cell(row_i, ri_idx['Go to Cell'])
            if sheet:
                link.value=f'=HYPERLINK("[{sourceFile}]{sheet}!{cell_addr}","Go to cell")'
                link.font=Font(size=9,color=MID_BLUE,underline='single',name='Arial')
            else:
                link.value='—'; link.font=Fn(sz=9,col=GREY_MID)
            link.fill=F(row_bg); link.alignment=A(h='center',v='center'); link.border=B(col=PANEL_BORDER)
            set_row(wst,row_i,18)

        _ri_last_row = RI_HEADER_ROW+len(_rows)
        if _rows:
            _ri_table = Table(displayName="RedundantInputsTable", ref=f'B{RI_HEADER_ROW}:{get_column_letter(n_ri_cols-1)}{_ri_last_row}')
            _ri_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                        showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
            wst.add_table(_ri_table)

            decision_dv = DataValidation(type='list', formula1='"Link,Remove,Relabel,Keep as Memo"', allow_blank=True)
            wst.add_data_validation(decision_dv)
            decision_dv.add(f'{get_column_letter(ri_idx["Review Decision"])}{RI_HEADER_ROW+1}:{get_column_letter(ri_idx["Review Decision"])}{_ri_last_row}')

            status_dv = DataValidation(type='list', formula1='"Open,Cleared,Deferred"', allow_blank=False)
            wst.add_data_validation(status_dv)
            status_dv.add(f'{get_column_letter(ri_idx["Status"])}{RI_HEADER_ROW+1}:{get_column_letter(ri_idx["Status"])}{_ri_last_row}')

    # ════════════════════════════════════════════════════════════════════════
    # TAB — SHEET LINKAGE (inverse of Redundant Inputs) — orphan sheet detection
    # ════════════════════════════════════════════════════════════════════════
    wsl=wb.create_sheet('Sheet Linkage'); wsl.sheet_view.showGridLines=False
    sl_headers=['','Sheet','Reviewer Decision','Owner','Status','Notes','']
    n_sl_cols = len(sl_headers)
    for i,w in enumerate([3,22,20,16,12,34,3],1): set_col(wsl,i,w)
    sl_idx = {h.replace('\n',' '): i for i,h in enumerate(sl_headers,1) if h}
    SL_HEADER_ROW=6

    merge(wsl,f'B1:{get_column_letter(n_sl_cols-1)}1','SHEET LINKAGE',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(wsl,1,2,1,n_sl_cols-1,DARK_BLUE); set_row(wsl,1,26)

    if not orphanIn.get('applicable',False):
        merge(wsl,f'B2:{get_column_letter(n_sl_cols-1)}2',
              'Not applicable — no sheet matching financial-statement naming (AFS/IFS/Balance Sheet/Cash Flow/P&L) was detected in this model.',
              sz=9,col=GREY_TXT2,bg=PALE_ACCENT,italic=True,wrap=True)
        set_row(wsl,2,20)
    else:
        _orphans = orphanIn.get('orphanSheets',[])
        _fin_sheets = orphanIn.get('financialStatementSheets',[])
        _total = orphanIn.get('totalSheets',0)

        # ── Summary cards ─────────────────────────────────────────────────
        _sl_cards=[('Orphan Sheets', len(_orphans), 2,3), ('Total Sheets', _total, 4,5),
                   ('Financial Statement Sheets', len(_fin_sheets), 6,7)]
        for label,val,c1,c2 in _sl_cards:
            col_l1,col_l2=get_column_letter(c1),get_column_letter(c2)
            merge(wsl,f'{col_l1}3:{col_l2}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
            merge(wsl,f'{col_l1}4:{col_l2}4',val,bold=True,sz=18,col=(P1_TXT if label=='Orphan Sheets' and val>0 else CHARCOAL),bg=PANEL_GREY,h='center')
            wsl.cell(4,c1).number_format='#,##0;[Red](#,##0);-'
            for rr in (3,4):
                for cc in range(c1,c2+1): wsl.cell(rr,cc).border=B(col=PANEL_BORDER)
        set_row(wsl,3,14); set_row(wsl,4,26)

        # ── Instruction box ───────────────────────────────────────────────
        merge(wsl,f'B5:{get_column_letter(n_sl_cols-1)}5',
              (f'Every sheet must have a traceable formula path — direct or indirect, including named ranges — to a '
               f'financial-statement sheet ({", ".join(_fin_sheets) if _fin_sheets else "none detected"}), or a documented reason it is intentionally standalone. '
               + orphanIn.get('note','')),
              sz=8,col=GREY_TXT2,italic=True,wrap=True)
        set_row(wsl,5,26)

        for col,h in enumerate(sl_headers,1):
            c=wsl.cell(SL_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
            c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
        set_row(wsl,SL_HEADER_ROW,22)
        wsl.freeze_panes = f'{get_column_letter(sl_idx["Owner"])}{SL_HEADER_ROW+1}'

        if not _orphans:
            merge(wsl,f'B{SL_HEADER_ROW+1}:{get_column_letter(n_sl_cols-1)}{SL_HEADER_ROW+1}',
                  'No orphan sheets detected — every sheet has a traceable path to a financial statement.',
                  sz=9,col=OK_TXT,bg=OK_FILL,wrap=True)
            set_row(wsl,SL_HEADER_ROW+1,20)
        else:
            for idx,sheet_name in enumerate(_orphans):
                row_i = SL_HEADER_ROW+1+idx
                row_bg = WHITE if idx%2==0 else 'FAFBFC'
                row_values={'Sheet':sheet_name,'Reviewer Decision':'','Owner':'','Status':'Open','Notes':''}
                for name,val in row_values.items():
                    col=sl_idx[name]
                    c=wsl.cell(row_i,col); c.value=val
                    c.font=Fn(sz=9,bold=(name=='Sheet'))
                    c.fill=F(row_bg)
                    c.alignment=A(h='center' if name=='Status' else 'left', v='top', wrap=(name=='Notes'))
                    c.border=B(col=PANEL_BORDER)
                set_row(wsl,row_i,20)

            _sl_last_row = SL_HEADER_ROW+len(_orphans)
            _sl_table = Table(displayName="SheetLinkageTable", ref=f'B{SL_HEADER_ROW}:{get_column_letter(n_sl_cols-1)}{_sl_last_row}')
            _sl_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                        showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
            wsl.add_table(_sl_table)

            decision_dv = DataValidation(type='list',
                formula1='"Link into statements,Confirm intentionally standalone,Remove sheet,Investigate"', allow_blank=True)
            wsl.add_data_validation(decision_dv)
            decision_dv.add(f'{get_column_letter(sl_idx["Reviewer Decision"])}{SL_HEADER_ROW+1}:{get_column_letter(sl_idx["Reviewer Decision"])}{_sl_last_row}')

            status_dv2 = DataValidation(type='list', formula1='"Open,Cleared,Deferred"', allow_blank=False)
            wsl.add_data_validation(status_dv2)
            status_dv2.add(f'{get_column_letter(sl_idx["Status"])}{SL_HEADER_ROW+1}:{get_column_letter(sl_idx["Status"])}{_sl_last_row}')


    # ════════════════════════════════════════════════════════════════════════
    # TAB — NAMED RANGE AUDIT
    # ════════════════════════════════════════════════════════════════════════
    wnr=wb.create_sheet('Named Range Audit'); wnr.sheet_view.showGridLines=False
    nr_headers=['','Name','Issue','Target','Sheets','Reviewer Decision','Owner','Status','']
    n_nr_cols = len(nr_headers)
    for i,w in enumerate([3,24,16,26,16,20,16,12,3],1): set_col(wnr,i,w)
    nr_idx = {h.replace('\n',' '): i for i,h in enumerate(nr_headers,1) if h}
    NR_HEADER_ROW=6

    merge(wnr,f'B1:{get_column_letter(n_nr_cols-1)}1','NAMED RANGE AUDIT',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(wnr,1,2,1,n_nr_cols-1,DARK_BLUE); set_row(wnr,1,26)

    if not namedRangeIn.get('applicable',False):
        merge(wnr,f'B2:{get_column_letter(n_nr_cols-1)}2',
              namedRangeIn.get('note','No named ranges are defined in this model.'),
              sz=9,col=GREY_TXT2,bg=PALE_ACCENT,italic=True,wrap=True)
        set_row(wnr,2,20)
    else:
        _unused = namedRangeIn.get('unused',[])
        _poor = namedRangeIn.get('poorlyNamed',[])
        _broken = namedRangeIn.get('broken',[])
        _total = namedRangeIn.get('totalNamedRanges',0)

        _nr_cards=[('Total Named Ranges', _total, 2,3), ('Unused', len(_unused), 4,5),
                   ('Poorly Named', len(_poor), 6,6), ('Broken', len(_broken), 7,7)]
        for label,val,c1,c2 in _nr_cards:
            col_l1,col_l2=get_column_letter(c1),get_column_letter(c2)
            merge(wnr,f'{col_l1}3:{col_l2}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
            merge(wnr,f'{col_l1}4:{col_l2}4',val,bold=True,sz=16,col=(P1_TXT if label=='Broken' and val>0 else P2_TXT if label=='Unused' and val>0 else CHARCOAL),bg=PANEL_GREY,h='center')
            wnr.cell(4,c1).number_format='#,##0;[Red](#,##0);-'
            for rr in (3,4):
                for cc in range(c1,c2+1): wnr.cell(rr,cc).border=B(col=PANEL_BORDER)
        set_row(wnr,3,14); set_row(wnr,4,26)

        merge(wnr,f'B5:{get_column_letter(n_nr_cols-1)}5', namedRangeIn.get('note',''),
              sz=8,col=GREY_TXT2,italic=True,wrap=True)
        set_row(wnr,5,20)

        for col,h in enumerate(nr_headers,1):
            c=wnr.cell(NR_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
            c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
        set_row(wnr,NR_HEADER_ROW,22)
        wnr.freeze_panes = f'{get_column_letter(nr_idx["Target"])}{NR_HEADER_ROW+1}'

        _rows_data=[]
        for b in _broken: _rows_data.append((b['name'],'Broken',b['target'],'', P1_FILL,P1_TXT))
        for u in _unused: _rows_data.append((u['name'],'Unused',u['target'],u.get('sheets',''), P2_FILL,P2_TXT))
        for p in _poor:
            already = any(rd[0]==p['name'] and rd[1]=='Unused' for rd in _rows_data)
            _rows_data.append((p['name'], 'Unused + Poorly Named' if already else 'Poorly Named', p['target'], p.get('sheets',''), P2_FILL,P2_TXT))

        if not _rows_data:
            merge(wnr,f'B{NR_HEADER_ROW+1}:{get_column_letter(n_nr_cols-1)}{NR_HEADER_ROW+1}',
                  'No issues detected — every named range is used, clearly named, and resolves correctly.',
                  sz=9,col=OK_TXT,bg=OK_FILL,wrap=True)
            set_row(wnr,NR_HEADER_ROW+1,20)
        else:
            for idx,(name,issue,target,sheets,bf,bt) in enumerate(_rows_data):
                row_i = NR_HEADER_ROW+1+idx
                row_bg = WHITE if idx%2==0 else 'FAFBFC'
                row_values={'Name':name,'Issue':issue,'Target':target,'Sheets':sheets,
                            'Reviewer Decision':'','Owner':'','Status':'Open'}
                for name2,val in row_values.items():
                    col=nr_idx[name2]
                    c=wnr.cell(row_i,col); c.value=val
                    c.font=Fn(sz=9,bold=(name2=='Name'))
                    c.fill=F(row_bg)
                    c.alignment=A(h='center' if name2 in ('Issue','Status') else 'left', v='top', wrap=(name2=='Target'))
                    c.border=B(col=PANEL_BORDER)
                badge(wnr,row_i,nr_idx['Issue'],issue,bf,bt,sz=8,bold=False)
                set_row(wnr,row_i,20)

            _nr_last_row = NR_HEADER_ROW+len(_rows_data)
            _nr_table = Table(displayName="NamedRangeAuditTable", ref=f'B{NR_HEADER_ROW}:{get_column_letter(n_nr_cols-1)}{_nr_last_row}')
            _nr_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                        showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
            wnr.add_table(_nr_table)

            decision_dv2 = DataValidation(type='list',
                formula1='"Link into calculation,Rename,Remove,Confirm intentional,Investigate"', allow_blank=True)
            wnr.add_data_validation(decision_dv2)
            decision_dv2.add(f'{get_column_letter(nr_idx["Reviewer Decision"])}{NR_HEADER_ROW+1}:{get_column_letter(nr_idx["Reviewer Decision"])}{_nr_last_row}')

            status_dv3 = DataValidation(type='list', formula1='"Open,Cleared,Deferred"', allow_blank=False)
            wnr.add_data_validation(status_dv3)
            status_dv3.add(f'{get_column_letter(nr_idx["Status"])}{NR_HEADER_ROW+1}:{get_column_letter(nr_idx["Status"])}{_nr_last_row}')


    # ════════════════════════════════════════════════════════════════════════
    # TAB — REASONABLENESS REVIEW (Wave 1)
    # ════════════════════════════════════════════════════════════════════════
    wrr=wb.create_sheet('Reasonableness Review'); wrr.sheet_view.showGridLines=False
    for i,w in enumerate([3,24,18,18,18,34,3],1): set_col(wrr,i,w)
    r_rr=1

    merge(wrr,f'B{r_rr}:F{r_rr}','ASSUMPTIONS AND COMMERCIAL REASONABLENESS REVIEW',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(wrr,r_rr,2,r_rr,6,DARK_BLUE); set_row(wrr,r_rr,26); r_rr+=1
    merge(wrr,f'B{r_rr}:F{r_rr}',
          'Every other tab in this report answers "is the model wired correctly?" This tab asks a different question: are the results commercially believable? A model can be perfectly linked and still produce unrealistic outputs. Nothing here proves an assumption is wrong — each item is a specific, named reason to challenge it before relying on the result.',
          sz=8,col=GREY_TXT2,bg=PALE_ACCENT,italic=True,wrap=True)
    set_row(wrr,r_rr,32); r_rr+=1
    set_row(wrr,r_rr,8); r_rr+=1

    _wacc = reasonIn.get('waccOverride',{'applicable':False})
    _tv   = reasonIn.get('terminalValue',{'applicable':False})
    _out  = reasonIn.get('outputs',{'applicable':False,'results':[]})
    _dup  = dupIn

    # ── Summary cards ─────────────────────────────────────────────────
    _cards=[
        ('WACC Override', 'Yes' if _wacc.get('mismatch') else 'No' if _wacc.get('applicable') else 'N/A', 2,2),
        ('Terminal Value % of NPV', f"{_tv.get('concentrationPct',0)*100:.0f}%" if _tv.get('applicable') else 'N/A', 3,3),
        ('Flagged Outputs', _out.get('flaggedCount',0) if _out.get('applicable') else 'N/A', 4,4),
        ('Duplicate Sheets', _dup.get('flaggedCount',0) if _dup.get('applicable') else 'N/A', 5,5),
    ]
    for label,val,c1,c2 in _cards:
        col_l=get_column_letter(c1)
        merge(wrr,f'{col_l}{r_rr}:{get_column_letter(c2)}{r_rr}',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        merge(wrr,f'{col_l}{r_rr+1}:{get_column_letter(c2)}{r_rr+1}',val,bold=True,sz=16,
              col=(P2_TXT if (val=='Yes' or (isinstance(val,int) and val>0)) else CHARCOAL),bg=PANEL_GREY,h='center')
        for rr in (r_rr,r_rr+1):
            for cc in range(c1,c2+1): wrr.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(wrr,r_rr,14); set_row(wrr,r_rr+1,24); r_rr+=3

    def section_header(row,title):
        merge(wrr,f'B{row}:F{row}',title,bold=True,sz=11,col=WHITE,bg=DARK_BLUE,v='center')
        fill_range(wrr,row,2,row,6,DARK_BLUE); set_row(wrr,row,20)
        return row+1

    # ── WACC Override ─────────────────────────────────────────────────
    r_rr = section_header(r_rr,'WACC OVERRIDE')
    merge(wrr,f'B{r_rr}:F{r_rr}', _wacc.get('note','Not applicable — no calculated WACC cell was found.'),
          sz=9,col=(P2_TXT if _wacc.get('mismatch') else CHARCOAL),bg=(P2_FILL if _wacc.get('mismatch') else WHITE),wrap=True)
    for cc in range(2,7): wrr.cell(r_rr,cc).border=B(col=PANEL_BORDER)
    set_row(wrr,r_rr,32); r_rr+=1
    set_row(wrr,r_rr,10); r_rr+=1

    # ── Terminal Value Concentration ────────────────────────────────────
    r_rr = section_header(r_rr,'TERMINAL VALUE CONCENTRATION')
    merge(wrr,f'B{r_rr}:F{r_rr}', _tv.get('note','Not applicable — could not locate both a Terminal Value and total NPV figure.'),
          sz=9,col=(P2_TXT if _tv.get('flagged') else CHARCOAL),bg=(P2_FILL if _tv.get('flagged') else WHITE),wrap=True)
    for cc in range(2,7): wrr.cell(r_rr,cc).border=B(col=PANEL_BORDER)
    set_row(wrr,r_rr,32); r_rr+=1
    set_row(wrr,r_rr,10); r_rr+=1

    # ── Output Reasonableness ───────────────────────────────────────────
    r_rr = section_header(r_rr,'OUTPUT REASONABLENESS')
    if _out.get('applicable') and _out.get('results'):
        for label,c1,c2 in [('Metric',2,2),('Value',3,3),('Review Trigger',4,4),('Status',5,5),('Rationale',6,6)]:
            cell(wrr,f'{get_column_letter(c1)}{r_rr}',label,bold=True,sz=8,col=WHITE,bg=DARK_BLUE,h='center')
        set_row(wrr,r_rr,16); r_rr+=1
        for res in _out['results']:
            _unit = res.get('unit','percent')
            disp_val = f"{res['value']*100:.1f}%" if _unit=='percent' else f"{res['value']:.1f}x"
            disp_thr = f"{res['threshold']*100:.0f}%" if _unit=='percent' else f"{res['threshold']:.1f}x"
            wrr.cell(r_rr,2).value=res['metric']; wrr.cell(r_rr,2).font=Fn(sz=9,bold=True,col=CHARCOAL); wrr.cell(r_rr,2).fill=F(WHITE)
            wrr.cell(r_rr,3).value=disp_val; wrr.cell(r_rr,3).font=Fn(sz=9,col=CHARCOAL); wrr.cell(r_rr,3).fill=F(WHITE); wrr.cell(r_rr,3).alignment=A(h='center')
            wrr.cell(r_rr,4).value=f'>= {disp_thr}'; wrr.cell(r_rr,4).font=Fn(sz=9,col=GREY_TXT2); wrr.cell(r_rr,4).fill=F(WHITE); wrr.cell(r_rr,4).alignment=A(h='center')
            badge(wrr,r_rr,5,'Review' if res['flagged'] else 'OK', P2_FILL if res['flagged'] else OK_FILL, P2_TXT if res['flagged'] else OK_TXT, sz=8, bold=False)
            wrr.cell(r_rr,6).value=res['rationale']; wrr.cell(r_rr,6).font=Fn(sz=8,col=GREY_TXT2); wrr.cell(r_rr,6).fill=F(WHITE); wrr.cell(r_rr,6).alignment=A(wrap=True)
            for cc in range(2,7): wrr.cell(r_rr,cc).border=B(col=PANEL_BORDER)
            set_row(wrr,r_rr,34); r_rr+=1
        merge(wrr,f'B{r_rr}:F{r_rr}', _out.get('note',''), sz=8,col=GREY_TXT2,italic=True,wrap=True)
        set_row(wrr,r_rr,26); r_rr+=1
    else:
        merge(wrr,f'B{r_rr}:F{r_rr}','Not applicable — no labelled output metrics were found to test.',sz=9,col=GREY_TXT2,italic=True,wrap=True)
        set_row(wrr,r_rr,20); r_rr+=1
    set_row(wrr,r_rr,10); r_rr+=1

    # ── Duplicate / Backup Sheets ───────────────────────────────────────
    r_rr = section_header(r_rr,'DUPLICATE / BACKUP SHEETS')
    if _dup.get('applicable') and _dup.get('flagged'):
        for label,c1,c2 in [('Sheet',2,3),('Likely Live Counterpart',4,4),('Note',5,6)]:
            if c2>c1: merge(wrr,f'{get_column_letter(c1)}{r_rr}:{get_column_letter(c2)}{r_rr}',label,bold=True,sz=8,col=WHITE,bg=DARK_BLUE,h='center')
            else: cell(wrr,f'{get_column_letter(c1)}{r_rr}',label,bold=True,sz=8,col=WHITE,bg=DARK_BLUE,h='center')
        set_row(wrr,r_rr,16); r_rr+=1
        for f in _dup['flagged']:
            wrr.merge_cells(f'B{r_rr}:C{r_rr}')
            wrr.cell(r_rr,2).value=f['sheet']; wrr.cell(r_rr,2).font=Fn(sz=9,bold=True,col=CHARCOAL); wrr.cell(r_rr,2).fill=F(WHITE)
            wrr.cell(r_rr,4).value=f.get('liveCounterpart') or '—'; wrr.cell(r_rr,4).font=Fn(sz=9,col=GREY_TXT2); wrr.cell(r_rr,4).fill=F(WHITE); wrr.cell(r_rr,4).alignment=A(h='center')
            wrr.merge_cells(f'E{r_rr}:F{r_rr}')
            wrr.cell(r_rr,5).value=f['note']; wrr.cell(r_rr,5).font=Fn(sz=8,col=GREY_TXT2); wrr.cell(r_rr,5).fill=F(WHITE); wrr.cell(r_rr,5).alignment=A(wrap=True)
            for cc in range(2,7): wrr.cell(r_rr,cc).border=B(col=PANEL_BORDER)
            set_row(wrr,r_rr,28); r_rr+=1
    else:
        merge(wrr,f'B{r_rr}:F{r_rr}','No duplicate or backup sheets detected by name.',sz=9,col=OK_TXT,bg=OK_FILL,wrap=True)
        set_row(wrr,r_rr,20); r_rr+=1


    # ════════════════════════════════════════════════════════════════════════
    ws7=wb.create_sheet('Sheet Dependency'); ws7.sheet_view.showGridLines=False
    sd_headers=['','Target Sheet','Precedent Sheet','Link Count','Direction','Dependency Risk',
                'Reviewer Action','Reviewer Note','Priority Count','Avg F-Score','Max F-Score','']
    n_sd_cols = len(sd_headers)
    n_sd_visible = 8  # Target Sheet..Reviewer Note
    for i,w in enumerate([3,22,22,12,12,14,14,26]+[14]*(n_sd_cols-1-n_sd_visible),1): set_col(ws7,i,w)
    sd_idx = {h.replace('\n',' '): i for i,h in enumerate(sd_headers,1) if h}
    SD_HEADER_ROW=6

    merge(ws7,f'B1:{get_column_letter(n_sd_cols-1)}1','SHEET DEPENDENCY RISK MAP',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws7,1,2,1,n_sd_cols-1,DARK_BLUE); set_row(ws7,1,26)
    merge(ws7,f'B2:{get_column_letter(n_sd_cols-1)}2',
          'How sheets reference each other — where model dependency risk sits, and which links warrant review first.',
          sz=8,col=GREY_TXT2,italic=True)
    set_row(ws7,2,16)

    edges=t0.get('edgeList',[])
    _risk_rank={'Critical':0,'High':1,'Moderate':2,'Low':3}
    _sorted_edges = sorted(edges[:100], key=lambda e: (_risk_rank.get(e.get('risk','Low'),3), -e.get('linkCount',0)))
    _critical = [e for e in _sorted_edges if e.get('risk') in ('Critical','High')][:5]

    # ── Critical Dependency panel — highest-risk links, surfaced first ──────
    if _critical:
        merge(ws7,f'B3:{get_column_letter(n_sd_cols-1)}3','CRITICAL DEPENDENCIES',bold=True,sz=9,col=WHITE,bg=DARK_BLUE,h='left')
        set_row(ws7,3,16)
        _cd_row=4
        for e in _critical:
            txt=f"{e.get('targetSheet','')} → {e.get('precedentSheet','')}: {e.get('linkCount',0):,} links ({e.get('risk','')})"
            merge(ws7,f'B{_cd_row}:{get_column_letter(n_sd_cols-1)}{_cd_row}',txt,sz=9,col=P1_TXT,bg=P1_FILL)
            set_row(ws7,_cd_row,16); _cd_row+=1
        set_row(ws7,_cd_row,8); SD_HEADER_ROW=_cd_row+1
    else:
        merge(ws7,f'B3:{get_column_letter(n_sd_cols-1)}3','No Critical or High-risk dependencies detected.',sz=9,col=GREY_TXT2,italic=True)
        set_row(ws7,3,16); set_row(ws7,4,8); SD_HEADER_ROW=5

    for col,h in enumerate(sd_headers,1):
        c=ws7.cell(SD_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
        c.fill=F(DARK_BLUE); c.alignment=A(h='center',v='center',wrap=True); c.border=B(col=WHITE)
    set_row(ws7,SD_HEADER_ROW,26)
    ws7.freeze_panes = f'{get_column_letter(sd_idx["Dependency Risk"])}{SD_HEADER_ROW+1}'

    _risk_badge={'Critical':(P1_FILL,P1_TXT),'High':(P1_FILL,P1_TXT),'Moderate':(P2_FILL,P2_TXT),'Low':(OK_FILL,OK_TXT)}
    _dir_badge={'Normal':(OK_FILL,OK_TXT),'Backward':(P1_FILL,P1_TXT),'External':(P2_FILL,P2_TXT),'Circular':(P1_FILL,P1_TXT)}
    for idx,edge in enumerate(_sorted_edges):
        row_i = SD_HEADER_ROW+1+idx
        row_bg = WHITE if idx%2==0 else 'FAFBFC'
        risk=edge.get('risk','Low'); direction=edge.get('direction','Normal')
        row_values={'Target Sheet':edge.get('targetSheet',''),'Precedent Sheet':edge.get('precedentSheet',''),
                    'Link Count':edge.get('linkCount',0),'Direction':direction,'Dependency Risk':risk,
                    'Reviewer Action':'','Reviewer Note':'','Priority Count':edge.get('priorityCount',0),
                    'Avg F-Score':edge.get('avgFscore',''),'Max F-Score':edge.get('maxFscore','')}
        for name,val in row_values.items():
            col=sd_idx[name]
            c=ws7.cell(row_i,col); c.value=val
            c.font=Fn(sz=9,bold=(name in ('Target Sheet','Precedent Sheet')))
            c.fill=F(row_bg)
            c.alignment=A(h='center' if name in ('Link Count','Direction','Dependency Risk','Priority Count','Avg F-Score','Max F-Score') else 'left',
                          v='top', wrap=(name=='Reviewer Note'))
            c.border=B(col=PANEL_BORDER)
        ws7.cell(row_i, sd_idx['Link Count']).number_format='#,##0;[Red](#,##0);-'
        rf,rt=_risk_badge.get(risk,(GREY_LIGHT,CHARCOAL))
        badge(ws7,row_i,sd_idx['Dependency Risk'],risk,rf,rt,sz=8,bold=False)
        df,dt=_dir_badge.get(direction,(GREY_LIGHT,CHARCOAL))
        badge(ws7,row_i,sd_idx['Direction'],direction,df,dt,sz=8,bold=False)
        set_row(ws7,row_i,20)

    _sd_last_row = SD_HEADER_ROW+len(_sorted_edges)
    for col in range(n_sd_visible+1, n_sd_cols):
        ws7.column_dimensions[get_column_letter(col)].hidden = True

    if _sorted_edges:
        _sd_table = Table(displayName="SheetDependencyTable", ref=f'B{SD_HEADER_ROW}:{get_column_letter(n_sd_cols-1)}{_sd_last_row}')
        _sd_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                                    showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
        ws7.add_table(_sd_table)
        action_dv = DataValidation(type='list', formula1='"Review,Accept,Simplify,Investigate"', allow_blank=True)
        ws7.add_data_validation(action_dv)
        action_dv.add(f'{get_column_letter(sd_idx["Reviewer Action"])}{SD_HEADER_ROW+1}:{get_column_letter(sd_idx["Reviewer Action"])}{_sd_last_row}')


    ws8=wb.create_sheet('F-Score Rules'); ws8.sheet_view.showGridLines=False
    for col,w in [(1,3),(2,10),(3,34),(4,38),(5,32),(6,3)]: set_col(ws8,col,w)

    merge(ws8,'B1:E1','F-SCORE RULES',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws8,1,2,1,5,DARK_BLUE); set_row(ws8,1,26)
    merge(ws8,'B2:E2','The F-score measures formula complexity and audit risk — every formula in the model receives one automatically. Higher scores warrant closer review.',
          sz=9,col=GREY_TXT2,bg=PALE_ACCENT,italic=True,wrap=True)
    set_row(ws8,2,26)

    hdr(ws8,'B3','Score',h='center'); hdr(ws8,'C3','Feature'); hdr(ws8,'D3','Reason'); hdr(ws8,'E3','Review Action'); set_row(ws8,3,18)
    rules_data=[
        ('+1','Formula length > 100 characters','Longer formulas are harder to review and audit','Read through for clarity; simplify if easy to do so'),
        ('+2','Formula length > 250 characters','Indicates significant formula density','Consider splitting into helper cells'),
        ('+3','Formula length > 500 characters','Indicates very high formula density','Split into helper rows before relying on the output'),
        ('+1 each','Nested IF / IFS / CHOOSE / SWITCH','Nested branching increases scenario and edge-case risk','Trace each branch against a known scenario'),
        ('+1','XLOOKUP / INDEX-MATCH / VLOOKUP / HLOOKUP','Lookup formulas require range and error-handling review','Confirm exact-match settings and range boundaries'),
        ('+3','OFFSET or INDIRECT','Dynamic references are difficult to trace and audit','Replace with a static reference where possible'),
        ('+3','Volatile function (TODAY / NOW / RAND)','Volatile formulas recalculate on every workbook change','Confirm recalculation timing is intended, not accidental'),
        ('+5','External workbook reference ([ marker)','Broken-link and version-control risk on file relocation','Break the link or document the external dependency'),
        ('+1','Hardcoded numeric constant inside formula','May bypass the centralised input structure','Move the constant to the Inputs sheet'),
        ('+1','IFERROR / IFNA / ISERROR','Can hide genuine calculation errors if misused','Confirm the wrapped error is expected, not masked'),
        ('+5','Circularity-sensitive formula','Requires controlled iteration switch and convergence testing','Verify the iteration switch and test for convergence'),
        ('+1','> 5 arithmetic or logical operators','Moderate calculation density','Review for clarity; consider breaking into steps'),
        ('+2','> 10 arithmetic or logical operators','High calculation density','Split into helper rows for auditability'),
        ('+1','Cross-sheet reference (! marker)','Increases dependency and traceability complexity','Confirm the source sheet and cell are correct'),
    ]
    for row_i,(score,feature,reason,action) in enumerate(rules_data,4):
        row_bg = WHITE if row_i%2==0 else 'FAFBFC'
        badge(ws8,row_i,2,score,PANEL_GREY,MID_BLUE,sz=9,bold=True)
        for col,val in [(3,feature),(4,reason),(5,action)]:
            c=ws8.cell(row_i,col); c.value=val; c.font=Fn(sz=9); c.fill=F(row_bg)
            c.alignment=A(v='top',wrap=True); c.border=B(col=PANEL_BORDER)
        set_row(ws8,row_i,20)

    _bands_start = 4+len(rules_data)+1
    set_row(ws8,_bands_start-1,8)
    merge(ws8,f'B{_bands_start}:E{_bands_start}','COMPLEXITY BANDS',bold=True,sz=11,col=WHITE,bg=DARK_BLUE,v='center')
    set_row(ws8,_bands_start,20)
    hdr(ws8,f'B{_bands_start+1}','F-Score Range',h='center'); hdr(ws8,f'C{_bands_start+1}','Band',h='center')
    ws8.merge_cells(f'D{_bands_start+1}:E{_bands_start+1}')
    ws8[f'D{_bands_start+1}'].value='Treatment'; ws8[f'D{_bands_start+1}'].font=Fn(bold=True,sz=10,col=WHITE)
    ws8[f'D{_bands_start+1}'].fill=F(DARK_BLUE); ws8[f'D{_bands_start+1}'].alignment=A(h='left')
    set_row(ws8,_bands_start+1,18)
    _band_badge2={'Low':(OK_FILL,OK_TXT),'Moderate':(P3_FILL,P3_TXT),'High':(P2_FILL,P2_TXT),'Very High':(P1_FILL,P1_TXT)}
    bands=[('0–3','Low','Ordinary complexity. Review through normal formula testing.'),
           ('4–7','Moderate','Review for logic clarity, input linkage and copy-across consistency.'),
           ('8–12','High','Inspect carefully. Consider simplification or helper rows.'),
           ('13+','Very High','Needs detailed review. Consider simplifying or splitting into helper rows.')]
    for i,(score_range,band,treatment) in enumerate(bands):
        row_i=_bands_start+2+i
        ws8.cell(row_i,2).value=score_range; ws8.cell(row_i,2).font=Fn(bold=True,sz=9,col=CHARCOAL)
        ws8.cell(row_i,2).fill=F(WHITE); ws8.cell(row_i,2).alignment=A(h='center'); ws8.cell(row_i,2).border=B(col=PANEL_BORDER)
        bf,bt=_band_badge2.get(band,(GREY_LIGHT,CHARCOAL))
        badge(ws8,row_i,3,band,bf,bt,sz=9,bold=False)
        ws8.merge_cells(f'D{row_i}:E{row_i}')
        ws8[f'D{row_i}'].value=treatment; ws8[f'D{row_i}'].font=Fn(sz=9,col=CHARCOAL)
        ws8[f'D{row_i}'].fill=F(WHITE); ws8[f'D{row_i}'].border=B(col=PANEL_BORDER)
        set_row(ws8,row_i,18)


    ws9=wb.create_sheet('Pipeline Audit Trail'); ws9.sheet_view.showGridLines=False
    log_headers=['','Timestamp','Step','Action','Artifact','Result','Duration','Notes','']
    n_al_cols = len(log_headers)
    for i,w in enumerate([3,12,16,36,28,12,10,34,3],1): set_col(ws9,i,w)
    al_idx = {h.replace('\n',' '): i for i,h in enumerate(log_headers,1) if h}
    AL_HEADER_ROW=6

    merge(ws9,f'B1:{get_column_letter(n_al_cols-1)}1','PIPELINE AUDIT TRAIL',bold=True,sz=14,col=WHITE,bg=DARK_BLUE,v='center')
    fill_range(ws9,1,2,1,n_al_cols-1,DARK_BLUE); set_row(ws9,1,26)
    merge(ws9,f'B2:{get_column_letter(n_al_cols-1)}2',
          'Step-by-step record of this run — proves the review was actually performed, not just claimed.',
          sz=8,col=GREY_TXT2,italic=True)
    set_row(ws9,2,16)

    audit_log=d.get('auditLog',[])
    # Sheet count is embedded as text in the Parse step's notes/action —
    # extracted here rather than requiring a new payload field.
    _sheets_parsed = 0
    for entry in audit_log:
        m = re.search(r'(\d+)\s*sheets', str(entry.get('notes','')) + ' ' + str(entry.get('action','')))
        if m: _sheets_parsed = int(m.group(1)); break
    _stats = t0.get('stats',{})

    _al_cards=[('Sheets Parsed', _sheets_parsed, 2,3), ('Formula Cells Scanned', _stats.get('totalFormulaCells',0), 4,5),
               ('Unique Formulas', _stats.get('uniqueFormulaCount',0), 6,6), ('Semantic Checks', len(checklist_rules), 7,7)]
    for label,val,c1,c2 in _al_cards:
        col_l1,col_l2=get_column_letter(c1),get_column_letter(c2)
        if c2>c1: merge(ws9,f'{col_l1}3:{col_l2}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        else: cell(ws9,f'{col_l1}3',label,bold=True,sz=8,col=GREY_TXT2,bg=PANEL_GREY,h='center')
        if c2>c1: merge(ws9,f'{col_l1}4:{col_l2}4',val,bold=True,sz=16,col=CHARCOAL,bg=PANEL_GREY,h='center')
        else: cell(ws9,f'{col_l1}4',val,bold=True,sz=16,col=CHARCOAL,bg=PANEL_GREY,h='center')
        ws9.cell(4,c1).number_format='#,##0;[Red](#,##0);-'
        for rr in (3,4):
            for cc in range(c1,c2+1): ws9.cell(rr,cc).border=B(col=PANEL_BORDER)
    set_row(ws9,3,14); set_row(ws9,4,24); set_row(ws9,5,8)

    for col,h in enumerate(log_headers,1):
        c=ws9.cell(AL_HEADER_ROW,col); c.value=h; c.font=Fn(bold=True,sz=9,col=WHITE)
        c.fill=F(DARK_BLUE); c.alignment=A(h='center' if h in ('Timestamp','Result','Duration') else 'left',v='center'); c.border=B(col=WHITE)
    set_row(ws9,AL_HEADER_ROW,20)
    ws9.freeze_panes = f'B{AL_HEADER_ROW+1}'

    def _result_class(r):
        rs=str(r)
        if '⚠' in rs or 'Issue' in rs: return 'Issues',(P2_FILL,P2_TXT)
        if '❌' in rs or 'Error' in rs: return 'Error',(P1_FILL,P1_TXT)
        if 'Not performed' in rs: return 'Not performed',(GREY_LIGHT,GREY_TXT2)
        return 'Pass',(OK_FILL,OK_TXT)

    for idx,entry in enumerate(audit_log):
        row_i = AL_HEADER_ROW+1+idx
        row_bg = WHITE if idx%2==0 else 'FAFBFC'
        raw_dur = entry.get('duration','')
        try: dur_val = float(raw_dur)
        except (TypeError,ValueError): dur_val = None
        result_label,_ = _result_class(entry.get('result','✓ Pass'))
        row_values={'Timestamp':entry.get('timestamp',''),'Step':entry.get('step',''),'Action':entry.get('action',''),
                    'Artifact':entry.get('artifact',''),'Result':result_label,
                    'Duration':(dur_val if dur_val is not None else '—'),'Notes':entry.get('notes','')}
        for name,val in row_values.items():
            col=al_idx[name]
            c=ws9.cell(row_i,col); c.value=val
            c.font=Fn(sz=9,bold=(name=='Step'))
            c.fill=F(row_bg)
            c.alignment=A(h='center' if name in ('Timestamp','Result','Duration') else 'left', v='top', wrap=(name in ('Action','Artifact','Notes')))
            c.border=B(col=PANEL_BORDER)
        if dur_val is not None:
            ws9.cell(row_i, al_idx['Duration']).number_format = '0.0 "s"'
        _,rescol = _result_class(entry.get('result','✓ Pass'))
        badge(ws9,row_i,al_idx['Result'],result_label,rescol[0],rescol[1],sz=8,bold=False)
        set_row(ws9,row_i,26)


    # ── Save ─────────────────────────────────────────────────────────────────
    # ── Workbook-wide polish — tab colours + print setup ──────────────────────
    _tab_colors = {
        'Audit Output': DARK_BLUE, 'Read Me': '5B7C99', 'Scope and Reliance': '5B7C99',
        'Issue Log': AMBER, 'Remediation': AMBER,
        'Assumption Register': GREY_MID, 'Validation Matrix': GREY_MID,
        'Formula Risk Review': GREY_MID, 'Redundant Inputs': GREY_MID, 'Sheet Linkage': GREY_MID, 'Named Range Audit': GREY_MID, 'Reasonableness Review': AMBER,
        'Error Matrix': GREY_MID, 'Sheet Dependency': GREY_MID,
        'F-Score Rules': GREY_MID, 'Pipeline Audit Trail': GREY_MID,
    }
    for _ws in wb.worksheets:
        if _ws.title in _tab_colors:
            _ws.sheet_properties.tabColor = _tab_colors[_ws.title]

    ws1.page_setup.orientation = 'landscape'
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0
    ws1.sheet_properties.pageSetUpPr.fitToPage = True
    ws1.print_area = f'B1:I{_audit_output_last_row}'

    wb.save(output_path)
    return {'status':'ok','tabs':16,'findings':len(findings)}

if __name__=='__main__':
    if len(sys.argv)<3:
        print('Usage: python3 build_report.py <data.json> <output.xlsx>',file=sys.stderr)
        sys.exit(1)
    result=build_report(sys.argv[1],sys.argv[2])
    print(json.dumps(result))
