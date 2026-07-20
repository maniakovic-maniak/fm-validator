#!/usr/bin/env python3
"""
recalc_check.py — A1: real formula recalculation vs. cached values,
using Formualizer (github.com/PSU3D0/formualizer, verified real and MIT
licensed).

Three real, distinct problems were found and fixed via direct testing
against real files, in order of discovery:

1. CIRCULARITY CONFIGURATION. Formualizer's high-level recalculate_file()
   convenience function uses cycle_detection="static" by default, which
   stamps EVERY circular reference as a #CIRC! error rather than
   attempting Excel-style iterative resolution. Against The Bend (a real
   project-finance model with a genuine, intentional iterative
   circularity — see this project's own G7 check), this produced 916
   entirely false "errors" on the first test run. This script ALWAYS
   configures cycle_policy="iterate" explicitly via the lower-level
   Workbook API.

2. EXTERNAL WORKBOOK REFERENCES. A formula like '[1]SheetName'!$A$1
   references another file this workbook doesn't have loaded.
   Formualizer 0.7.1's evaluate_all() treats hitting even ONE such
   unresolvable reference as fatal for the ENTIRE evaluation — every
   other cell comes back None afterward, even cells unrelated to the
   external link. Confirmed via testing against Hidden Gem, a real
   mining model. Fixed by pre-scanning with openpyxl for the '[N]'
   bracket pattern and neutralizing those specific cells via
   Formualizer's own set_value() mutation API before evaluate_all() runs.
   Neutralized cells are tracked separately and always excluded from the
   mismatch comparison — their value is fabricated, never a genuine
   recalculation.

3. SCALE. Also discovered via Hidden Gem, a real mining model with
   1,152,793 formula cells (vs. 15,690 for The Bend): two design choices
   that work fine at Bend's scale fail badly at Hidden Gem's.
     a. openpyxl's DEFAULT (non-read_only) mode builds a full in-memory
        object tree — combined with Formualizer's own ~1.4GB load for
        the same file, this caused an out-of-memory kill with no error
        output at all. read_only=True uses a streaming parser instead,
        confirmed via direct testing to use ~55MB combined for both
        loads on the same file — now the PRIMARY approach, not a
        fallback. It also happens to use a more lenient XML parser that
        resolved a separate "invalid XML" failure on this same file.
     b. Looping through individual wb.evaluate_cell() calls is fine for
        15,690 cells (~20s) but would take over 40 MINUTES at Hidden
        Gem's scale (confirmed: ~2.26ms/cell). The batch wb.evaluate_cells()
        call is ~300x faster (confirmed: ~0.008ms/cell) — the same
        1.15M-cell comparison drops to single-digit seconds. Similarly,
        RANDOM cell access into a read_only-mode workbook (e.g.
        ws['J75'].value repeatedly) is catastrophically slow in
        read_only mode (confirmed: ~12.6ms per access — over 4 HOURS at
        1.15M cells) because it isn't built for random access; PARALLEL
        SEQUENTIAL iteration over both workbooks via zip(iter_rows(),
        iter_rows()) is the correct pattern and is fast (confirmed:
        ~2.3s for 15,690 cells).

With all three fixes applied: The Bend (15,690 cells, 20 genuine
circular groups) and Carlsberg (5,665 cells) both return 0 mismatches
and 0 errors — a clean baseline. Testing against Hidden Gem specifically
motivated fixes 2 and 3; re-confirm against it after any future change
to this script, since it is the only real file tested so far that
exercises the external-reference and scale code paths at all.

4. UNSUPPORTED MODERN EXCEL FUNCTIONS. Formualizer 0.7.1's function
   registry does not recognize several functions introduced with
   Excel's dynamic-array era. Confirmed individually, each via multiple
   independent test shapes (not a one-off argument-shape mistake):
     a. _xlfn.SINGLE — Excel's serialized form of the "@" implicit-
        intersection operator. Formualizer DOES support the "@" operator
        itself (confirmed: "=@A1" evaluates correctly) but does not
        recognize the literal function name "_xlfn.SINGLE" as an alias
        for it — {'type': 'Error', 'kind': 'Name'}. Root-caused live
        against Hidden Gem: this is what silently zeroed out Ops!U193
        and 50 related cells (see the masking-BFS fix history below) —
        Timing!E7 = _xlfn.SINGLE(Ops_end) failed with a Name error, and
        that error propagated through several formula hops before being
        silently absorbed by a SUM()/SUMIF() into a plausible-looking
        0.0, rather than surfacing as an error at the top level.
     b. LAMBDA — returns Formualizer's own explicit
        {'type': 'Error', 'kind': 'NImpl'} (Not Implemented) for the
        realistic inline-invoked usage pattern ("=LAMBDA(x,x*2)(A1)") —
        a genuinely different signal from "unrecognized name": Formualizer
        knows about LAMBDA and has explicitly marked it unimplemented,
        not merely unrecognized.
     c. EXPAND — {'type': 'Error', 'kind': 'Name'}, confirmed across
        three independent argument shapes (with and without the optional
        pad-value argument, and a 1x1-to-2x2 expansion).
   Any model using one of these will have that formula silently produce
   a wrong-but-plausible recalculated value if the error is absorbed by
   a downstream SUM()/SUMIF() before reaching a target cell being
   compared — exactly the class of bug the masking-BFS fixes below exist
   to correctly attribute rather than report as an unexplained mismatch.
   No code-level workaround applied here (unlike fixes 1-3 above) — this
   is a genuine Formualizer limitation to track, not something this
   script can neutralize the way external references are neutralized,
   since (unlike an external reference) there's no way to know in
   advance which specific cells in a real file use one of these
   functions without already having tried to evaluate them.

Usage:
    python3 recalc_check.py path/to/workbook.xlsx
    (outputs a JSON summary to stdout)
"""

import sys
import json
import time
import re
import os
import math
import warnings

# Confirmed via direct inspection of openpyxl's own source
# (openpyxl/worksheet/_reader.py): when a date-formatted cell's cached
# value is NaN (see the _cast_number patch below for why that value can
# legitimately be NaN in the first place), openpyxl tries to convert it
# to a date via from_excel(), which correctly raises on NaN — and
# openpyxl's OWN try/except catches that internally, sets the cell to
# "#VALUE!", and emits this warning purely for information. It is not a
# sign of a problem this script needs to react to. At real-file scale
# (hundreds+ of affected cells in one sheet observed in testing) this
# was flooding stderr with no functional benefit — suppressed here.
warnings.filterwarnings('ignore', message=r'.*is marked as a date but the serial value.*')

try:
    import openpyxl
    # openpyxl's internal number parser (_cast_number) only recognizes a
    # float by the presence of '.', 'E', or 'e' in the text, and falls
    # through to int(value) otherwise — which crashes on 'NaN',
    # 'Infinity', and '-Infinity', all valid IEEE-754 values Excel can
    # legitimately cache as a formula's result (e.g. certain invalid
    # statistical operations). Confirmed via a real crash against Hidden
    # Gem: this happens deep inside openpyxl's internal row-parsing
    # generator, where a per-cell try/except in THIS script can't catch
    # it without losing every other cell in the same row generator —
    # patched at the source instead, before any workbook is loaded.
    import openpyxl.worksheet._reader as _oxl_reader

    _original_cast_number = _oxl_reader._cast_number

    def _patched_cast_number(value):
        if value in ('NaN', 'nan'):
            return float('nan')
        if value in ('Infinity', 'INF'):
            return float('inf')
        if value in ('-Infinity', '-INF'):
            return float('-inf')
        return _original_cast_number(value)

    _oxl_reader._cast_number = _patched_cast_number
except ImportError:
    openpyxl = None  # handled by the existing "unavailable" check below

try:
    import formualizer as fz
except ImportError:
    print(json.dumps({"status": "unavailable", "reason": "formualizer not installed"}))
    sys.exit(0)

if openpyxl is None:
    print(json.dumps({"status": "unavailable", "reason": "openpyxl not installed"}))
    sys.exit(0)

EXTERNAL_REF_RE = re.compile(r'\[\d+\]')

# Confirmed via real testing against Hidden Gem: neutralizing an
# external-reference cell to 0 (see fix #2 in the module docstring)
# correctly avoids the crash, but that fabricated 0 propagates through
# the dependency graph — any cell that references a neutralized cell,
# directly or transitively, inherits the same fabricated value. Without
# tracing this, such cells show up as "mismatches" against their real
# cached value (which reflects the actual external data) even though
# nothing is genuinely wrong with their formula — a real, confirmed
# false-positive class (4,155 cells on a single sheet in one real test,
# all recalculating to exactly 0.0). This section traces forward from
# each neutralized cell to find everything downstream and excludes it
# from the comparison too, not just the neutralized cells themselves.

_CELL_REF_RE = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?\$?([A-Za-z]{1,3})\$?(\d+)(?::\$?([A-Za-z]{1,3})\$?(\d+))?"
)

# FIX (found via mining "Advanced Excel", Jordan — L23): 3D references
# like =SUM(Jan:Dec!B5), aggregating across every sheet from Jan to Dec
# inclusive, were not recognized by _CELL_REF_RE at all. Confirmed via
# direct testing: '=SUM(Jan:Dec!B5)' extracted only ('Dec','B',5) —
# "Jan:" doesn't match anything so the engine just skips past it. Worse,
# '=SUM(Sheet1:Sheet3!B3)' produced an outright CORRUPTED ref —
# ('Summary','EET',1) — because "Sheet1" doesn't match the sheet-prefix
# pattern (no "!" immediately after it) or the plain cell pattern (too
# many letters), but the SUBSTRING "eet1" within it accidentally does
# match \$?([A-Za-z]{1,3})\$?(\d+) using the current sheet as an implicit
# prefix. This regex is matched and MASKED OUT of the formula text before
# _CELL_REF_RE runs, so the existing single-sheet logic never sees (and
# can't corrupt-match into) a 3D-reference span.
_SHEET_RANGE_REF_RE = re.compile(
    r"(?:'([^']+)'|([A-Za-z0-9_]+)):(?:'([^']+)'|([A-Za-z0-9_]+))!"
    r"\$?([A-Za-z]{1,3})\$?(\d+)(?::\$?([A-Za-z]{1,3})\$?(\d+))?"
)


def _col_to_num(col):
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def _num_to_col(n):
    s = ''
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def _sheets_in_span(sheet1, sheet2, sheet_order):
    """Every sheet from sheet1 to sheet2 inclusive, in actual tab order —
    NOT alphabetical or numeric order, since that's what "Jan:Dec!B5"
    genuinely means in Excel (whatever sheets sit between Jan's and
    Dec's tab positions, in whichever direction). Falls back to just the
    two named sheets if sheet_order wasn't supplied or either name isn't
    found in it — degraded (won't catch sheets strictly between them),
    but still correct for those two, and no longer silently drops the
    first one or corrupts the parse the way the old behavior did."""
    if sheet_order and sheet1 in sheet_order and sheet2 in sheet_order:
        i1, i2 = sheet_order.index(sheet1), sheet_order.index(sheet2)
        lo, hi = (i1, i2) if i1 <= i2 else (i2, i1)
        return sheet_order[lo:hi + 1]
    return [sheet1, sheet2] if sheet1 != sheet2 else [sheet1]


def _extract_refs(formula, current_sheet, sheet_order=None, max_range_cells=2000):
    """Best-effort cell/range reference extraction — doesn't need to be a
    complete Excel-grammar parser, just reliable enough to catch real
    references for taint propagation. Pathologically large ranges are
    skipped rather than fully expanded (not worth the cost here).

    sheet_order: the workbook's sheets in actual tab order (e.g.
    wb.sheetnames), used to correctly expand 3D references like
    "Jan:Dec!B5". Optional — falls back to a degraded-but-not-corrupted
    two-sheet interpretation if not supplied."""
    refs = set()

    # Pass 1: 3D (multi-sheet) references — matched and masked out of the
    # formula text before the single-sheet regex runs.
    masked_formula = formula
    for m in _SHEET_RANGE_REF_RE.finditer(formula):
        q1, b1, q2, b2, col1, row1, col2, row2 = m.groups()
        sheet1 = (q1 or b1).strip()
        sheet2 = (q2 or b2).strip()
        for sheet in _sheets_in_span(sheet1, sheet2, sheet_order):
            if col2 and row2:
                c1, c2 = sorted([_col_to_num(col1.upper()), _col_to_num(col2.upper())])
                r1, r2 = sorted([int(row1), int(row2)])
                if (c2 - c1 + 1) * (r2 - r1 + 1) > max_range_cells:
                    continue
                for c in range(c1, c2 + 1):
                    for r in range(r1, r2 + 1):
                        refs.add((sheet, _num_to_col(c), r))
            else:
                refs.add((sheet, col1.upper(), int(row1)))
        start, end = m.span()
        masked_formula = masked_formula[:start] + (' ' * (end - start)) + masked_formula[end:]

    # Pass 2: ordinary single-sheet (or same-sheet) references, over the
    # masked text so 3D-reference spans can't be re-matched or
    # corrupt-matched here.
    for m in _CELL_REF_RE.finditer(masked_formula):
        quoted_sheet, bare_sheet, col1, row1, col2, row2 = m.groups()
        sheet = (quoted_sheet or bare_sheet or current_sheet).strip()
        if col2 and row2:
            c1, c2 = sorted([_col_to_num(col1.upper()), _col_to_num(col2.upper())])
            r1, r2 = sorted([int(row1), int(row2)])
            if (c2 - c1 + 1) * (r2 - r1 + 1) > max_range_cells:
                continue
            for c in range(c1, c2 + 1):
                for r in range(r1, r2 + 1):
                    refs.add((sheet, _num_to_col(c), r))
        else:
            refs.add((sheet, col1.upper(), int(row1)))
    return refs

# Confirmed via direct testing (the system's own OOM killer log) that
# Formualizer's evaluate_all() was killed by the OS after growing to
# ~3.9GB while processing a real 1,152,789-formula-cell mining model —
# Formualizer's own initial load alone took ~1.4GB for that file, before
# evaluate_all() pushed memory past the available ceiling. That test ran
# in a 3.9GB sandbox; local machines and production servers can have
# very different headroom, so this is configurable via the
# RECALC_CHECK_MAX_FORMULA_CELLS environment variable rather than a
# fixed constant — set it high (or to 0) for local runs on a machine
# with plenty of memory, leave the default in place for a
# resource-constrained host (e.g. fm-validator's current CloudLinux/
# shared-hosting production environment, which has separately confirmed
# resource constraints — see the planned VPS migration). 0 or a negative
# value disables the check entirely. An unset or malformed value falls
# back to the default below.
DEFAULT_MAX_FORMULA_CELLS = 500000


def get_max_formula_cells():
    raw = os.environ.get('RECALC_CHECK_MAX_FORMULA_CELLS')
    if raw is None:
        return DEFAULT_MAX_FORMULA_CELLS
    try:
        return int(raw)
    except ValueError:
        return DEFAULT_MAX_FORMULA_CELLS


def count_formula_cells(path):
    """Cheap pre-check (~18s for 1.15M cells in testing, no Formualizer
    load at all) so an oversized file can be skipped gracefully before
    the expensive, memory-intensive part of this script even starts."""
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    count = 0
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    count += 1
    return count


def _progress(msg):
    print(f"[recalc_check] {msg}", file=sys.stderr, flush=True)


def run(path, relative_tolerance=0.001, absolute_tolerance=1.0):
    t_start = time.time()

    _progress("Counting formula cells...")
    try:
        formula_count = count_formula_cells(path)
    except Exception as e:
        return {"status": "cached_value_read_failed", "error": str(e), "elapsed_s": round(time.time() - t_start, 2)}
    _progress(f"{formula_count:,} formula cells found ({round(time.time()-t_start,1)}s elapsed).")

    max_cells = get_max_formula_cells()
    if max_cells > 0 and formula_count > max_cells:
        return {"status": "skipped_too_large", "formula_cells": formula_count,
                "threshold": max_cells,
                "reason": f"{formula_count:,} formula cells exceeds the {max_cells:,}-cell safety threshold (RECALC_CHECK_MAX_FORMULA_CELLS) — skipped to avoid a memory-related crash. Raise this via the environment variable if this machine has more headroom, or set it to 0 to disable the check entirely.",
                "elapsed_s": round(time.time() - t_start, 2)}

    cfg = fz.EvaluationConfig()
    cfg.cycle_policy = "iterate"  # fix #1 — see module docstring, never omit this

    _progress("Loading workbook into Formualizer...")
    try:
        wb = fz.Workbook.load_path(path, config=fz.WorkbookConfig(eval_config=cfg))
    except Exception as e:
        return {"status": "load_or_eval_failed", "error": str(e), "elapsed_s": round(time.time() - t_start, 2)}
    _progress(f"Formualizer load complete ({round(time.time()-t_start,1)}s elapsed).")

    # fix #3a — read_only=True is primary, not a fallback. See module docstring.
    _progress("Loading workbook into openpyxl (for cached values)...")
    try:
        wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
        wb_cached = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        try:
            wb_formulas = openpyxl.load_workbook(path, data_only=False, keep_vba=path.lower().endswith('.xlsm'))
            wb_cached = openpyxl.load_workbook(path, data_only=True, keep_vba=path.lower().endswith('.xlsm'))
        except Exception as e2:
            return {"status": "cached_value_read_failed", "error": str(e), "fallback_error": str(e2), "elapsed_s": round(time.time() - t_start, 2)}
    _progress(f"openpyxl loads complete ({round(time.time()-t_start,1)}s elapsed).")

    # L23 fix: the workbook's actual sheet tab order, needed to correctly
    # expand 3D references like "Jan:Dec!B5" into every sheet in between —
    # "between" means tab position, not alphabetical or creation order.
    sheet_order = list(wb_formulas.sheetnames)

    # Single parallel-sequential pass (fix #3b) that both (i) identifies
    # and neutralizes external-reference cells (fix #2) and (ii) collects
    # the full target list + cached values + formula text for the batch
    # comparison and taint-propagation trace below. Never uses random
    # access (ws[coordinate]) — confirmed catastrophically slow in
    # read_only mode.
    external_ref_cells = set()
    targets = []          # [(sheet, row, col), ...]
    target_keys = []      # ["Sheet!A1", ...] parallel to targets
    cached_values = []    # parallel to targets
    formula_texts = []    # parallel to targets — needed for the taint trace below
    sheet_of_target = {}  # "Sheet!A1" -> sheet name, for the taint trace

    _progress("Scanning formulas and cached values (single pass)...")
    for sheet_name in wb_formulas.sheetnames:
        if sheet_name not in wb_cached.sheetnames:
            continue
        ws_f = wb_formulas[sheet_name]
        ws_c = wb_cached[sheet_name]
        for row_f, row_c in zip(ws_f.iter_rows(), ws_c.iter_rows()):
            for cell_f, cell_c in zip(row_f, row_c):
                if cell_f.data_type != 'f':
                    continue
                # FIX (found via real testing against Hidden Gem at
                # production scale): openpyxl represents a legacy CSE
                # array formula as an ArrayFormula object (attributes
                # .ref, .text), not a plain string. Two consequences if
                # left unnormalized: (1) the EXTERNAL_REF_RE check just
                # below used str(cell_f.value), which on an ArrayFormula
                # stringifies to its object repr, not the actual formula
                # text — silently missing any "[N]" external reference
                # living inside an array formula; (2) the $-stripped
                # pre-filter added for the taint-BFS fixes calls
                # .replace('$', ''), which doesn't exist on ArrayFormula
                # at all and crashed with AttributeError on this exact
                # file. Normalizing once here, at the single point
                # formula_texts is populated, means every downstream
                # consumer (both BFS passes, IFERROR/IFNA regex,
                # _extract_refs, mismatch reporting) always sees plain
                # formula text or None — never a wrapper object.
                formula_value = cell_f.value
                if not isinstance(formula_value, str):
                    formula_value = getattr(formula_value, 'text', None)
                if formula_value and EXTERNAL_REF_RE.search(formula_value):
                    try:
                        wb.sheet(sheet_name).set_value(cell_f.row, cell_f.column, 0)
                        external_ref_cells.add(f"{sheet_name}!{cell_f.coordinate}")
                    except Exception:
                        pass  # if neutralizing itself fails, evaluate_all() below will surface it clearly
                    continue  # never compare a neutralized cell — its value is fabricated
                key = f"{sheet_name}!{cell_f.coordinate}"
                targets.append((sheet_name, cell_f.row, cell_f.column))
                target_keys.append(key)
                cached_values.append(cell_c.value)
                formula_texts.append(formula_value)
                sheet_of_target[key] = sheet_name

    # Taint propagation: find every cell that references a neutralized
    # cell, directly or transitively, and exclude it from the comparison
    # too. Fast string-contains pre-filter (cheap) before the more
    # expensive regex-based reference parsing (only run on formulas that
    # plausibly could reference a tainted cell) — necessary for this to
    # stay tractable at real scale (1M+ formula cells). Bounded to a
    # reasonable number of BFS waves rather than iterating to an
    # unbounded fixed point, since each wave requires another pass over
    # every remaining formula.
    _progress(f"Scan complete: {len(targets):,} target(s), {len(external_ref_cells)} external-reference cell(s) found ({round(time.time()-t_start,1)}s elapsed).")

    # FIX #2 (found via a real Hidden Gem run: the masking BFS below hit
    # its old 30-wave cap with 7 cells STILL newly tainted on wave 30 —
    # i.e. genuinely not converged, silently truncated). The same fixed-
    # cap risk applies here too, even though this particular run's 4
    # external-reference cells happened to have zero downstream
    # dependents. Rather than leave one BFS capped and hope a future
    # file's external-ref chain never happens to be long, both passes
    # now use the same reverse-dependency-index architecture: parse
    # every formula's references once, build a reverse index once, then
    # propagate via the actual frontier with NO fixed wave limit — safe
    # to run to true convergence because each wave's cost is bounded by
    # the frontier size, not by rescanning every target cell.
    _progress(f"Parsing references for {len(target_keys):,} target cell(s) "
               f"for external-reference taint tracing...")
    _ext_refs_by_key = {}
    for i, formula in enumerate(formula_texts):
        if formula:
            _ext_refs_by_key[target_keys[i]] = _extract_refs(formula, sheet_of_target[target_keys[i]], sheet_order=sheet_order)
    _ext_dependents_of = {}
    for key, refs in _ext_refs_by_key.items():
        for (rsheet, rcol, rrow) in refs:
            _ext_dependents_of.setdefault(f"{rsheet}!{rcol}{rrow}", []).append(key)
    _progress(f"  reference parsing and reverse index complete ({round(time.time()-t_start,1)}s elapsed).")

    tainted = set(external_ref_cells)
    if tainted:
        _progress(f"Tracing dependents of {len(external_ref_cells)} external-reference cell(s)...")
        frontier = set(tainted)
        _wave = 0
        while True:
            _wave += 1
            newly_tainted = set()
            for tkey in frontier:
                for dep_key in _ext_dependents_of.get(tkey, ()):
                    if dep_key not in tainted:
                        newly_tainted.add(dep_key)
            if not newly_tainted:
                break
            tainted |= newly_tainted
            frontier = newly_tainted
            _progress(f"  wave {_wave}: {len(newly_tainted):,} newly tainted, {len(tainted):,} total ({round(time.time()-t_start,1)}s elapsed).")

        if len(tainted) > len(external_ref_cells):
            # Filter targets/keys/cached_values/formula_texts to drop
            # every tainted cell before the batch comparison below.
            keep_indices = [i for i, key in enumerate(target_keys) if key not in tainted]
            targets = [targets[i] for i in keep_indices]
            target_keys = [target_keys[i] for i in keep_indices]
            cached_values = [cached_values[i] for i in keep_indices]
            formula_texts = [formula_texts[i] for i in keep_indices]

    tainted_downstream_count = len(tainted) - len(external_ref_cells)
    if tainted_downstream_count > 0:
        _progress(f"Taint trace complete: {tainted_downstream_count:,} downstream cell(s) excluded ({round(time.time()-t_start,1)}s elapsed).")

    _progress(f"Running evaluate_all() ({len(targets):,} target cells to compare)...")
    try:
        wb.evaluate_all()
    except Exception as e:
        return {"status": "load_or_eval_failed", "error": str(e), "elapsed_s": round(time.time() - t_start, 2),
                "note": f"{len(external_ref_cells)} external-reference cell(s) were pre-neutralized but evaluation still failed — a second, different blocking issue exists beyond external references."}
    _progress(f"evaluate_all() complete ({round(time.time()-t_start,1)}s elapsed).")

    telemetry = wb.last_cycle_telemetry()
    _progress(f"Circularity: {telemetry.iterated_sccs} genuine group(s), {telemetry.converged_sccs} converged, {telemetry.capped_sccs} unconverged.")

    # fix #3b — one batched call instead of len(targets) individual ones.
    _progress(f"Running batch evaluate_cells() for {len(targets):,} target(s) — this is usually fast (~0.008ms/cell observed in testing), but is a single call with no incremental progress once started. If this step runs far longer than {len(targets)*0.001:.0f}s-{len(targets)*0.01:.0f}s, something is genuinely different about this file's formulas, not just scale.")
    try:
        recalculated_values = wb.evaluate_cells(targets)
    except Exception as e:
        return {"status": "batch_evaluation_failed", "error": str(e), "elapsed_s": round(time.time() - t_start, 2),
                "target_count": len(targets)}
    _progress(f"Batch evaluation complete ({round(time.time()-t_start,1)}s elapsed). Comparing against cached values...")

    mismatches = []
    unresolved_errors = []

    for key, cached_val, recalced_val in zip(target_keys, cached_values, recalculated_values):
        sheet_name, cell_addr = key.split('!', 1)

        is_string_error = isinstance(recalced_val, str) and recalced_val.startswith('#')
        is_dict_error = isinstance(recalced_val, dict) and recalced_val.get('type') == 'Error'
        if is_string_error or is_dict_error:
            error_repr = recalced_val if is_string_error else f"{recalced_val.get('kind', 'Unknown')} error"
            unresolved_errors.append({"sheet": sheet_name, "cell": cell_addr, "recalculated_error": error_repr})
            continue

        cached_is_nan = isinstance(cached_val, float) and math.isnan(cached_val)
        recalced_is_nan = isinstance(recalced_val, float) and math.isnan(recalced_val)
        if cached_is_nan or recalced_is_nan:
            # IEEE-754 NaN comparisons are always False in Python (nan > x
            # is never true), so the normal mismatch check below would
            # silently treat this as "no mismatch" — but a formula
            # genuinely producing NaN is itself usually a real problem
            # (e.g. an invalid statistical operation), not something to
            # hide by falling through the tolerance check.
            unresolved_errors.append({"sheet": sheet_name, "cell": cell_addr,
                                       "cached_is_nan": cached_is_nan, "recalculated_is_nan": recalced_is_nan})
            continue

        if isinstance(cached_val, (int, float)) and isinstance(recalced_val, (int, float)):
            tolerance = max(absolute_tolerance, abs(cached_val) * relative_tolerance)
            if abs(cached_val - recalced_val) > tolerance:
                mismatches.append({
                    "sheet": sheet_name, "cell": cell_addr,
                    "cached": cached_val, "recalculated": recalced_val,
                    "diff": recalced_val - cached_val,
                })

    # Separate genuinely clean mismatches from a chain of IFERROR/IFNA-
    # masked upstream failures. Found via real testing against Hidden
    # Gem: a Dashboard cell wraps 'Unit Economics'!F12 in IFERROR(...,0);
    # that cell's own formula is ALSO wrapped in an error-masking
    # function referencing something further upstream, and so on — each
    # layer silently returns its fallback rather than propagating a
    # visible error, so the single-hop reference-based taint trace above
    # can't see through more than one layer. This runs entirely on data
    # already in memory (formula text, the mismatches just built) — no
    # need to re-run the expensive Formualizer evaluation again.
    # Separate genuinely clean mismatches from ones caused by the same
    # upstream root cause as a confirmed genuine failure. Originally
    # this only looked for IFERROR/IFNA-wrapped mismatches referencing
    # each other, but real testing against Hidden Gem found a case that
    # breaks that assumption: =AA12-AA13, plain arithmetic with no
    # error-masking function at all, inheriting the same 0.0 result as
    # its precedents simply by referencing them. The correct, more
    # general test is reference-based, not function-based: seed from
    # cells we KNOW are genuinely broken (external-reference cells,
    # and any cell Formualizer itself reports as an unresolved error),
    # then propagate to any OTHER mismatch that references a seed or an
    # already-tainted cell — regardless of what function, if any, that
    # mismatch's own formula uses. Runs entirely on data already in
    # memory (formula text, the mismatches/unresolved_errors just
    # built) — no need to re-run the expensive Formualizer evaluation.
    _IFERROR_IFNA_RE = re.compile(r'\bIFERROR\s*\(|\bIFNA\s*\(', re.IGNORECASE)
    mismatch_keys = {f"{m['sheet']}!{m['cell']}" for m in mismatches}

    formula_by_key = {key: formula_texts[i] for i, key in enumerate(target_keys)}
    root_seeds = set(external_ref_cells) | {f"{e['sheet']}!{e['cell']}" for e in unresolved_errors}

    # FIX #2 (found via a real 8+ hour run against Hidden Gem that had to
    # be interrupted — the earlier fix below was itself broken at scale).
    # That version used a "cheap" substring pre-filter — any(pat in
    # formula for pat in addr_patterns) — on the assumption addr_patterns
    # would stay small, as it does for the external-ref-only BFS above
    # (seeded from a handful of cells). But THIS masking BFS seeds from
    # root_seeds = external_ref_cells | unresolved_errors, and Hidden
    # Gem's unresolved_error_count was 264,055 — so addr_patterns started
    # near that size and only grew every wave. Measured directly: a
    # substring check against ~113K patterns costs ~200 microseconds PER
    # FORMULA; doing that for ~1.15M target cells across multiple waves
    # extrapolates to hours, matching what actually happened. The real
    # fix is architectural, not a smaller constant: parse every formula's
    # references ONCE (not once per wave), build a reverse index of
    # "which cells reference cell X" ONCE, then propagate by walking only
    # the actual frontier each wave — a proper BFS over the dependency
    # graph, bounded by the number of (cell, reference) edges overall,
    # rather than rescanning all target cells' formula text on every
    # wave regardless of whether they're anywhere near the taint.
    _progress(f"Parsing references for {len(target_keys):,} target cell(s) "
               f"(one-time cost, not repeated per wave)...")
    refs_by_key = {}
    for key in target_keys:
        formula = formula_by_key.get(key)
        if formula:
            refs_by_key[key] = _extract_refs(formula, key.split('!', 1)[0], sheet_order=sheet_order)
    _progress(f"  reference parsing complete ({round(time.time()-t_start,1)}s elapsed).")

    dependents_of = {}
    for key, refs in refs_by_key.items():
        for (rsheet, rcol, rrow) in refs:
            dependents_of.setdefault(f"{rsheet}!{rcol}{rrow}", []).append(key)
    _progress(f"  reverse dependency index built: {len(dependents_of):,} referenced cell(s) "
               f"({round(time.time()-t_start,1)}s elapsed).")

    tainted_by_error = set()
    for key in target_keys:
        formula = formula_by_key.get(key)
        if not formula:
            continue
        # Directly references a confirmed-broken cell, or uses an
        # explicit error-masking function as a supplementary signal
        # (still valid — just no longer the ONLY signal).
        if _IFERROR_IFNA_RE.search(formula):
            tainted_by_error.add(key)
            continue
        refs = refs_by_key.get(key, ())
        if any(f"{rsheet}!{rcol}{rrow}" in root_seeds for (rsheet, rcol, rrow) in refs):
            tainted_by_error.add(key)
    _progress(f"  initial seeding pass complete: {len(tainted_by_error):,} cell(s) tainted "
               f"({round(time.time()-t_start,1)}s elapsed).")

    # FIX #2 continued: same reasoning as the external-reference BFS
    # above — no fixed wave cap. Confirmed this was NOT just theoretical:
    # a real run against Hidden Gem hit the old 30-wave cap with 7 cells
    # still newly tainted on wave 30, meaning propagation was silently
    # truncated before converging. Safe to run uncapped now that each
    # wave only touches the actual frontier via the reverse index.
    frontier = set(tainted_by_error)
    _wave = 0
    while True:
        _wave += 1
        newly_tainted = set()
        for tkey in frontier:
            for dep_key in dependents_of.get(tkey, ()):
                if dep_key not in tainted_by_error:
                    newly_tainted.add(dep_key)
        _progress(f"  masking wave {_wave}: {len(newly_tainted):,} newly tainted, "
                  f"{len(tainted_by_error)+len(newly_tainted):,} total ({round(time.time()-t_start,1)}s elapsed).")
        if not newly_tainted:
            break
        tainted_by_error |= newly_tainted
        frontier = newly_tainted

    masked = tainted_by_error & mismatch_keys

    mismatches_clean = [m for m in mismatches if f"{m['sheet']}!{m['cell']}" not in masked]
    mismatches_likely_masked_upstream_error = [m for m in mismatches if f"{m['sheet']}!{m['cell']}" in masked]

    return {
        "status": "success",
        "elapsed_s": round(time.time() - t_start, 2),
        "formula_cells_checked": len(targets),
        "external_reference_cells_excluded": len(external_ref_cells),
        "tainted_downstream_cells_excluded": tainted_downstream_count,
        "genuine_circular_groups": telemetry.iterated_sccs,
        "converged_circular_groups": telemetry.converged_sccs,
        "unconverged_circular_groups": telemetry.capped_sccs,
        "mismatches": mismatches_clean,
        "mismatch_count": len(mismatches_clean),
        "mismatches_likely_masked_upstream_error": mismatches_likely_masked_upstream_error,
        "mismatches_likely_masked_upstream_error_count": len(mismatches_likely_masked_upstream_error),
        "unresolved_errors": unresolved_errors,
        "unresolved_error_count": len(unresolved_errors),
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "reason": "usage: recalc_check.py <path>"}))
        sys.exit(1)
    result = run(sys.argv[1])
    print(json.dumps(result))
