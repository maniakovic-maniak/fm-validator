"""
test_recalc_sanitize.py — tests for the sanitize-and-retry fix in
recalc_check.py, addressing a real Calamine (Formualizer's backend)
parsing failure on defined names containing "?".

Investigation summary (see recalc_check.py's own docstrings for the
full detail):
  1. A minimal, isolated reproduction (a single-name workbook) confirmed
     "?" is precisely the trigger character, not something incidental
     to a more complex file.
  2. Formualizer's load_path() backend is "currently fixed to calamine"
     per its own docstring — no swappable-backend escape hatch exists.
  3. Removing the "?"-containing name via openpyxl and re-saving lets
     Formualizer load the file successfully.
  4. Running the full recalculation against a real, previously-crashing
     file after sanitization produced zero genuine mismatches, and
     confirmed only a small fraction of formula cells (692 of 59,228,
     ~1.2%) reference the removed names at all.
"""
import sys
import os
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))
import openpyxl
import recalc_check

all_pass = True


def check(desc, passed):
    global all_pass
    print(f"{'PASS' if passed else 'FAIL'}: {desc}")
    if not passed:
        all_pass = False


# ── Case 1: the minimal reproduction — confirms "?" alone is the
# precise trigger, and that Formualizer genuinely cannot load such a
# file without intervention. ──
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1['A1'] = 5
wb1.defined_names['CF_Report?'] = openpyxl.workbook.defined_name.DefinedName(
    'CF_Report?', attr_text="Sheet!$A$1"
)
path1 = tempfile.mktemp(suffix='.xlsx')
wb1.save(path1)

import formualizer as fz
try:
    fz.Workbook.load_path(path1)
    check('minimal reproduction: Formualizer fails to load a "?"-named defined name directly (confirms the bug is real before testing the fix)', False)
except Exception as e:
    check('minimal reproduction: Formualizer fails to load a "?"-named defined name directly (confirms the bug is real before testing the fix)',
          'Invalid name' in str(e))

# ── Case 2: the sanitize function itself — confirms it finds and
# removes the problematic name, and produces a file Formualizer can load. ──
temp_path, removed, affected = recalc_check._sanitize_problematic_defined_names(path1)
check('sanitize function finds and removes the "?"-named defined name', removed == ['CF_Report?'])
check('sanitize function produces a temp file', temp_path is not None and os.path.exists(temp_path))

try:
    wb_check = fz.Workbook.load_path(temp_path)
    check('the sanitized temp file loads successfully via Formualizer', True)
except Exception as e:
    check(f'the sanitized temp file loads successfully via Formualizer (got: {e})', False)
os.remove(temp_path)
os.remove(path1)

# ── Case 3: a normal file with NO problematic names — must be a
# genuine no-op, not attempt any sanitization. ──
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3['A1'] = 5
wb3.defined_names['NormalName'] = openpyxl.workbook.defined_name.DefinedName(
    'NormalName', attr_text="Sheet!$A$1"
)
path3 = tempfile.mktemp(suffix='.xlsx')
wb3.save(path3)
temp_path3, removed3, affected3 = recalc_check._sanitize_problematic_defined_names(path3)
check('a normal file with no "?"-named ranges is correctly left alone (genuine no-op, no temp file created)',
      temp_path3 is None and removed3 == [] and affected3 == 0)
os.remove(path3)

# ── Case 4: multiple problematic names, plus a formula that genuinely
# references one of them — confirms the "affected formula cell count"
# reporting is accurate, and that a formula referencing a removed name
# doesn't crash the whole load (matching what was confirmed on the
# real file: it just becomes one unresolved cell, not a fatal error). ──
wb4 = openpyxl.Workbook()
ws4 = wb4.active
ws4['A1'] = 10
ws4['B1'] = '=IF(Toggle?,A1,0)'  # genuinely references a name that will be removed
ws4['C1'] = '=A1*2'  # unrelated, must be unaffected
wb4.defined_names['Toggle?'] = openpyxl.workbook.defined_name.DefinedName('Toggle?', attr_text="Sheet!$A$1")
wb4.defined_names['Other?'] = openpyxl.workbook.defined_name.DefinedName('Other?', attr_text="Sheet!$A$1")
path4 = tempfile.mktemp(suffix='.xlsx')
wb4.save(path4)
temp_path4, removed4, affected4 = recalc_check._sanitize_problematic_defined_names(path4)
check('multiple problematic names are all found and removed', sorted(removed4) == ['Other?', 'Toggle?'])
check('the formula genuinely referencing a removed name is correctly counted as affected', affected4 == 1)
try:
    wb_check4 = fz.Workbook.load_path(temp_path4)
    check('a file with a formula referencing a removed name still loads successfully (does not crash the whole workbook)', True)
except Exception as e:
    check(f'a file with a formula referencing a removed name still loads successfully (got: {e})', False)
os.remove(temp_path4)
os.remove(path4)

# ── Case 5: end-to-end — the actual run() function, called directly
# against the exact real file that originally crashed, with no manual
# pre-sanitization. This is the real, practical test: does the
# documented bug actually get fixed automatically end-to-end. ──
real_file = '/mnt/user-data/uploads/The_Bend_Precinct_Model_26_7_2026_Pre_audit.xlsm'
if os.path.exists(real_file):
    result = recalc_check.run(real_file)
    check('end-to-end: the real, originally-crashing file now completes successfully via run() with no manual intervention',
          result.get('status') == 'success')
    check('end-to-end: zero genuine mismatches — every formula that resolved matches Excel\'s own cached value exactly',
          result.get('mismatch_count') == 0)
    check('end-to-end: the result transparently reports what was sanitized',
          result.get('sanitized_defined_names_count', 0) > 0)
else:
    print('SKIPPED: end-to-end real-file test (file not present in this environment)')

print('')
print('ALL TESTS PASSED' if all_pass else 'SOME TESTS FAILED')
if not all_pass:
    sys.exit(1)
